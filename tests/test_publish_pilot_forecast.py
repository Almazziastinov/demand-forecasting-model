from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from scripts.publish_pilot_forecast import (
    MISSING_KRATNOST_LABEL,
    MISSING_STOCK_LABEL,
    PILOT_BAKERY_IDS,
    PRODUCT_NAME_OVERRIDES,
    _build_excel,
    _enrich_forecast_product_metadata,
    _find_bakeries_with_unavailable_stock,
    _round_up_kratnost,
    _production_plan_with_optional_kratnost,
)


def test_base_pilot_contains_ten_bakeries_without_kulagina() -> None:
    assert len(PILOT_BAKERY_IDS) == 10
    assert 16 not in PILOT_BAKERY_IDS


def test_round_up_kratnost_after_stock_subtraction() -> None:
    forecast = 23.8
    yesterday_stock = 8.0

    net_need = max(forecast - yesterday_stock, 0.0)

    assert net_need == 15.8
    assert _round_up_kratnost(net_need, 10) == 20


def test_stock_can_cover_full_forecast() -> None:
    assert _round_up_kratnost(max(5.0 - 7.0, 0.0), 10) == 0


def test_sales_without_recorded_production_marks_stock_unavailable() -> None:
    events = pd.DataFrame(
        [
            {"bakery_id": 229, "product_id": 1071, "qty_produced": 0, "qty_sold": 18},
            {"bakery_id": 229, "product_id": 36, "qty_produced": 0, "qty_sold": 5},
            {"bakery_id": 23, "product_id": 1071, "qty_produced": 30, "qty_sold": 20},
        ]
    )

    assert _find_bakeries_with_unavailable_stock(events) == {229}


def test_missing_kratnost_keeps_sku_with_explicit_label() -> None:
    production_plan, kratnost = _production_plan_with_optional_kratnost(15.8, None)

    assert production_plan == 16
    assert kratnost == "нет данных по кратности"
    assert kratnost == MISSING_KRATNOST_LABEL


def test_known_kratnost_still_rounds_to_batch() -> None:
    production_plan, kratnost = _production_plan_with_optional_kratnost(15.8, 10)

    assert production_plan == 20
    assert kratnost == 10


def test_temporary_pletenka_name_overrides() -> None:
    assert PRODUCT_NAME_OVERRIDES == {
        11615: "Плетенка кленовая",
        11616: "Плетенка с черникой",
        11617: "Плетенка с земляникой",
    }


def test_enrich_forecast_product_metadata_fills_only_missing_values() -> None:
    forecast = pd.DataFrame(
        [
            {
                "product_id": 11615,
                "product_name": None,
                "category_name": None,
                "forecast_qty": 2.4,
            },
            {
                "product_id": 11616,
                "product_name": "Плетенка с черникой",
                "category_name": "Выпечка сладкая",
                "forecast_qty": 2.6,
            },
        ]
    )
    dimension = pd.DataFrame(
        [
            {
                "product_id": 11615,
                "product_name": "Сдоба Кленовый пекан",
                "category_name": "Выпечка сладкая",
            },
            {
                "product_id": 11616,
                "product_name": "Сдоба с черникой",
                "category_name": "Выпечка сладкая",
            },
        ]
    )

    result = _enrich_forecast_product_metadata(forecast, dimension)

    assert result.loc[0, "product_name"] == "Сдоба Кленовый пекан"
    assert result.loc[0, "category_name"] == "Выпечка сладкая"
    assert result.loc[1, "product_name"] == "Плетенка с черникой"


def test_excel_contains_stock_and_production_plan_columns() -> None:
    rows = [
        {
            "bakery_id": 16,
            "bakery_name": "Кулагина 4 Казань",
            "category": "Выпечка сладкая",
            "product_name": "Ватрушка",
            "forecast": 23.8,
            "yesterday_stock": 8.0,
            "net_need": 15.8,
            "production_plan": 20,
            "total_for_sale": 28.0,
            "kratnost": 10,
        }
    ]

    workbook = load_workbook(BytesIO(_build_excel(rows, "2026-07-29")), data_only=True)
    sheet = workbook["Прогноз"]

    assert [cell.value for cell in sheet[2]] == [
        "Пекарня",
        "Категория",
        "Номенклатура",
        "Прогноз",
        "Остаток со вчерашнего дня",
        "Чистая потребность",
        "План выпуска",
        "Итого на продажу",
        "Кратность",
    ]
    assert [cell.value for cell in sheet[3]] == [
        "Кулагина 4 Казань",
        "Выпечка сладкая",
        "Ватрушка",
        23.8,
        8.0,
        15.8,
        20,
        28.0,
        10,
    ]


def test_excel_renders_missing_kratnost_as_text() -> None:
    rows = [
        {
            "bakery_id": 16,
            "bakery_name": "Кулагина 4 Казань",
            "category": "Выпечка сладкая",
            "product_name": "Новая позиция",
            "forecast": 15.8,
            "yesterday_stock": 0.0,
            "net_need": 15.8,
            "production_plan": 16,
            "total_for_sale": 16.0,
            "kratnost": MISSING_KRATNOST_LABEL,
        }
    ]

    workbook = load_workbook(BytesIO(_build_excel(rows, "2026-09-01")), data_only=True)
    sheet = workbook["Прогноз"]

    assert sheet.cell(row=3, column=7).value == 16
    assert sheet.cell(row=3, column=9).value == "нет данных по кратности"


def test_excel_renders_unavailable_stock_as_text() -> None:
    rows = [
        {
            "bakery_id": 229,
            "bakery_name": "Лукина 5 Чебоксары",
            "category": "Выпечка сытная",
            "product_name": "Треугольник курица безд",
            "forecast": 12.0,
            "yesterday_stock": MISSING_STOCK_LABEL,
            "net_need": 12.0,
            "production_plan": 20,
            "total_for_sale": 20.0,
            "kratnost": 20,
        }
    ]

    workbook = load_workbook(BytesIO(_build_excel(rows, "2026-09-02")), data_only=True)

    assert workbook["Прогноз"].cell(row=3, column=5).value == "нет данных по остатку"
