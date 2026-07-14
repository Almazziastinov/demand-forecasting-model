"""Tests for the template-driven baking plan (window assignment read from
the xlsx template, not computed — see docs/baking_plan_implementation.md).
"""

from __future__ import annotations

# ruff: noqa: E501
import sys
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "apps"))

from baking_plan import allocation, rendering, templates  # noqa: E402
from baking_plan.allocation import ScheduledColumn, allocate_template_row  # noqa: E402
from baking_plan.templates import Window  # noqa: E402

WINDOWS = {
    3: Window(column=3, label="4:00-7:00", start_hour=4, end_hour=7),
    4: Window(column=4, label="7:00-8:00", start_hour=7, end_hour=8),
    5: Window(column=5, label="8:00-9:00", start_hour=8, end_hour=9),
    8: Window(column=8, label="11:00-12:00", start_hour=11, end_hour=12),
    9: Window(column=9, label="12:00-13:00", start_hour=12, end_hour=13),
    12: Window(column=12, label="15:00-16:00", start_hour=15, end_hour=16),
}


def _schedule(*columns, defrost=()):
    return [
        ScheduledColumn(
            window=WINDOWS[c],
            column=c,
            is_defrost=c in defrost,
            note="20 (ночная дефр)" if c in defrost else None,
        )
        for c in columns
    ]


def _schedule_with_quantities(items, defrost=()):
    return [
        ScheduledColumn(
            window=WINDOWS[c],
            column=c,
            is_defrost=c in defrost,
            note="20 (ночная дефр)" if c in defrost else None,
            quantity=q,
        )
        for c, q in items
    ]


def test_coverage_hours_tiles_multi_window_without_gap_or_overlap():
    windows = [WINDOWS[3], WINDOWS[8], WINDOWS[12]]
    cov = allocation.coverage_hours(windows)
    assert cov["4:00-7:00"] == [6, 7, 8, 9, 10, 11]
    assert cov["11:00-12:00"][0] == 12
    all_hours = [h for hours in cov.values() for h in hours]
    assert len(all_hours) == len(set(all_hours))  # no overlap
    assert set(all_hours) == set(range(6, 24))  # no gap


def test_coverage_hours_single_late_window_absorbs_full_day():
    assert allocation.coverage_hours([WINDOWS[12]])["15:00-16:00"] == list(range(6, 24))


def test_allocate_only_fills_scheduled_columns():
    hourly = {6: 3, 7: 12, 8: 15, 9: 15, 10: 20, 11: 30, 12: 27}
    result = allocate_template_row(schedule=_schedule(3, 8), hourly=hourly)
    assert set(result.keys()) == {3, 8}
    assert result[3] == 3 + 12 + 15 + 15 + 20 + 30  # hours 6..11
    assert result[8] == 27  # hour 12..end


def test_schedule_round_to_uses_template_batch_gcd():
    assert allocation.schedule_round_to(_schedule_with_quantities([(3, 20), (8, 20)])) == 20
    assert allocation.schedule_round_to(_schedule_with_quantities([(3, 20), (8, 10), (12, 20)])) == 10


def test_allocate_rounds_up_and_carries_surplus_to_next_bakes():
    hourly = {6: 21, 12: 25, 16: 1}
    result = allocate_template_row(
        schedule=_schedule_with_quantities([(3, 20), (8, 20), (12, 20)]),
        hourly=hourly,
    )
    assert result == {3: 40, 8: 20}


def test_defrost_cell_detection_keys_off_cell_value_not_sku_name():
    assert allocation.is_defrost_cell("20 (ночная дефр)") is True
    assert allocation.is_defrost_cell(20) is False
    assert allocation.normalize_sku_name("Треугольник курица (тесто ночного брожжения)") == "треугольник курица"


def test_defrost_sized_from_next_day_early_window_with_annotation():
    today = {7: 50}
    next_day = {7: 30, 8: 5, 15: 99}  # hour 15 is late, excluded

    result = allocate_template_row(
        schedule=_schedule(3, 12, defrost=(12,)),
        hourly=today,
        next_day_hourly=next_day,
    )
    assert result[12] == "35 (ночная дефр)"
    assert result[3] == 50  # defrost column never contributes to today's bake


def test_defrost_falls_back_to_today_when_next_day_absent():
    result = allocate_template_row(
        schedule=_schedule(3, 12, defrost=(12,)),
        hourly={7: 40},
        next_day_hourly=None,
    )
    assert result[12] == "40 (ночная дефр)"


def test_allocate_returns_empty_when_no_hourly_forecast():
    result = allocate_template_row(schedule=_schedule(3, 9), hourly={})
    assert result == {}


def test_revenue_bucket_matches_template_thresholds():
    assert templates.revenue_bucket(1_499_999) == "до 1,5 млн"
    assert templates.revenue_bucket(1_500_000) == "до 2,5 млн"
    assert templates.revenue_bucket(2_500_000) == "от 2,5 млн"
    assert templates.revenue_bucket(3_000_000) == "от 3млн"
    assert templates.revenue_bucket(None) == templates.DEFAULT_BUCKET


def test_select_sheet_name_falls_back_when_bucket_unknown():
    assert templates.select_sheet_name("до 2,5 млн", ["до 1,5 млн", "до 2,5 млн"]) == "до 2,5 млн"
    assert templates.select_sheet_name("нет такого", ["до 1,5 млн", "до 2,5 млн"]) == "до 2,5 млн"
    assert templates.select_sheet_name(None, ["только один"]) == "только один"


def test_sku_match_keys_include_known_forecast_aliases():
    assert "треугольник курица" in allocation.sku_match_keys("Треугольник курица безд")
    assert "жар пицца оригинальная" in allocation.sku_match_keys("ЖарПицца Оригинальная")
    assert "пирожок булочка с яблоками" in allocation.sku_match_keys("Пирожок яблоко")


def test_assortment_lookup_prefers_regular_product_over_order_variant():
    lookup = allocation.build_assortment_lookup(
        [
            {"product_id": "1", "product_name": "Клубника и банан ЗКЗ", "category_name": "Заказная продукция"},
            {"product_id": "2", "product_name": "Клубника и банан НОВЫЙ", "category_name": "Пироги сладкие"},
        ]
    )
    assert lookup["клубника банан"]["product_id"] == "2"


def test_template_path_for_bakery_uses_individual_override():
    assert templates.template_path_for_bakery(22).name == "22_sibirskiy_trakt_25.xlsx"
    assert templates.template_path_for_bakery(999999).name == "template.xlsx"


def test_base_template_has_revenue_tier_sheets_and_comments():
    workbook = load_workbook(templates.BASE_TEMPLATE_PATH, data_only=True)
    assert "комментарии" in workbook.sheetnames
    for bucket in ("до 1,5 млн", "до 2,5 млн", "от 2,5 млн", "от 3млн"):
        assert bucket in workbook.sheetnames
    windows = templates.parse_windows(workbook, "до 2,5 млн")
    assert len(windows) > 0
    assert all(w.column >= 3 for w in windows)
    meta = templates.parse_comments_sheet(workbook)
    assert len(meta) > 0


def test_write_plan_renders_category_header_total_column_and_unscheduled_rows():
    workbook = load_workbook(templates.BASE_TEMPLATE_PATH)
    selected = templates.select_sheet_name("до 1,5 млн", workbook.sheetnames)
    windows = templates.parse_windows(workbook, selected)
    sheet = workbook[selected]
    for worksheet in list(workbook.worksheets):
        if worksheet.title != selected:
            workbook.remove(worksheet)
    sheet.title = rendering.SHEET_TITLE

    total_column = 3 + len(windows)
    plan_rows = [
        {
            "snapshot": rendering.snapshot_row(sheet, templates.PLAN_START_ROW, total_column),
            "product_id": "1",
            "product_name": "Тестовый SKU",
            "category_name": "Выпечка сытная",
            "allocated": {windows[0].column: 20},
            "total": 20,
            "source_order": templates.PLAN_START_ROW,
        },
        {
            "snapshot": None,
            "product_id": "2",
            "product_name": "Новая позиция",
            "category_name": "Новая группа",
            "allocated": {},
            "total": 7,
            "source_order": 10**9,
        },
    ]

    rendering.write_plan(
        sheet=sheet,
        windows=windows,
        plan_rows=plan_rows,
        bakery_name="Тест",
        forecast_date="2026-06-20",
        selected_sheet_name=selected,
    )

    output = BytesIO()
    workbook.save(output)
    result_sheet = load_workbook(output)[rendering.SHEET_TITLE]

    names = {
        result_sheet.cell(row=row, column=2).value: row
        for row in range(1, result_sheet.max_row + 1)
        if result_sheet.cell(row=row, column=2).value
    }
    assert result_sheet.cell(row=templates.WINDOWS_HEADER_ROW, column=total_column).value == "Итого"
    assert "Тестовый SKU" in names
    assert "Новая позиция" in names

    new_row = names["Новая позиция"]
    assert all(result_sheet.cell(row=new_row, column=w.column).value is None for w in windows)
    assert result_sheet.cell(row=new_row, column=total_column).value == 7
    assert all(not result_sheet.row_dimensions[row].hidden for row in range(1, result_sheet.max_row + 1))
