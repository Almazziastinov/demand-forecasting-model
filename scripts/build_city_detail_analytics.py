"""Build analytics workbook with filled cities and detailed complaint categories."""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any
import json
import os
import re
import time

import openpyxl
import requests
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter


INPUT_FILE = Path(
    r"C:\Users\dns\Documents\Codex\2026-06-30\new-chat\outputs\Отзывы покупателей единый реестр.xlsx"
)
OUTPUT_FILE = Path(
    r"C:\Users\dns\Documents\Codex\2026-06-30\new-chat\outputs\Отзывы покупателей единый реестр аналитика города детали.xlsx"
)
PROJECT_DIR = Path(__file__).resolve().parents[1]
LEGACY_SCRIPT = PROJECT_DIR / "analyze_reviews.py"
CITY_API_CHECKPOINT = PROJECT_DIR / "outputs" / "remaining_city_api_checkpoint.json"
BASE_URL = "https://vibecode.bitrix24.tech/v1"
MODEL = "bitrix/bitrixgpt-5.5"

DETAIL_CATEGORIES = [
    "Низкое качество еды",
    "Невежливое обслуживание",
    "Несвежие продукты",
    "Проблемы с чистотой",
    "Недостаток ассортимента",
    "Ошибки при расчетах",
    "Наличие вредителей",
    "Плохая упаковка товаров",
    "Другое",
]


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_city(value: Any) -> str:
    city = clean_text(value)
    if not city or city == "Неизвестно":
        return "Неизвестно"
    city = city.replace("г. ", "").replace("г.", "").strip(" ,")
    if city.lower() == "набережные челны":
        return "Набережные Челны"
    return city


def read_rows(workbook: openpyxl.Workbook) -> list[dict[str, Any]]:
    sheet = workbook["Единый реестр отзывов"]
    headers = [clean_text(c.value) for c in sheet[1]]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, row))
        rows.append(item)
    return rows


def compact_address_key(value: str) -> str:
    value = clean_text(value).lower().replace("ё", "е")
    value = re.sub(r"\b(ул|улица|д|дом|проспект|пр-т|пр|корп|корпус)\b\.?", " ", value)
    value = re.sub(r"[^а-яa-z0-9/]+", " ", value)
    return " ".join(value.split())


def load_address_mapping(workbook: openpyxl.Workbook) -> dict[str, str]:
    if "Сцепка" not in workbook.sheetnames:
        return {}
    sheet = workbook["Сцепка"]
    mapping: dict[str, str] = {}
    for raw, normalized in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        if not normalized:
            continue
        normalized_value = clean_text(normalized)
        for value in [raw, normalized]:
            key = compact_address_key(clean_text(value))
            if key:
                mapping[key] = normalized_value
    return mapping


def normalize_address_with_mapping(address: str, mapping: dict[str, str]) -> str:
    address = clean_text(address)
    if not address or address == "Неизвестно":
        return address
    key = compact_address_key(address)
    if key in mapping:
        return mapping[key]
    # Conservative contains check: prefer longer keys to avoid accidental street-only matches.
    for map_key, normalized in sorted(mapping.items(), key=lambda item: len(item[0]), reverse=True):
        if len(map_key) < 8:
            continue
        if map_key in key or key in map_key:
            return normalized
    return address


def valid_bakery_address(address: str, city_patterns: list[str]) -> bool:
    address = clean_text(address)
    if not address or address == "Неизвестно":
        return False
    if normalize_city(address) != "Неизвестно" and len(address.split()) <= 2:
        return False
    if re.fullmatch(r"[А-Яа-яA-Za-z .]+\.?\s*\+?\d[\d\s()+-]{7,}", address):
        return False
    return bool(re.search(r"\d", address))


def known_city_patterns(rows: list[dict[str, Any]]) -> list[str]:
    counts = Counter()
    for row in rows:
        city = normalize_city(row.get("город"))
        if city != "Неизвестно":
            counts[city] += 1
    return sorted([city for city in counts if len(city) > 3], key=len, reverse=True)


def fill_city(row: dict[str, Any], patterns: list[str]) -> tuple[str, str]:
    current = normalize_city(row.get("город"))
    if current != "Неизвестно":
        return current, "исходный"

    extracted_address = extract_structured_address(clean_text(row.get("текст_отзыва")), patterns)
    extracted_city = city_from_text(extracted_address, patterns)
    if extracted_city != "Неизвестно":
        return extracted_city, "по адресу"

    haystack = f"{clean_text(row.get('адрес'))} {clean_text(row.get('текст_отзыва'))}".lower()
    for city in patterns:
        if city.lower() in haystack:
            return city, "по тексту"
    return "Неизвестно", "не найден"


def city_from_text(text: str, city_patterns: list[str]) -> str:
    low = clean_text(text).lower()
    for city in sorted(city_patterns, key=len, reverse=True):
        if city.lower() in low:
            return normalize_city(city)
    return "Неизвестно"


def looks_like_address(candidate: str, city_patterns: list[str]) -> bool:
    low = candidate.lower()
    has_city = any(city.lower() in low for city in city_patterns)
    has_digit = bool(re.search(r"\d", candidate))
    noisy_words = [
        "покупатель", "продавец", "сегодня", "очень", "вкусн", "сделайте", "деньги",
        "за что", "надоело", "здравствуйте", "спасибо", "обслуживание",
    ]
    too_long = len(candidate) > 90
    has_noise = any(word in low for word in noisy_words)
    return has_city and has_digit and not too_long and not has_noise


def normalize_extracted_address(candidate: str) -> str:
    candidate = re.sub(r"^(отзыв\s+)?(2гис|яндекс справочник|яндекс|горячая линия)\s*[:.-]?\s*", "", candidate, flags=re.I)
    candidate = re.sub(r"^\+?\d[\d\s()+-]{7,}\s*", "", candidate)
    candidate = re.sub(r"\s+", " ", candidate).strip(" .,-")
    candidate = re.sub(r"(\d)(Кофейня)", r"\1 \2", candidate)
    candidate = re.sub(r"(\d)(Казань|Москва|Чебоксары|Набережные)", r"\1 \2", candidate)
    return candidate


def extract_structured_address(text: str, city_patterns: list[str]) -> str:
    text = clean_text(text).replace("\n", " ")
    text = re.sub(r"\s+", " ", text)

    patterns = [
        # Отзыв 2ГИС. Фучика 90 Кофейня Казань. 03.04.2024. ...
        r"(?:Отзыв\s+)?(?:2ГИС|Яндекс Справочник|ВКонтакте|Вконтакте)[.,]\s*(?P<addr>.*?)(?:\.|\s+)(?=\d{1,2}[./]\d{1,2}[./]\d{2,4})",
        # Same source prefix, but date is missing: use only the short first segment.
        r"(?:Отзыв\s+)?(?:2ГИС|Яндекс Справочник|ВКонтакте|Вконтакте)[.,]\s*(?P<addr>[^.]{1,90})\.",
        # Жалоба горячая линия. 8967... Лево-Булачная 42/2 Казань. Покупатель ...
        r"Жалоба\s+горячая\s+линия\.\s*(?:\+?\d[\d\s()+-]{7,}\.?\s*)?(?P<addr>.*?)(?=\.?\s*Покупатель\b)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I)
        if not match:
            continue
        candidate = normalize_extracted_address(match.group("addr"))
        if looks_like_address(candidate, city_patterns):
            return candidate
    return ""


def fill_address(row: dict[str, Any], city_patterns: list[str]) -> tuple[str, str]:
    current = clean_text(row.get("адрес"))
    city = clean_text(row.get("город"))
    if current and current not in {"Неизвестно", city}:
        return current, "исходный"

    candidate = extract_structured_address(clean_text(row.get("текст_отзыва")), city_patterns)
    if candidate:
        return candidate, "по тексту"
    return current if current else "Неизвестно", "не найден" if not current else "исходный"


def detailed_category(row: dict[str, Any]) -> str:
    if clean_text(row.get("тип_отзыва")) != "Жалоба":
        return ""

    category = clean_text(row.get("категория_нормализованная"))
    text = clean_text(row.get("текст_отзыва")).lower().replace("ё", "е")

    if any(w in text for w in ["таракан", "насеком", "мыш", "крыс", "вредител"]):
        return "Наличие вредителей"
    if any(w in text for w in ["плес", "проср", "кисл", "прокис", "несвеж", "стар", "запах", "отрав"]):
        return "Несвежие продукты"
    if any(w in text for w in ["гряз", "мусор", "антисан", "санитар", "чистот", "стол", "пол"]):
        return "Проблемы с чистотой"
    if any(w in text for w in ["сдач", "чек", "расчет", "рассчет", "обсчит", "деньг", "ценник", "цена не", "неверно считают"]):
        return "Ошибки при расчетах"
    if any(w in text for w in ["упаков", "пакет", "развал", "помял", "порвал"]):
        return "Плохая упаковка товаров"
    if any(w in text for w in ["нет в наличии", "нет товара", "ассортимент", "законч", "выбор", "не было"]):
        return "Недостаток ассортимента"
    if category == "Сервис" or any(w in text for w in ["груб", "хам", "кассир", "продав", "персонал", "обслуж", "отношение", "очеред", "ждать"]):
        return "Невежливое обслуживание"
    if category in {"Качество продукции", "Просрочка"}:
        return "Низкое качество еды"
    if category == "Чистота":
        return "Проблемы с чистотой"
    if category == "Ассортимент":
        return "Недостаток ассортимента"
    return "Другое"


def enrich_rows(rows: list[dict[str, Any]], address_mapping: dict[str, str]) -> None:
    patterns = known_city_patterns(rows)
    for row in rows:
        address, address_source = fill_address(row, patterns)
        address = normalize_address_with_mapping(address, address_mapping)
        if not valid_bakery_address(address, patterns):
            address = "Неизвестно"
            address_source = "не найден"
        row["адрес"] = address
        row["адрес_источник"] = address_source
        city, source = fill_city(row, patterns)
        address_city = city_from_text(address, patterns)
        if address_city != "Неизвестно" and (source != "исходный" or city != address_city):
            city, source = address_city, "по адресу"
        row["город"] = city
        row["город_источник"] = source
        row["детальная_категория"] = detailed_category(row)


def load_api_key() -> str:
    env_key = os.getenv("VIBECODE_API_KEY")
    if env_key:
        return env_key
    text = LEGACY_SCRIPT.read_text(encoding="utf-8", errors="ignore")
    match = re.search(r'API_KEY\s*=\s*["\']([^"\']+)["\']', text)
    if not match:
        raise RuntimeError("API key was not found")
    return match.group(1)


def api_complete(prompt: str, api_key: str) -> str:
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    payload = {"model": MODEL, "messages": [{"role": "user", "content": prompt}], "max_tokens": 1600}
    attempt = 0
    while True:
        try:
            response = requests.post(f"{BASE_URL}/chat/completions", headers=headers, json=payload, timeout=120)
            if response.status_code == 429:
                wait = int(response.headers.get("Retry-After", 20))
                print(f"  rate limit, waiting {wait}s")
                time.sleep(wait)
                continue
            if response.status_code >= 500:
                wait = min(180, 15 + attempt * 15)
                print(f"  server error {response.status_code}, waiting {wait}s")
                time.sleep(wait)
                attempt += 1
                continue
            response.raise_for_status()
            return response.json()["choices"][0]["message"]["content"].strip()
        except requests.RequestException as exc:
            wait = min(180, 15 + attempt * 15)
            print(f"  request error, waiting {wait}s: {exc}")
            time.sleep(wait)
            attempt += 1


def strip_json(raw: str) -> str:
    text = raw.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    start = text.find("[")
    end = text.rfind("]")
    return text[start : end + 1] if start >= 0 and end > start else text


def fill_remaining_cities_with_api(rows: list[dict[str, Any]]) -> None:
    unresolved = [
        row for row in rows
        if row.get("город") == "Неизвестно" and row.get("город_источник") == "не найден"
    ]
    if not unresolved:
        return

    checkpoint = {}
    if CITY_API_CHECKPOINT.exists():
        checkpoint = json.loads(CITY_API_CHECKPOINT.read_text(encoding="utf-8"))

    api_key = load_api_key()
    todo = []
    for row in unresolved:
        key = f"{row.get('исходный_лист')}|{row.get('исходная_строка')}"
        if key not in checkpoint:
            todo.append((key, row))

    print(f"remaining city API: total={len(unresolved)} already={len(checkpoint)} todo={len(todo)}")
    for start in range(0, len(todo), 20):
        batch = todo[start : start + 20]
        lines = []
        for index, (key, row) in enumerate(batch, start=1):
            text = clean_text(row.get("текст_отзыва"))[:700].replace("\n", " ")
            address = clean_text(row.get("адрес"))
            lines.append(f"{index}. id={key}; адрес={address}; текст={text}")
        prompt = f"""Определи город по тексту отзыва и адресу.
Верни JSON-массив объектов в том же порядке.
Поля: "город".
Если город явно не указан, верни "Неизвестно".
Не выдумывай город по улице.

Строки:
{chr(10).join(lines)}
"""
        raw = api_complete(prompt, api_key)
        try:
            parsed = json.loads(strip_json(raw))
            if not isinstance(parsed, list):
                parsed = []
        except Exception:
            parsed = []
        for idx, (key, _row) in enumerate(batch):
            city = ""
            if idx < len(parsed) and isinstance(parsed[idx], dict):
                city = clean_text(parsed[idx].get("город"))
            checkpoint[key] = city if city else "Неизвестно"
        CITY_API_CHECKPOINT.parent.mkdir(parents=True, exist_ok=True)
        CITY_API_CHECKPOINT.write_text(json.dumps(checkpoint, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"  city api processed {min(start + len(batch), len(todo))}/{len(todo)}")

    for row in unresolved:
        key = f"{row.get('исходный_лист')}|{row.get('исходная_строка')}"
        city = clean_text(checkpoint.get(key))
        if city and city != "Неизвестно":
            row["город"] = normalize_city(city)
            row["город_источник"] = "API"


def style_header(sheet: openpyxl.worksheet.worksheet.Worksheet, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    for cell in sheet[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def border_table(sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    side = Side(style="thin", color="D9E2F3")
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def clear_sheets(workbook: openpyxl.Workbook) -> None:
    for name in ["Общая сводка", "Категории негатива", "По пекарням", "По городам", "Все отзывы", "Заполнение городов"]:
        if name in workbook.sheetnames:
            del workbook[name]


def rewrite_registry(workbook: openpyxl.Workbook, rows: list[dict[str, Any]]) -> None:
    old = workbook["Единый реестр отзывов"]
    index = workbook.worksheets.index(old)
    del workbook["Единый реестр отзывов"]
    sheet = workbook.create_sheet("Единый реестр отзывов", index)
    headers = [
        "дата", "год", "месяц", "тип_отзыва", "категория_исходная", "категория_нормализованная",
        "детальная_категория", "текст_отзыва", "город", "город_источник", "адрес", "адрес_источник", "РД",
        "управляющий", "источник", "формат_точки", "исходный_лист", "исходная_строка",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(h) for h in headers])
    style_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [14, 10, 10, 16, 24, 28, 28, 90, 22, 16, 34, 16, 18, 20, 22, 16, 22, 14]
    for i, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width
    for cell in sheet["A"]:
        if isinstance(cell.value, datetime):
            cell.number_format = "yyyy-mm-dd"


def write_summary(workbook: openpyxl.Workbook, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("Общая сводка")
    total = len(rows)
    type_counts = Counter(row.get("тип_отзыва") for row in rows)
    dates = [row.get("дата") for row in rows if isinstance(row.get("дата"), datetime)]
    sheet["A1"] = "Общая сводка отзывов"
    sheet["A1"].font = Font(bold=True, size=16)
    sheet.append([])
    sheet.append(["Тип отзыва", "Количество", "Доля", ""])
    for label in ["Благодарность", "Жалоба"]:
        count = type_counts[label]
        sheet.append([label, count, count / total if total else 0, ""])
    sheet.append(["Всего", total, 1, ""])
    sheet.append([])
    sheet.append(["Период", "", "", ""])
    sheet.append(["Минимальная дата", min(dates) if dates else "", "", ""])
    sheet.append(["Максимальная дата", max(dates) if dates else "", "", ""])
    style_header(sheet, 3)
    for r in range(4, 7):
        sheet.cell(r, 3).number_format = "0.0%"
    for r in range(9, 11):
        sheet.cell(r, 2).number_format = "yyyy-mm-dd"
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 14
    pie = PieChart()
    pie.title = "Распределение отзывов"
    pie.add_data(Reference(sheet, min_col=2, min_row=3, max_row=5), titles_from_data=True)
    pie.set_categories(Reference(sheet, min_col=1, min_row=4, max_row=5))
    sheet.add_chart(pie, "E3")


def write_negative_categories(workbook: openpyxl.Workbook, rows: list[dict[str, Any]]) -> list[str]:
    sheet = workbook.create_sheet("Категории негатива")
    complaints = [r for r in rows if r.get("тип_отзыва") == "Жалоба"]
    counts = Counter(r.get("детальная_категория") or "Другое" for r in complaints)
    total = len(complaints)
    sheet["A1"] = f"Категории проблем в жалобах (всего: {total})"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.append([])
    sheet.append(["Категория проблемы", "Количество", "Доля от жалоб"])
    for cat in DETAIL_CATEGORIES:
        count = counts[cat]
        if count:
            sheet.append([cat, count, count / total if total else 0])
    style_header(sheet, 3)
    for r in range(4, sheet.max_row + 1):
        sheet.cell(r, 3).number_format = "0.0%"
    sheet.column_dimensions["A"].width = 34
    chart = BarChart()
    chart.title = "Категории жалоб"
    chart.add_data(Reference(sheet, min_col=2, min_row=3, max_row=sheet.max_row), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=4, max_row=sheet.max_row))
    sheet.add_chart(chart, "E3")
    return [cat for cat in DETAIL_CATEGORIES if counts[cat]]


def write_pivot(workbook: openpyxl.Workbook, rows: list[dict[str, Any]], key: str, name: str, cats: list[str]) -> None:
    sheet = workbook.create_sheet(name)
    sheet.append([key, "Всего", "Благодарностей", "Жалоб", "% благодарностей", "% жалоб", *cats])
    grouped = defaultdict(list)
    for row in rows:
        grouped[clean_text(row.get(key)) or "Неизвестно"].append(row)
    out = []
    for value, group in grouped.items():
        total = len(group)
        thanks = sum(1 for r in group if r.get("тип_отзыва") == "Благодарность")
        complaints = total - thanks
        detail = Counter(r.get("детальная_категория") for r in group if r.get("тип_отзыва") == "Жалоба")
        out.append([value, total, thanks, complaints, thanks / total if total else 0, complaints / total if total else 0, *[detail[c] for c in cats]])
    for row in sorted(out, key=lambda x: x[1], reverse=True):
        sheet.append(row)
    style_header(sheet)
    border_table(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 42 if key == "адрес" else 24
    for c in range(2, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(c)].width = 16
    for r in range(2, sheet.max_row + 1):
        sheet.cell(r, 5).number_format = "0.0%"
        sheet.cell(r, 6).number_format = "0.0%"


def write_all_reviews(workbook: openpyxl.Workbook, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("Все отзывы")
    headers = ["Дата", "Год", "Источник", "Адрес", "Город", "Тип отзыва", "Категория", "Детальная категория", "Текст отзыва", "Исходный лист", "Исходная строка"]
    sheet.append(headers)
    for r in rows:
        sheet.append([r.get("дата"), r.get("год"), r.get("источник"), r.get("адрес"), r.get("город"), r.get("тип_отзыва"), r.get("категория_нормализованная"), r.get("детальная_категория"), r.get("текст_отзыва"), r.get("исходный_лист"), r.get("исходная_строка")])
    style_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [14, 10, 22, 34, 22, 18, 26, 28, 90, 22, 14]
    for i, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width


def write_city_fill(workbook: openpyxl.Workbook, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("Заполнение городов")
    counts = Counter(r.get("город_источник") for r in rows)
    address_counts = Counter(r.get("адрес_источник") for r in rows)
    sheet.append(["Источник города", "Количество"])
    for k, v in counts.most_common():
        sheet.append([k, v])
    sheet.append([])
    sheet.append(["Источник адреса", "Количество"])
    for k, v in address_counts.most_common():
        sheet.append([k, v])
    style_header(sheet)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 14


def main() -> None:
    wb = openpyxl.load_workbook(INPUT_FILE)
    rows = read_rows(wb)
    address_mapping = load_address_mapping(wb)
    enrich_rows(rows, address_mapping)
    fill_remaining_cities_with_api(rows)
    clear_sheets(wb)
    rewrite_registry(wb, rows)
    write_summary(wb, rows)
    cats = write_negative_categories(wb, rows)
    write_pivot(wb, rows, "адрес", "По пекарням", cats)
    write_pivot(wb, rows, "город", "По городам", cats)
    write_all_reviews(wb, rows)
    write_city_fill(wb, rows)
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    city_counts = Counter(r.get("город_источник") for r in rows)
    print(f"rows={len(rows)}")
    print(f"city_sources={dict(city_counts)}")
    print(f"output={OUTPUT_FILE}")


if __name__ == "__main__":
    main()
