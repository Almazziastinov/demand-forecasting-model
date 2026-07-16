"""Build analytics workbook with filled cities and detailed complaint categories."""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any
import json
import os
import re
import time

import openpyxl
import requests
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter


INPUT_FILE = Path(
    r"C:\Users\dns\Documents\Codex\2026-06-30\new-chat\outputs\РћС‚Р·С‹РІС‹ РїРѕРєСѓРїР°С‚РµР»РµР№ РµРґРёРЅС‹Р№ СЂРµРµСЃС‚СЂ.xlsx"
)
OUTPUT_FILE = Path(
    r"C:\Users\dns\Documents\Codex\2026-06-30\new-chat\outputs\РћС‚Р·С‹РІС‹ РїРѕРєСѓРїР°С‚РµР»РµР№ РµРґРёРЅС‹Р№ СЂРµРµСЃС‚СЂ Р°РЅР°Р»РёС‚РёРєР° РіРѕСЂРѕРґР° РґРµС‚Р°Р»Рё.xlsx"
)
PROJECT_DIR = Path(__file__).resolve().parents[1]
LEGACY_SCRIPT = PROJECT_DIR / "analyze_reviews.py"
CITY_API_CHECKPOINT = PROJECT_DIR / "outputs" / "remaining_city_api_checkpoint.json"
BASE_URL = "https://vibecode.bitrix24.tech/v1"
MODEL = "bitrix/bitrixgpt-5.5"

DETAIL_CATEGORIES = [
    "РќРёР·РєРѕРµ РєР°С‡РµСЃС‚РІРѕ РµРґС‹",
    "РќРµРІРµР¶Р»РёРІРѕРµ РѕР±СЃР»СѓР¶РёРІР°РЅРёРµ",
    "РќРµСЃРІРµР¶РёРµ РїСЂРѕРґСѓРєС‚С‹",
    "РџСЂРѕР±Р»РµРјС‹ СЃ С‡РёСЃС‚РѕС‚РѕР№",
    "РќРµРґРѕСЃС‚Р°С‚РѕРє Р°СЃСЃРѕСЂС‚РёРјРµРЅС‚Р°",
    "РћС€РёР±РєРё РїСЂРё СЂР°СЃС‡РµС‚Р°С…",
    "РќР°Р»РёС‡РёРµ РІСЂРµРґРёС‚РµР»РµР№",
    "РџР»РѕС…Р°СЏ СѓРїР°РєРѕРІРєР° С‚РѕРІР°СЂРѕРІ",
    "Р”СЂСѓРіРѕРµ",
]


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_city(value: Any) -> str:
    city = clean_text(value)
    if not city or city == "РќРµРёР·РІРµСЃС‚РЅРѕ":
        return "РќРµРёР·РІРµСЃС‚РЅРѕ"
    city = city.replace("Рі. ", "").replace("Рі.", "").strip(" ,")
    if city.lower() == "РЅР°Р±РµСЂРµР¶РЅС‹Рµ С‡РµР»РЅС‹":
        return "РќР°Р±РµСЂРµР¶РЅС‹Рµ Р§РµР»РЅС‹"
    return city


def read_rows(workbook: openpyxl.Workbook) -> list[dict[str, Any]]:
    sheet = workbook["Р•РґРёРЅС‹Р№ СЂРµРµСЃС‚СЂ РѕС‚Р·С‹РІРѕРІ"]
    headers = [clean_text(c.value) for c in sheet[1]]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, row))
        rows.append(item)
    return rows


def known_city_patterns(rows: list[dict[str, Any]]) -> list[str]:
    counts = Counter()
    for row in rows:
        city = normalize_city(row.get("РіРѕСЂРѕРґ"))
        if city != "РќРµРёР·РІРµСЃС‚РЅРѕ":
            counts[city] += 1
    return sorted([city for city in counts if len(city) > 3], key=len, reverse=True)


def fill_city(row: dict[str, Any], patterns: list[str]) -> tuple[str, str]:
    current = normalize_city(row.get("РіРѕСЂРѕРґ"))
    if current != "РќРµРёР·РІРµСЃС‚РЅРѕ":
        return current, "РёСЃС…РѕРґРЅС‹Р№"

    haystack = f"{clean_text(row.get('Р°РґСЂРµСЃ'))} {clean_text(row.get('С‚РµРєСЃС‚_РѕС‚Р·С‹РІР°'))}".lower()
    for city in patterns:
        if city.lower() in haystack:
            return city, "РїРѕ С‚РµРєСЃС‚Сѓ"
    return "РќРµРёР·РІРµСЃС‚РЅРѕ", "РЅРµ РЅР°Р№РґРµРЅ"


def split_sentences(text: str) -> list[str]:
    parts = re.split(r"\.\s*", text)
    return [part.strip(" .") for part in parts if part and part.strip(" .")]


def looks_like_address(candidate: str, city_patterns: list[str]) -> bool:
    low = candidate.lower()
    has_city = any(city.lower() in low for city in city_patterns)
    has_digit = bool(re.search(r"\d", candidate))
    bad_markers = ["РѕС‚Р·С‹РІ 2РіРёСЃ", "РѕС‚Р·С‹РІ СЏРЅРґРµРєСЃ", "Р¶Р°Р»РѕР±Р° РіРѕСЂСЏС‡Р°СЏ", "СЏРЅРґРµРєСЃ СЃРїСЂР°РІРѕС‡РЅРёРє", "2РіРёСЃ"]
    is_only_source = any(marker == low for marker in bad_markers)
    return has_city and has_digit and not is_only_source


def normalize_extracted_address(candidate: str) -> str:
    candidate = re.sub(r"^(РѕС‚Р·С‹РІ\s+)?(2РіРёСЃ|СЏРЅРґРµРєСЃ СЃРїСЂР°РІРѕС‡РЅРёРє|СЏРЅРґРµРєСЃ|РіРѕСЂСЏС‡Р°СЏ Р»РёРЅРёСЏ)\s*[:.-]?\s*", "", candidate, flags=re.I)
    candidate = re.sub(r"^\+?\d[\d\s()+-]{7,}\s*", "", candidate)
    candidate = re.sub(r"\s+", " ", candidate).strip(" .,-")
    return candidate


def fill_address(row: dict[str, Any], city_patterns: list[str]) -> tuple[str, str]:
    current = clean_text(row.get("Р°РґСЂРµСЃ"))
    city = clean_text(row.get("РіРѕСЂРѕРґ"))
    if current and current not in {"РќРµРёР·РІРµСЃС‚РЅРѕ", city}:
        return current, "РёСЃС…РѕРґРЅС‹Р№"

    text = clean_text(row.get("С‚РµРєСЃС‚_РѕС‚Р·С‹РІР°"))
    sentences = split_sentences(text[:500])
    for sentence in sentences[:5]:
        candidate = normalize_extracted_address(sentence)
        if looks_like_address(candidate, city_patterns):
            return candidate, "РїРѕ С‚РµРєСЃС‚Сѓ"

    # If the address is embedded before the first date, inspect that compact prefix.
    date_match = re.search(r"\b\d{1,2}[./]\d{1,2}[./]\d{2,4}\b", text)
    if date_match:
        prefix = text[: date_match.start()]
        for sentence in reversed(split_sentences(prefix)):
            candidate = normalize_extracted_address(sentence)
            if looks_like_address(candidate, city_patterns):
                return candidate, "РїРѕ С‚РµРєСЃС‚Сѓ"

    return current if current else "РќРµРёР·РІРµСЃС‚РЅРѕ", "РЅРµ РЅР°Р№РґРµРЅ" if not current else "РёСЃС…РѕРґРЅС‹Р№"


def detailed_category(row: dict[str, Any]) -> str:
    if clean_text(row.get("С‚РёРї_РѕС‚Р·С‹РІР°")) != "Р–Р°Р»РѕР±Р°":
        return ""

    category = clean_text(row.get("РєР°С‚РµРіРѕСЂРёСЏ_РЅРѕСЂРјР°Р»РёР·РѕРІР°РЅРЅР°СЏ"))
    text = clean_text(row.get("С‚РµРєСЃС‚_РѕС‚Р·С‹РІР°")).lower().replace("С‘", "Рµ")

    if any(w in text for w in ["С‚Р°СЂР°РєР°РЅ", "РЅР°СЃРµРєРѕРј", "РјС‹С€", "РєСЂС‹СЃ", "РІСЂРµРґРёС‚РµР»"]):
        return "РќР°Р»РёС‡РёРµ РІСЂРµРґРёС‚РµР»РµР№"
    if any(w in text for w in ["РїР»РµСЃ", "РїСЂРѕСЃСЂ", "РєРёСЃР»", "РїСЂРѕРєРёСЃ", "РЅРµСЃРІРµР¶", "СЃС‚Р°СЂ", "Р·Р°РїР°С…", "РѕС‚СЂР°РІ"]):
        return "РќРµСЃРІРµР¶РёРµ РїСЂРѕРґСѓРєС‚С‹"
    if any(w in text for w in ["РіСЂСЏР·", "РјСѓСЃРѕСЂ", "Р°РЅС‚РёСЃР°РЅ", "СЃР°РЅРёС‚Р°СЂ", "С‡РёСЃС‚РѕС‚", "СЃС‚РѕР»", "РїРѕР»"]):
        return "РџСЂРѕР±Р»РµРјС‹ СЃ С‡РёСЃС‚РѕС‚РѕР№"
    if any(w in text for w in ["СЃРґР°С‡", "С‡РµРє", "СЂР°СЃС‡РµС‚", "СЂР°СЃСЃС‡РµС‚", "РѕР±СЃС‡РёС‚", "РґРµРЅСЊРі", "С†РµРЅРЅРёРє", "С†РµРЅР° РЅРµ", "РЅРµРІРµСЂРЅРѕ СЃС‡РёС‚Р°СЋС‚"]):
        return "РћС€РёР±РєРё РїСЂРё СЂР°СЃС‡РµС‚Р°С…"
    if any(w in text for w in ["СѓРїР°РєРѕРІ", "РїР°РєРµС‚", "СЂР°Р·РІР°Р»", "РїРѕРјСЏР»", "РїРѕСЂРІР°Р»"]):
        return "РџР»РѕС…Р°СЏ СѓРїР°РєРѕРІРєР° С‚РѕРІР°СЂРѕРІ"
    if any(w in text for w in ["РЅРµС‚ РІ РЅР°Р»РёС‡РёРё", "РЅРµС‚ С‚РѕРІР°СЂР°", "Р°СЃСЃРѕСЂС‚РёРјРµРЅС‚", "Р·Р°РєРѕРЅС‡", "РІС‹Р±РѕСЂ", "РЅРµ Р±С‹Р»Рѕ"]):
        return "РќРµРґРѕСЃС‚Р°С‚РѕРє Р°СЃСЃРѕСЂС‚РёРјРµРЅС‚Р°"
    if category == "РЎРµСЂРІРёСЃ" or any(w in text for w in ["РіСЂСѓР±", "С…Р°Рј", "РєР°СЃСЃРёСЂ", "РїСЂРѕРґР°РІ", "РїРµСЂСЃРѕРЅР°Р»", "РѕР±СЃР»СѓР¶", "РѕС‚РЅРѕС€РµРЅРёРµ", "РѕС‡РµСЂРµРґ", "Р¶РґР°С‚СЊ"]):
        return "РќРµРІРµР¶Р»РёРІРѕРµ РѕР±СЃР»СѓР¶РёРІР°РЅРёРµ"
    if category in {"РљР°С‡РµСЃС‚РІРѕ РїСЂРѕРґСѓРєС†РёРё", "РџСЂРѕСЃСЂРѕС‡РєР°"}:
        return "РќРёР·РєРѕРµ РєР°С‡РµСЃС‚РІРѕ РµРґС‹"
    if category == "Р§РёСЃС‚РѕС‚Р°":
        return "РџСЂРѕР±Р»РµРјС‹ СЃ С‡РёСЃС‚РѕС‚РѕР№"
    if category == "РђСЃСЃРѕСЂС‚РёРјРµРЅС‚":
        return "РќРµРґРѕСЃС‚Р°С‚РѕРє Р°СЃСЃРѕСЂС‚РёРјРµРЅС‚Р°"
    return "Р”СЂСѓРіРѕРµ"


def enrich_rows(rows: list[dict[str, Any]]) -> None:
    patterns = known_city_patterns(rows)
    for row in rows:
        city, source = fill_city(row, patterns)
        row["РіРѕСЂРѕРґ"] = city
        row["РіРѕСЂРѕРґ_РёСЃС‚РѕС‡РЅРёРє"] = source
        address, address_source = fill_address(row, patterns)
        row["Р°РґСЂРµСЃ"] = address
        row["Р°РґСЂРµСЃ_РёСЃС‚РѕС‡РЅРёРє"] = address_source
        row["РґРµС‚Р°Р»СЊРЅР°СЏ_РєР°С‚РµРіРѕСЂРёСЏ"] = detailed_category(row)


def load_api_key() -> str:
    env_key = os.getenv("VIBECODE_API_KEY")
    if env_key:
        return env_key
    text = LEGACY_SCRIPT.read_text(encoding="utf-8", errors="ignore")
    match = re.search(r'API_KEY\s*=\s*["\']([^"\']+)["\']', text)
    if not match:
        raise RuntimeError("API key was not found")
    return match.group(1)


def api_complete(prompt: str, api_key: str) -> str:
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    payload = {"model": MODEL, "messages": [{"role": "user", "content": prompt}], "max_tokens": 1600}
    attempt = 0
    while True:
        try:
            response = requests.post(f"{BASE_URL}/chat/completions", headers=headers, json=payload, timeout=120)
            if response.status_code == 429:
                wait = int(response.headers.get("Retry-After", 20))
                print(f"  rate limit, waiting {wait}s")
                time.sleep(wait)
                continue
            if response.status_code >= 500:
                wait = min(180, 15 + attempt * 15)
                print(f"  server error {response.status_code}, waiting {wait}s")
                time.sleep(wait)
                attempt += 1
                continue
            response.raise_for_status()
            return response.json()["choices"][0]["message"]["content"].strip()
        except requests.RequestException as exc:
            wait = min(180, 15 + attempt * 15)
            print(f"  request error, waiting {wait}s: {exc}")
            time.sleep(wait)
            attempt += 1


def strip_json(raw: str) -> str:
    text = raw.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    start = text.find("[")
    end = text.rfind("]")
    return text[start : end + 1] if start >= 0 and end > start else text


def fill_remaining_cities_with_api(rows: list[dict[str, Any]]) -> None:
    unresolved = [
        row for row in rows
        if row.get("РіРѕСЂРѕРґ") == "РќРµРёР·РІРµСЃС‚РЅРѕ" and row.get("РіРѕСЂРѕРґ_РёСЃС‚РѕС‡РЅРёРє") == "РЅРµ РЅР°Р№РґРµРЅ"
    ]
    if not unresolved:
        return

    checkpoint = {}
    if CITY_API_CHECKPOINT.exists():
        checkpoint = json.loads(CITY_API_CHECKPOINT.read_text(encoding="utf-8"))

    api_key = load_api_key()
    todo = []
    for row in unresolved:
        key = f"{row.get('РёСЃС…РѕРґРЅС‹Р№_Р»РёСЃС‚')}|{row.get('РёСЃС…РѕРґРЅР°СЏ_СЃС‚СЂРѕРєР°')}"
        if key not in checkpoint:
            todo.append((key, row))

    print(f"remaining city API: total={len(unresolved)} already={len(checkpoint)} todo={len(todo)}")
    for start in range(0, len(todo), 20):
        batch = todo[start : start + 20]
        lines = []
        for index, (key, row) in enumerate(batch, start=1):
            text = clean_text(row.get("С‚РµРєСЃС‚_РѕС‚Р·С‹РІР°"))[:700].replace("\n", " ")
            address = clean_text(row.get("Р°РґСЂРµСЃ"))
            lines.append(f"{index}. id={key}; Р°РґСЂРµСЃ={address}; С‚РµРєСЃС‚={text}")
        prompt = f"""РћРїСЂРµРґРµР»Рё РіРѕСЂРѕРґ РїРѕ С‚РµРєСЃС‚Сѓ РѕС‚Р·С‹РІР° Рё Р°РґСЂРµСЃСѓ.
Р’РµСЂРЅРё JSON-РјР°СЃСЃРёРІ РѕР±СЉРµРєС‚РѕРІ РІ С‚РѕРј Р¶Рµ РїРѕСЂСЏРґРєРµ.
РџРѕР»СЏ: "РіРѕСЂРѕРґ".
Р•СЃР»Рё РіРѕСЂРѕРґ СЏРІРЅРѕ РЅРµ СѓРєР°Р·Р°РЅ, РІРµСЂРЅРё "РќРµРёР·РІРµСЃС‚РЅРѕ".
РќРµ РІС‹РґСѓРјС‹РІР°Р№ РіРѕСЂРѕРґ РїРѕ СѓР»РёС†Рµ.

РЎС‚СЂРѕРєРё:
{chr(10).join(lines)}
"""
        raw = api_complete(prompt, api_key)
        try:
            parsed = json.loads(strip_json(raw))
            if not isinstance(parsed, list):
                parsed = []
        except Exception:
            parsed = []
        for idx, (key, _row) in enumerate(batch):
            city = ""
            if idx < len(parsed) and isinstance(parsed[idx], dict):
                city = clean_text(parsed[idx].get("РіРѕСЂРѕРґ"))
            checkpoint[key] = city if city else "РќРµРёР·РІРµСЃС‚РЅРѕ"
        CITY_API_CHECKPOINT.parent.mkdir(parents=True, exist_ok=True)
        CITY_API_CHECKPOINT.write_text(json.dumps(checkpoint, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"  city api processed {min(start + len(batch), len(todo))}/{len(todo)}")

    for row in unresolved:
        key = f"{row.get('РёСЃС…РѕРґРЅС‹Р№_Р»РёСЃС‚')}|{row.get('РёСЃС…РѕРґРЅР°СЏ_СЃС‚СЂРѕРєР°')}"
        city = clean_text(checkpoint.get(key))
        if city and city != "РќРµРёР·РІРµСЃС‚РЅРѕ":
            row["РіРѕСЂРѕРґ"] = normalize_city(city)
            row["РіРѕСЂРѕРґ_РёСЃС‚РѕС‡РЅРёРє"] = "API"


def style_header(sheet: openpyxl.worksheet.worksheet.Worksheet, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    for cell in sheet[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def border_table(sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    side = Side(style="thin", color="D9E2F3")
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def clear_sheets(workbook: openpyxl.Workbook) -> None:
    for name in ["РћР±С‰Р°СЏ СЃРІРѕРґРєР°", "РљР°С‚РµРіРѕСЂРёРё РЅРµРіР°С‚РёРІР°", "РџРѕ РїРµРєР°СЂРЅСЏРј", "РџРѕ РіРѕСЂРѕРґР°Рј", "Р’СЃРµ РѕС‚Р·С‹РІС‹", "Р—Р°РїРѕР»РЅРµРЅРёРµ РіРѕСЂРѕРґРѕРІ"]:
        if name in workbook.sheetnames:
            del workbook[name]


def rewrite_registry(workbook: openpyxl.Workbook, rows: list[dict[str, Any]]) -> None:
    old = workbook["Р•РґРёРЅС‹Р№ СЂРµРµСЃС‚СЂ РѕС‚Р·С‹РІРѕРІ"]
    index = workbook.worksheets.index(old)
    del workbook["Р•РґРёРЅС‹Р№ СЂРµРµСЃС‚СЂ РѕС‚Р·С‹РІРѕРІ"]
    sheet = workbook.create_sheet("Р•РґРёРЅС‹Р№ СЂРµРµСЃС‚СЂ РѕС‚Р·С‹РІРѕРІ", index)
    headers = [
        "РґР°С‚Р°", "РіРѕРґ", "РјРµСЃСЏС†", "С‚РёРї_РѕС‚Р·С‹РІР°", "РєР°С‚РµРіРѕСЂРёСЏ_РёСЃС…РѕРґРЅР°СЏ", "РєР°С‚РµРіРѕСЂРёСЏ_РЅРѕСЂРјР°Р»РёР·РѕРІР°РЅРЅР°СЏ",
        "РґРµС‚Р°Р»СЊРЅР°СЏ_РєР°С‚РµРіРѕСЂРёСЏ", "С‚РµРєСЃС‚_РѕС‚Р·С‹РІР°", "РіРѕСЂРѕРґ", "РіРѕСЂРѕРґ_РёСЃС‚РѕС‡РЅРёРє", "Р°РґСЂРµСЃ", "Р°РґСЂРµСЃ_РёСЃС‚РѕС‡РЅРёРє", "Р Р”",
        "СѓРїСЂР°РІР»СЏСЋС‰РёР№", "РёСЃС‚РѕС‡РЅРёРє", "С„РѕСЂРјР°С‚_С‚РѕС‡РєРё", "РёСЃС…РѕРґРЅС‹Р№_Р»РёСЃС‚", "РёСЃС…РѕРґРЅР°СЏ_СЃС‚СЂРѕРєР°",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(h) for h in headers])
    style_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [14, 10, 10, 16, 24, 28, 28, 90, 22, 16, 34, 16, 18, 20, 22, 16, 22, 14]
    for i, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width
    for cell in sheet["A"]:
        if isinstance(cell.value, datetime):
            cell.number_format = "yyyy-mm-dd"


def write_summary(workbook: openpyxl.Workbook, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("РћР±С‰Р°СЏ СЃРІРѕРґРєР°")
    total = len(rows)
    type_counts = Counter(row.get("С‚РёРї_РѕС‚Р·С‹РІР°") for row in rows)
    dates = [row.get("РґР°С‚Р°") for row in rows if isinstance(row.get("РґР°С‚Р°"), datetime)]
    sheet["A1"] = "РћР±С‰Р°СЏ СЃРІРѕРґРєР° РѕС‚Р·С‹РІРѕРІ"
    sheet["A1"].font = Font(bold=True, size=16)
    sheet.append([])
    sheet.append(["РўРёРї РѕС‚Р·С‹РІР°", "РљРѕР»РёС‡РµСЃС‚РІРѕ", "Р”РѕР»СЏ", ""])
    for label in ["Р‘Р»Р°РіРѕРґР°СЂРЅРѕСЃС‚СЊ", "Р–Р°Р»РѕР±Р°"]:
        count = type_counts[label]
        sheet.append([label, count, count / total if total else 0, ""])
    sheet.append(["Р’СЃРµРіРѕ", total, 1, ""])
    sheet.append([])
    sheet.append(["РџРµСЂРёРѕРґ", "", "", ""])
    sheet.append(["РњРёРЅРёРјР°Р»СЊРЅР°СЏ РґР°С‚Р°", min(dates) if dates else "", "", ""])
    sheet.append(["РњР°РєСЃРёРјР°Р»СЊРЅР°СЏ РґР°С‚Р°", max(dates) if dates else "", "", ""])
    style_header(sheet, 3)
    for r in range(4, 7):
        sheet.cell(r, 3).number_format = "0.0%"
    for r in range(9, 11):
        sheet.cell(r, 2).number_format = "yyyy-mm-dd"
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 14
    pie = PieChart()
    pie.title = "Р Р°СЃРїСЂРµРґРµР»РµРЅРёРµ РѕС‚Р·С‹РІРѕРІ"
    pie.add_data(Reference(sheet, min_col=2, min_row=3, max_row=5), titles_from_data=True)
    pie.set_categories(Reference(sheet, min_col=1, min_row=4, max_row=5))
    sheet.add_chart(pie, "E3")


def write_negative_categories(workbook: openpyxl.Workbook, rows: list[dict[str, Any]]) -> list[str]:
    sheet = workbook.create_sheet("РљР°С‚РµРіРѕСЂРёРё РЅРµРіР°С‚РёРІР°")
    complaints = [r for r in rows if r.get("С‚РёРї_РѕС‚Р·С‹РІР°") == "Р–Р°Р»РѕР±Р°"]
    counts = Counter(r.get("РґРµС‚Р°Р»СЊРЅР°СЏ_РєР°С‚РµРіРѕСЂРёСЏ") or "Р”СЂСѓРіРѕРµ" for r in complaints)
    total = len(complaints)
    sheet["A1"] = f"РљР°С‚РµРіРѕСЂРёРё РїСЂРѕР±Р»РµРј РІ Р¶Р°Р»РѕР±Р°С… (РІСЃРµРіРѕ: {total})"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.append([])
    sheet.append(["РљР°С‚РµРіРѕСЂРёСЏ РїСЂРѕР±Р»РµРјС‹", "РљРѕР»РёС‡РµСЃС‚РІРѕ", "Р”РѕР»СЏ РѕС‚ Р¶Р°Р»РѕР±"])
    for cat in DETAIL_CATEGORIES:
        count = counts[cat]
        if count:
            sheet.append([cat, count, count / total if total else 0])
    style_header(sheet, 3)
    for r in range(4, sheet.max_row + 1):
        sheet.cell(r, 3).number_format = "0.0%"
    sheet.column_dimensions["A"].width = 34
    chart = BarChart()
    chart.title = "РљР°С‚РµРіРѕСЂРёРё Р¶Р°Р»РѕР±"
    chart.add_data(Reference(sheet, min_col=2, min_row=3, max_row=sheet.max_row), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=4, max_row=sheet.max_row))
    sheet.add_chart(chart, "E3")
    return [cat for cat in DETAIL_CATEGORIES if counts[cat]]


def write_pivot(workbook: openpyxl.Workbook, rows: list[dict[str, Any]], key: str, name: str, cats: list[str]) -> None:
    sheet = workbook.create_sheet(name)
    sheet.append([key, "Р’СЃРµРіРѕ", "Р‘Р»Р°РіРѕРґР°СЂРЅРѕСЃС‚РµР№", "Р–Р°Р»РѕР±", "% Р±Р»Р°РіРѕРґР°СЂРЅРѕСЃС‚РµР№", "% Р¶Р°Р»РѕР±", *cats])
    grouped = defaultdict(list)
    for row in rows:
        grouped[clean_text(row.get(key)) or "РќРµРёР·РІРµСЃС‚РЅРѕ"].append(row)
    out = []
    for value, group in grouped.items():
        total = len(group)
        thanks = sum(1 for r in group if r.get("С‚РёРї_РѕС‚Р·С‹РІР°") == "Р‘Р»Р°РіРѕРґР°СЂРЅРѕСЃС‚СЊ")
        complaints = total - thanks
        detail = Counter(r.get("РґРµС‚Р°Р»СЊРЅР°СЏ_РєР°С‚РµРіРѕСЂРёСЏ") for r in group if r.get("С‚РёРї_РѕС‚Р·С‹РІР°") == "Р–Р°Р»РѕР±Р°")
        out.append([value, total, thanks, complaints, thanks / total if total else 0, complaints / total if total else 0, *[detail[c] for c in cats]])
    for row in sorted(out, key=lambda x: x[1], reverse=True):
        sheet.append(row)
    style_header(sheet)
    border_table(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 42 if key == "Р°РґСЂРµСЃ" else 24
    for c in range(2, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(c)].width = 16
    for r in range(2, sheet.max_row + 1):
        sheet.cell(r, 5).number_format = "0.0%"
        sheet.cell(r, 6).number_format = "0.0%"


def write_all_reviews(workbook: openpyxl.Workbook, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("Р’СЃРµ РѕС‚Р·С‹РІС‹")
    headers = ["Р”Р°С‚Р°", "Р“РѕРґ", "РСЃС‚РѕС‡РЅРёРє", "РђРґСЂРµСЃ", "Р“РѕСЂРѕРґ", "РўРёРї РѕС‚Р·С‹РІР°", "РљР°С‚РµРіРѕСЂРёСЏ", "Р”РµС‚Р°Р»СЊРЅР°СЏ РєР°С‚РµРіРѕСЂРёСЏ", "РўРµРєСЃС‚ РѕС‚Р·С‹РІР°", "РСЃС…РѕРґРЅС‹Р№ Р»РёСЃС‚", "РСЃС…РѕРґРЅР°СЏ СЃС‚СЂРѕРєР°"]
    sheet.append(headers)
    for r in rows:
        sheet.append([r.get("РґР°С‚Р°"), r.get("РіРѕРґ"), r.get("РёСЃС‚РѕС‡РЅРёРє"), r.get("Р°РґСЂРµСЃ"), r.get("РіРѕСЂРѕРґ"), r.get("С‚РёРї_РѕС‚Р·С‹РІР°"), r.get("РєР°С‚РµРіРѕСЂРёСЏ_РЅРѕСЂРјР°Р»РёР·РѕРІР°РЅРЅР°СЏ"), r.get("РґРµС‚Р°Р»СЊРЅР°СЏ_РєР°С‚РµРіРѕСЂРёСЏ"), r.get("С‚РµРєСЃС‚_РѕС‚Р·С‹РІР°"), r.get("РёСЃС…РѕРґРЅС‹Р№_Р»РёСЃС‚"), r.get("РёСЃС…РѕРґРЅР°СЏ_СЃС‚СЂРѕРєР°")])
    style_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [14, 10, 22, 34, 22, 18, 26, 28, 90, 22, 14]
    for i, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width


def write_city_fill(workbook: openpyxl.Workbook, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("Р—Р°РїРѕР»РЅРµРЅРёРµ РіРѕСЂРѕРґРѕРІ")
    counts = Counter(r.get("РіРѕСЂРѕРґ_РёСЃС‚РѕС‡РЅРёРє") for r in rows)
    address_counts = Counter(r.get("Р°РґСЂРµСЃ_РёСЃС‚РѕС‡РЅРёРє") for r in rows)
    sheet.append(["РСЃС‚РѕС‡РЅРёРє РіРѕСЂРѕРґР°", "РљРѕР»РёС‡РµСЃС‚РІРѕ"])
    for k, v in counts.most_common():
        sheet.append([k, v])
    sheet.append([])
    sheet.append(["РСЃС‚РѕС‡РЅРёРє Р°РґСЂРµСЃР°", "РљРѕР»РёС‡РµСЃС‚РІРѕ"])
    for k, v in address_counts.most_common():
        sheet.append([k, v])
    style_header(sheet)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 14


def main() -> None:
    wb = openpyxl.load_workbook(INPUT_FILE)
    rows = read_rows(wb)
    enrich_rows(rows)
    fill_remaining_cities_with_api(rows)
    clear_sheets(wb)
    rewrite_registry(wb, rows)
    write_summary(wb, rows)
    cats = write_negative_categories(wb, rows)
    write_pivot(wb, rows, "Р°РґСЂРµСЃ", "РџРѕ РїРµРєР°СЂРЅСЏРј", cats)
    write_pivot(wb, rows, "РіРѕСЂРѕРґ", "РџРѕ РіРѕСЂРѕРґР°Рј", cats)
    write_all_reviews(wb, rows)
    write_city_fill(wb, rows)
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    city_counts = Counter(r.get("РіРѕСЂРѕРґ_РёСЃС‚РѕС‡РЅРёРє") for r in rows)
    print(f"rows={len(rows)}")
    print(f"city_sources={dict(city_counts)}")
    print(f"output={OUTPUT_FILE}")


if __name__ == "__main__":
    main()

