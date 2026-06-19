from __future__ import annotations

import argparse
import re
from difflib import SequenceMatcher
from pathlib import Path

import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_PRODUCTS = ROOT / "reports" / "baking_plan_templates" / "products.csv"
DEFAULT_DIM = ROOT / "reports" / "required_assortment" / "dim_products_lookup.csv"
DEFAULT_ASSORTMENT = (
    ROOT / "reports" / "required_assortment" / "assortment_city_products.csv"
)
DEFAULT_OUT = (
    ROOT / "reports" / "baking_plan_templates" / "template_assortment_audit.csv"
)


ALIASES: dict[str, str] = {
    "ватрушка в ассортменте": "ватрушка",
    "ватрушка в ассортименте": "ватрушка",
    "пирожок булочка с яблоками тесто ночное": "пирожок яблоко",
    "пирожок булочка с яблоками": "пирожок яблоко",
    "булочка с яблоками тесто ночное": "пирожок яблоко",
    "булочка с яблоками": "пирожок яблоко",
    "пирожок с яблоками тесто ночное": "пирожок яблоко",
    "пирожок с яблоками": "пирожок яблоко",
    "пирожок капуста курица": "пирожок капуста и курица",
    "пирог капуста курица": "капуста и курица",
    "капуста курица": "капуста и курица",
    "конвертик с курицей": "конвертик курица",
    "пирог капустный": "капустный",
    "пирог капуста мясо": "капуста и мясо",
    "пирог горбуша саго": "горбуша саго",
    "жар киш с курицей": "жар киш курица",
    "киш с курицей": "киш курица",
    "треугольник курица": "треугольник курица безд",
    "элеш": "элеш с курицей",
    "трехслойник": "трехслойник новый",
    "клубника банан": "клубника и банан новый",
    "кыстыбый": "кыстыбый п",
    "хуплу чебоксары": "хуплу",
    "пирог хуплу чебоксары": "пирог хуплу",
    "пирог картофель мсо чебоксары": "картофель и мясо",
}


STRIP_PHRASES = (
    "тесто ночного брожжения",
    "тесто ночного брожения",
    "тесто ночное",
    "ночная дефростация",
    "чебоксары",
)


def normalize(value: object, *, strip_phrases: bool = True) -> str:
    text = "" if pd.isna(value) else str(value)
    text = text.lower().replace("ё", "е")
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"[/\\.,;:!?\-+_]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    if strip_phrases:
        for phrase in STRIP_PHRASES:
            text = text.replace(phrase, " ")
        text = re.sub(r"\s+", " ", text).strip()
    return text


def infer_city(row: pd.Series) -> str:
    haystack = f"{row.get('folder', '')} {row.get('file_name', '')}".lower()
    if "курск" in haystack:
        return "Курск"
    if "иркутск" in haystack:
        return "Иркутск"
    kazan_markers = ("рт", "казань", "сибирский тракт", "мира 45", "парковая 7")
    if any(marker in haystack for marker in kazan_markers):
        return "Казань"
    return ""


def build_lookup(
    dim: pd.DataFrame,
) -> tuple[dict[str, dict[str, object]], list[dict[str, object]]]:
    dim = dim.copy()
    dim["product_key_audit"] = dim["product_name"].map(normalize)
    dim["active_rank"] = (
        dim["is_inactive_product"]
        .astype(str)
        .str.lower()
        .isin(["true", "1"])
        .astype(int)
    )
    dim = dim.sort_values(["product_key_audit", "active_rank", "product_id"])
    unique = dim.drop_duplicates(["product_key_audit", "product_id"]).to_dict("records")
    exact = {}
    for record in unique:
        exact.setdefault(record["product_key_audit"], record)
    return exact, unique


def match_product(
    name: str,
    exact: dict[str, dict[str, object]],
    products: list[dict[str, object]],
) -> dict[str, object]:
    key_raw = normalize(name, strip_phrases=False)
    key = normalize(name)
    alias_target = ALIASES.get(key_raw) or ALIASES.get(key)
    if alias_target:
        alias_key = normalize(alias_target)
        if alias_key in exact:
            result = dict(exact[alias_key])
            result["match_method"] = "alias_exact"
            result["match_score"] = 1.0
            return result

    if key in exact:
        result = dict(exact[key])
        result["match_method"] = "exact_norm"
        result["match_score"] = 1.0
        return result

    best: dict[str, object] | None = None
    best_score = 0.0
    for product in products:
        product_key = str(product["product_key_audit"])
        if not product_key:
            continue
        score = SequenceMatcher(None, key, product_key).ratio()
        key_words = set(key.split())
        product_words = set(product_key.split())
        if key_words and key_words.issubset(product_words):
            score = max(score, 0.92)
        if score > best_score:
            best = product
            best_score = score

    if best is not None and best_score >= 0.86:
        result = dict(best)
        result["match_method"] = "fuzzy_norm"
        result["match_score"] = round(best_score, 4)
        return result

    return {
        "product_id": "",
        "product_name": "",
        "category_name": "",
        "is_inactive_product": "",
        "match_method": "not_found",
        "match_score": 0.0,
    }


def load_active_pairs(path: Path) -> set[tuple[str, str]]:
    assortment = pd.read_csv(path, dtype=str).fillna("")
    active = assortment[assortment["is_active"].astype(str).isin(["1", "True", "true"])]
    return set(zip(active["city"], active["product_id"]))


def classify(
    row: pd.Series,
    active_pairs: set[tuple[str, str]],
    known_cities: set[str],
) -> str:
    if row["is_service_row"]:
        return "Служебная строка"
    if row["match_method"] == "not_found":
        return "Нет в dim_products"
    if str(row["is_inactive_product"]).lower() in {"true", "1"}:
        return "В dim_products выведено"
    if not row["inferred_city"]:
        return "Не определен город шаблона"
    if row["inferred_city"] not in known_cities:
        return "Город вне текущей таблицы ассортимента"
    if (row["inferred_city"], row["matched_product_id"]) not in active_pairs:
        return "Нет в актуальном ассортименте города"
    return "OK"


def is_service_row(source: pd.Series) -> bool:
    role = str(source.get("table_role", "")).strip()
    quantity = pd.to_numeric(source.get("qty_sum_in_template", 0), errors="coerce")
    return not role and (pd.isna(quantity) or float(quantity) == 0.0)


def build_audit(
    products_path: Path,
    dim_path: Path,
    assortment_path: Path,
) -> pd.DataFrame:
    products = pd.read_csv(products_path, dtype=str).fillna("")
    dim = pd.read_csv(dim_path, dtype=str).fillna("")
    exact, dim_products = build_lookup(dim)
    active_pairs = load_active_pairs(assortment_path)
    known_cities = {city for city, _ in active_pairs}

    records = []
    for _, source in products.iterrows():
        matched = match_product(source["product_name"], exact, dim_products)
        inferred_city = infer_city(source)
        record = {
            "file_name": source["file_name"],
            "folder": source["folder"],
            "sheet_name": source["sheet_name"],
            "row": source["row"],
            "table_role": source["table_role"],
            "template_product_name": source["product_name"],
            "qty_sum_in_template": source["qty_sum_in_template"],
            "inferred_city": inferred_city,
            "matched_product_id": matched["product_id"],
            "matched_product_name": matched["product_name"],
            "matched_category_name": matched["category_name"],
            "is_inactive_product": matched["is_inactive_product"],
            "match_method": matched["match_method"],
            "match_score": matched["match_score"],
            "is_service_row": is_service_row(source),
        }
        record["is_active_in_city_assortment"] = (
            bool(inferred_city)
            and bool(record["matched_product_id"])
            and (inferred_city, str(record["matched_product_id"])) in active_pairs
        )
        records.append(record)

    audit = pd.DataFrame(records)
    audit["problem"] = audit.apply(
        classify,
        axis=1,
        active_pairs=active_pairs,
        known_cities=known_cities,
    )
    audit["recommendation"] = audit["problem"].map(
        {
            "OK": "Оставить",
            "Служебная строка": "Оставить без изменения",
            "Нет в dim_products": "Нужно ручное соответствие или удаление из шаблона",
            "В dim_products выведено": (
                "Удалить из шаблона или заменить на актуальную номенклатуру"
            ),
            "Не определен город шаблона": "Уточнить город/назначение шаблона",
            "Город вне текущей таблицы ассортимента": (
                "Проверить отдельно после загрузки ассортимента этого города"
            ),
            "Нет в актуальном ассортименте города": (
                "Удалить из шаблона или добавить в актуальный ассортимент города"
            ),
        }
    )
    return audit


def write_xlsx(audit: pd.DataFrame, out_path: Path) -> None:
    summary = (
        audit.groupby(["inferred_city", "problem"], dropna=False)
        .size()
        .reset_index(name="rows")
        .sort_values(["inferred_city", "problem"])
    )
    by_product = (
        audit.groupby(
            [
                "problem",
                "template_product_name",
                "matched_product_id",
                "matched_product_name",
                "matched_category_name",
            ],
            dropna=False,
        )
        .agg(files=("file_name", "nunique"), rows=("file_name", "size"))
        .reset_index()
        .sort_values(
            ["problem", "files", "template_product_name"],
            ascending=[True, False, True],
        )
    )
    problem_rows = audit[audit["problem"] != "OK"].copy()

    xlsx_path = out_path.with_suffix(".xlsx")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="summary")
        by_product.to_excel(writer, index=False, sheet_name="by_product")
        problem_rows.to_excel(writer, index=False, sheet_name="problem_rows")
        audit.to_excel(writer, index=False, sheet_name="all_rows")

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    wb = load_workbook(xlsx_path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for column in ws.columns:
            letter = column[0].column_letter
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 48)
            ws.column_dimensions[letter].width = width
    wb.save(xlsx_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--products", type=Path, default=DEFAULT_PRODUCTS)
    parser.add_argument("--dim", type=Path, default=DEFAULT_DIM)
    parser.add_argument("--assortment", type=Path, default=DEFAULT_ASSORTMENT)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()

    audit = build_audit(args.products, args.dim, args.assortment)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    audit.to_csv(args.out, index=False, encoding="utf-8-sig")
    write_xlsx(audit, args.out)

    print(f"Rows: {len(audit)}")
    print(audit["problem"].value_counts().to_string())
    print(f"CSV: {args.out}")
    print(f"XLSX: {args.out.with_suffix('.xlsx')}")


if __name__ == "__main__":
    main()
