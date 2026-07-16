"""Build analytics sheets for the unified buyer reviews registry."""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT_FILE = Path(
    r"C:\Users\dns\Documents\Codex\2026-06-30\new-chat\outputs\Отзывы покупателей единый реестр.xlsx"
)
OUTPUT_FILE = Path(
    r"C:\Users\dns\Documents\Codex\2026-06-30\new-chat\outputs\Отзывы покупателей единый реестр аналитика.xlsx"
)

ANALYTIC_SHEETS = [
    "Общая сводка",
    "Категории негатива",
    "По пекарням",
    "По городам",
    "Все отзывы",
]


def normalize_blank(value: Any, fallback: str = "Неизвестно") -> str:
    text = "" if value is None else str(value).strip()
    return text if text and text.lower() != "none" else fallback


def read_registry(workbook: openpyxl.Workbook) -> list[dict[str, Any]]:
    sheet = workbook["Единый реестр отзывов"]
    headers = [str(cell.value).strip() for cell in sheet[1]]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, row))
        item["тип_отзыва"] = normalize_blank(item.get("тип_отзыва"), "Жалоба")
        item["категория_нормализованная"] = normalize_blank(
            item.get("категория_нормализованная"), "Без категории"
        )
        item["город"] = normalize_blank(item.get("город"))
        item["адрес"] = normalize_blank(item.get("адрес"))
        item["текст_отзыва"] = normalize_blank(item.get("текст_отзыва"), "")
        rows.append(item)
    return rows


def clear_analytics(workbook: openpyxl.Workbook) -> None:
    for sheet_name in ANALYTIC_SHEETS:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]


def style_header(sheet: openpyxl.worksheet.worksheet.Worksheet, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    for cell in sheet[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def set_table_style(sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_summary(workbook: openpyxl.Workbook, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("Общая сводка")
    total = len(rows)
    type_counts = Counter(row["тип_отзыва"] for row in rows)
    dates = [row["дата"] for row in rows if isinstance(row.get("дата"), datetime)]

    sheet["A1"] = "Общая сводка отзывов"
    sheet["A1"].font = Font(bold=True, size=16)
    sheet.append([])
    sheet.append(["Тип отзыва", "Количество", "Доля", ""])
    for label in ["Благодарность", "Жалоба"]:
        count = type_counts.get(label, 0)
        sheet.append([label, count, count / total if total else 0, ""])
    sheet.append(["Всего", total, 1, ""])
    sheet.append([])
    sheet.append(["Период", "", "", ""])
    sheet.append(["Минимальная дата", min(dates) if dates else "", "", ""])
    sheet.append(["Максимальная дата", max(dates) if dates else "", "", ""])

    style_header(sheet, 3)
    for row in range(4, 7):
        sheet.cell(row=row, column=3).number_format = "0.0%"
    for row in range(9, 11):
        sheet.cell(row=row, column=2).number_format = "yyyy-mm-dd"
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 14

    pie = PieChart()
    labels = Reference(sheet, min_col=1, min_row=4, max_row=5)
    data = Reference(sheet, min_col=2, min_row=3, max_row=5)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Распределение отзывов"
    sheet.add_chart(pie, "E3")


def write_negative_categories(workbook: openpyxl.Workbook, rows: list[dict[str, Any]]) -> list[str]:
    sheet = workbook.create_sheet("Категории негатива")
    complaints = [row for row in rows if row["тип_отзыва"] == "Жалоба"]
    counts = Counter(row["категория_нормализованная"] for row in complaints)
    total = len(complaints)

    sheet["A1"] = f"Категории проблем в жалобах (всего: {total})"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.append([])
    sheet.append(["Категория проблемы", "Количество", "Доля от жалоб"])
    top_categories = [category for category, _ in counts.most_common(8)]
    for category, count in counts.most_common():
        sheet.append([category, count, count / total if total else 0])

    style_header(sheet, 3)
    for row in range(4, sheet.max_row + 1):
        sheet.cell(row=row, column=3).number_format = "0.0%"
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 16

    if sheet.max_row >= 4:
        chart = BarChart()
        chart.title = "Топ категорий жалоб"
        chart.y_axis.title = "Количество"
        labels = Reference(sheet, min_col=1, min_row=4, max_row=min(sheet.max_row, 13))
        data = Reference(sheet, min_col=2, min_row=3, max_row=min(sheet.max_row, 13))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.height = 7
        chart.width = 14
        sheet.add_chart(chart, "E3")

    return top_categories


def build_group_rows(
    rows: list[dict[str, Any]], group_key: str, top_categories: list[str]
) -> list[list[Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[normalize_blank(row.get(group_key))].append(row)

    output = []
    for key, group in grouped.items():
        total = len(group)
        thanks = sum(1 for row in group if row["тип_отзыва"] == "Благодарность")
        complaints = sum(1 for row in group if row["тип_отзыва"] == "Жалоба")
        cat_counts = Counter(
            row["категория_нормализованная"]
            for row in group
            if row["тип_отзыва"] == "Жалоба"
        )
        output.append(
            [
                key,
                total,
                thanks,
                complaints,
                thanks / total if total else 0,
                complaints / total if total else 0,
                *[cat_counts.get(category, 0) for category in top_categories],
            ]
        )
    output.sort(key=lambda item: item[1], reverse=True)
    return output


def write_pivot(
    workbook: openpyxl.Workbook,
    rows: list[dict[str, Any]],
    group_key: str,
    sheet_name: str,
    top_categories: list[str],
) -> None:
    sheet = workbook.create_sheet(sheet_name)
    first_header = "адрес" if group_key == "адрес" else "город"
    headers = [
        first_header,
        "Всего",
        "Благодарностей",
        "Жалоб",
        "% благодарностей",
        "% жалоб",
        *top_categories,
    ]
    sheet.append(headers)
    for row in build_group_rows(rows, group_key, top_categories):
        sheet.append(row)

    style_header(sheet)
    set_table_style(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 42 if group_key == "адрес" else 24
    for col in range(2, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 16
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=5).number_format = "0.0%"
        sheet.cell(row=row, column=6).number_format = "0.0%"


def write_all_reviews(workbook: openpyxl.Workbook, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("Все отзывы")
    headers = [
        "Дата",
        "Год",
        "Источник",
        "Адрес",
        "Город",
        "Тип отзыва",
        "Категория",
        "Текст отзыва",
        "Исходный лист",
        "Исходная строка",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append(
            [
                row.get("дата"),
                row.get("год"),
                row.get("источник"),
                row.get("адрес"),
                row.get("город"),
                row.get("тип_отзыва"),
                row.get("категория_нормализованная"),
                row.get("текст_отзыва"),
                row.get("исходный_лист"),
                row.get("исходная_строка"),
            ]
        )

    style_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [14, 10, 22, 34, 22, 18, 26, 90, 22, 14]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for cell in sheet["A"]:
        if isinstance(cell.value, datetime):
            cell.number_format = "yyyy-mm-dd"
    for row in sheet.iter_rows(min_row=2):
        row[7].alignment = Alignment(wrap_text=True, vertical="top")


def main() -> None:
    workbook = openpyxl.load_workbook(INPUT_FILE)
    rows = read_registry(workbook)
    clear_analytics(workbook)
    write_summary(workbook, rows)
    top_categories = write_negative_categories(workbook, rows)
    write_pivot(workbook, rows, "адрес", "По пекарням", top_categories)
    write_pivot(workbook, rows, "город", "По городам", top_categories)
    write_all_reviews(workbook, rows)

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_FILE)
    print(f"rows={len(rows)}")
    print(f"output={OUTPUT_FILE}")


if __name__ == "__main__":
    main()
