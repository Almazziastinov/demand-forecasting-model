"""Apply available AI enrichment checkpoint to the unified reviews workbook."""

from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_FILE = Path(
    r"C:\Users\dns\Documents\Codex\2026-06-30\new-chat\outputs\Отзывы покупателей единый реестр.xlsx"
)
CHECKPOINT_FILE = Path(
    r"C:\Users\dns\Desktop\Projects\demand-forecasting-model\outputs\unified_reviews_ai_enrichment_checkpoint.json"
)
OUTPUT_FILE = Path(
    r"C:\Users\dns\Documents\Codex\2026-06-30\new-chat\outputs\Отзывы покупателей единый реестр AI частично.xlsx"
)

AI_COLUMNS = [
    "ai_обработано",
    "подкатегория_нормализованная",
    "подкатегория_описание",
    "критичность",
    "риск_здоровью",
    "продукт",
    "краткая_суть",
    "требует_реакции",
]


def load_checkpoint() -> dict[str, dict[str, str]]:
    return json.loads(CHECKPOINT_FILE.read_text(encoding="utf-8"))


def build_summary(workbook: openpyxl.Workbook, total_rows: int, enriched: dict[str, dict[str, str]]) -> None:
    sheet = workbook.create_sheet("AI сводка частичная")
    criticality = Counter(item.get("критичность") for item in enriched.values())
    health = Counter(item.get("риск_здоровью") for item in enriched.values())
    reaction = Counter(item.get("требует_реакции") for item in enriched.values())
    subcats = Counter(item.get("подкатегория_нормализованная") for item in enriched.values())

    rows = [
        ["Показатель", "Значение"],
        ["Всего отзывов в реестре", total_rows],
        ["AI обработано", len(enriched)],
        ["AI не обработано", total_rows - len(enriched)],
        ["", ""],
        ["Критичность", ""],
    ]
    rows.extend([key, value] for key, value in criticality.most_common())
    rows.append(["", ""])
    rows.append(["Риск здоровью", ""])
    rows.extend([key, value] for key, value in health.most_common())
    rows.append(["", ""])
    rows.append(["Требует реакции", ""])
    rows.extend([key, value] for key, value in reaction.most_common())
    rows.append(["", ""])
    rows.append(["Топ подкатегорий", ""])
    rows.extend([key, value] for key, value in subcats.most_common(30))

    for row in rows:
        sheet.append(row)

    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(bold=True, color="FFFFFF")
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 18


def main() -> None:
    enriched = load_checkpoint()
    workbook = openpyxl.load_workbook(INPUT_FILE)
    sheet = workbook["Единый реестр отзывов"]

    headers = [cell.value for cell in sheet[1]]
    source_sheet_col = headers.index("исходный_лист") + 1
    source_row_col = headers.index("исходная_строка") + 1
    start_col = sheet.max_column + 1

    for offset, header in enumerate(AI_COLUMNS):
        cell = sheet.cell(row=1, column=start_col + offset)
        cell.value = header
        cell.fill = PatternFill("solid", fgColor="7030A0")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    processed = 0
    for row_idx in range(2, sheet.max_row + 1):
        source_sheet = str(sheet.cell(row=row_idx, column=source_sheet_col).value or "").strip()
        source_row = str(sheet.cell(row=row_idx, column=source_row_col).value or "").strip()
        key = f"{source_sheet}|{source_row}"
        item = enriched.get(key)
        values = ["нет", "", "", "", "", "", "", ""]
        if item:
            processed += 1
            values = [
                "да",
                item.get("подкатегория_нормализованная", ""),
                item.get("подкатегория_описание", ""),
                item.get("критичность", ""),
                item.get("риск_здоровью", ""),
                item.get("продукт", ""),
                item.get("краткая_суть", ""),
                item.get("требует_реакции", ""),
            ]
        for offset, value in enumerate(values):
            sheet.cell(row=row_idx, column=start_col + offset).value = value

    widths = [14, 28, 30, 14, 14, 24, 60, 16]
    for offset, width in enumerate(widths):
        letter = get_column_letter(start_col + offset)
        sheet.column_dimensions[letter].width = width

    for row in sheet.iter_rows(min_row=2, min_col=start_col, max_col=start_col + len(AI_COLUMNS) - 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    build_summary(workbook, sheet.max_row - 1, enriched)
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_FILE)
    print(f"processed={processed}")
    print(f"output={OUTPUT_FILE}")


if __name__ == "__main__":
    main()
