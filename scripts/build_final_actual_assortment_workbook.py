"""Build final actual assortment workbook with ClickHouse product names."""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from scripts.build_required_assortment_contract import normalize_text  # noqa: E402

REPORT_DIR = ROOT / "reports" / "required_assortment"
ASSORTMENT_PATH = REPORT_DIR / "assortment_city_products.csv"
FEEDBACK_AUDIT_PATH = REPORT_DIR / "partner_feedback_audit.csv"
OUTPUT_PATH = REPORT_DIR / "actual_assortment_final.xlsx"


NO_PLAN_ALIASES = {
    "Жар пиццы с салями": ["жарпицца с салями", "пицца с салями п"],
    "Маковки нет": ["маковка"],
    "Сметанник маковый": ["сметанник маковый"],
    "киш грибы курица": ["киш грибы курица"],
    "пирожок зеленый лук": ["пирожок зеленый лук"],
    "корзинка ягодная": ["корзинка ягодная"],
    "пирог Ягодный": ["пирог ягодный"],
    "пирог с манго": ["пирог с манго"],
    "пирог с черносливом и грецким орехом": [
        "пирог с черносливом и грец орехом",
        "пирог с черносливом и грецким орехом",
    ],
    "жар киш грибы и курица": ["жар киш грибы курица"],
    "сэендвич мюнхенский": ["сэндвич мюнхенский"],
    "Московская плюшка": ["московская плюшка п", "московская плюшка"],
}

DISCONTINUED_ALIASES = {
    "Пирожок капуста курица": [
        "пирожок капуста курица",
        "пирожок капуста и курица",
    ],
    "Ватрушка в ассортименте": ["ватрушка в ассортименте"],
    "Пирожок/Булочка с яблоками (тесто ночное)": [
        "булочка с яблоком",
        "пирожок с яблоком п",
    ],
    "Булочка с вишней (тесто ночное)": ["булочка с вишней"],
    "Вишневый": ["вишневый"],
    "Мандариновый пай": ["мандариновый пай"],
}


def norm_aliases(values: list[str]) -> set[str]:
    return {normalize_text(value) for value in values}


def mark_alias_group(
    df: pd.DataFrame,
    *,
    aliases: dict[str, list[str]],
    flag_column: str,
    note_column: str,
) -> pd.DataFrame:
    work = df.copy()
    work[flag_column] = 0
    work[note_column] = ""
    for raw_name, names in aliases.items():
        keys = norm_aliases(names)
        mask = work["product_key_text"].isin(keys)
        work.loc[mask, flag_column] = 1
        work.loc[mask, note_column] = raw_name
    return work


def build_workbook() -> Path:
    assortment = pd.read_csv(ASSORTMENT_PATH, dtype={"product_id": str})
    assortment["product_key_text"] = assortment["product_name"].map(normalize_text)
    assortment = mark_alias_group(
        assortment,
        aliases=NO_PLAN_ALIASES,
        flag_column="no_baking_plan",
        note_column="no_baking_plan_source_name",
    )
    assortment = mark_alias_group(
        assortment,
        aliases=DISCONTINUED_ALIASES,
        flag_column="is_discontinued_feedback",
        note_column="discontinued_source_name",
    )

    excluded = assortment[assortment["is_discontinued_feedback"].eq(1)].copy()
    final = assortment[~assortment["is_discontinued_feedback"].eq(1)].copy()

    final_columns = [
        "city",
        "product_id",
        "product_name",
        "category_name",
        "no_baking_plan",
        "no_baking_plan_source_name",
        "source",
        "source_file",
        "valid_from",
        "comment",
    ]
    final = final[final_columns].sort_values(["city", "category_name", "product_name"])

    city_matrix = (
        final.assign(in_assortment=1)
        .pivot_table(
            index=[
                "product_id",
                "product_name",
                "category_name",
                "no_baking_plan",
                "no_baking_plan_source_name",
            ],
            columns="city",
            values="in_assortment",
            aggfunc="max",
            fill_value=0,
        )
        .reset_index()
    )
    city_matrix.columns.name = None

    summary = (
        final.groupby(["city"], as_index=False)
        .agg(
            products=("product_id", "nunique"),
            no_baking_plan=("no_baking_plan", "sum"),
        )
        .sort_values("city")
    )
    summary.loc[len(summary)] = {
        "city": "ИТОГО уникальных product_id",
        "products": final["product_id"].nunique(),
        "no_baking_plan": final.loc[
            final["no_baking_plan"].eq(1),
            "product_id",
        ].nunique(),
    }

    feedback = pd.read_csv(FEEDBACK_AUDIT_PATH)
    unmatched_feedback = feedback[
        feedback["match_status"].eq("not_found")
        | (
            feedback["present_in_new_assortment"].fillna(False).eq(False)
            & feedback["present_in_prod_forecast"].fillna(False).eq(True)
        )
    ].copy()

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        city_matrix.to_excel(writer, sheet_name="Actual assortment matrix", index=False)
        final.to_excel(writer, sheet_name="Actual assortment rows", index=False)
        excluded.to_excel(writer, sheet_name="Excluded discontinued", index=False)
        unmatched_feedback.to_excel(
            writer,
            sheet_name="Feedback unresolved",
            index=False,
        )

    format_workbook(OUTPUT_PATH)
    return OUTPUT_PATH


def format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    warning_fill = PatternFill("solid", fgColor="FEF3C7")
    excluded_fill = PatternFill("solid", fgColor="FEE2E2")

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = border
        for row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
            max_col=worksheet.max_column,
        ):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        if worksheet.max_row >= 2 and worksheet.max_column >= 1:
            reference = (
                f"A1:{get_column_letter(worksheet.max_column)}"
                f"{worksheet.max_row}"
            )
            table_name = "T" + "".join(
                character for character in worksheet.title if character.isalnum()
            )[:20]
            table = Table(displayName=table_name, ref=reference)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)
        for column_idx in range(1, worksheet.max_column + 1):
            letter = get_column_letter(column_idx)
            values = [
                str(worksheet.cell(row=row_idx, column=column_idx).value or "")
                for row_idx in range(1, min(worksheet.max_row, 80) + 1)
            ]
            worksheet.column_dimensions[letter].width = min(
                max(max(len(value) for value in values) + 2, 10),
                55,
            )

    if "Actual assortment rows" in workbook.sheetnames:
        sheet = workbook["Actual assortment rows"]
        headers = [cell.value for cell in sheet[1]]
        if "no_baking_plan" in headers:
            flag_idx = headers.index("no_baking_plan") + 1
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=flag_idx).value == 1:
                    for cell in sheet[row]:
                        cell.fill = warning_fill

    if "Excluded discontinued" in workbook.sheetnames:
        sheet = workbook["Excluded discontinued"]
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.fill = excluded_fill

    workbook.save(path)


def main() -> None:
    output = build_workbook()
    print(output)


if __name__ == "__main__":
    main()
