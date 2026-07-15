"""
Compute per-(bakery, product, hour) stockout correction factors and write to
sku_hour_stockout_correction_embedded.

For each bakery with production data:
  1. Load baking-plan template windows -> coverage_hours map
  2. Detect stockout days: daily_sold / daily_produced >= STOCKOUT_RATIO
  3. Per (bakery, product, coverage_window): compute
       correction = 1 + stockout_rate * avg_missed / avg_sold_per_day
  4. Map each coverage window to its hours, write correction_factor per (bakery, product, hour)
  5. Hours not in any stockout-prone window get correction_factor = 1.0

Run (dev first):
  .venv/Scripts/python.exe -m scripts.build_stockout_correction --env-file .env.dev --write-table
  .venv/Scripts/python.exe -m scripts.build_stockout_correction --env-file .env.dev --write-table --dry-run
"""

from __future__ import annotations

import argparse
import sys
from datetime import date, datetime, timezone
from pathlib import Path

import pandas as pd
import clickhouse_connect
from openpyxl import load_workbook

try:
    from dotenv import dotenv_values
except ImportError:
    def dotenv_values(path):
        vals = {}
        for line in Path(path).read_text().splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            vals[k.strip()] = v.strip().strip('"').strip("'")
        return vals

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "apps"))

from baking_plan.templates import (
    parse_windows,
    COMMENTS_SHEET_NAME,
    INDIVIDUAL_TEMPLATES_DIR,
    BASE_TEMPLATE_PATH,
)
from baking_plan.allocation import coverage_hours

# ── config ────────────────────────────────────────────────────────────────────
CORRECTION_TABLE = "sku_hour_stockout_correction_embedded"

# Bakeries for the pilot — others will be skipped until more validation done.
PILOT_BAKERY_IDS = {16, 20, 21, 22, 257}

LOOKBACK_DAYS     = 60
STOCKOUT_RATIO    = 0.90   # продано/выпуск >= this -> stockout day
MIN_BAKERY_ACTIVE = 3.0    # bakery-total units/hour to confirm bakery was open
MIN_STOCKOUT_DAYS = 3      # require at least this many stockout events to apply correction
MIN_STOCKOUT_RATE = 0.15   # require at least 15% of days to be stockout days
MAX_CORRECTION    = 2.0    # cap

BAKEABLE_CATEGORIES = {
    "Пироги сытные", "Пироги сладкие",
    "Выпечка сытная", "Выпечка сладкая", "Фастфуд",
}

CREATE_TABLE_DDL = f"""
CREATE TABLE IF NOT EXISTS {CORRECTION_TABLE} (
    bakery_id        Int64,
    product_id       Int64,
    hour             Int16,
    correction_factor Float64,
    profile_version  String,
    generated_at     DateTime64(3)
) ENGINE = ReplacingMergeTree(generated_at)
ORDER BY (bakery_id, product_id, hour, profile_version)
"""


# ── template helpers ──────────────────────────────────────────────────────────

def _load_windows_for_bakery(
    bakery_id_int: int,
    revenue_bucket: str | None,
) -> dict[str, list[int]]:
    """Return coverage_hours map for a bakery."""
    # Try individual template first
    for path in sorted(INDIVIDUAL_TEMPLATES_DIR.glob(f"{bakery_id_int}_*.xlsx")):
        wb = load_workbook(path, read_only=True, data_only=True)
        data_sheets = [s for s in wb.sheetnames if s != COMMENTS_SHEET_NAME]
        if data_sheets:
            windows = parse_windows(wb, data_sheets[0])
            wb.close()
            return _dedup_coverage(windows)

    # Fall back to base template with revenue-tier sheet
    if BASE_TEMPLATE_PATH.exists():
        wb = load_workbook(BASE_TEMPLATE_PATH, read_only=True, data_only=True)
        bucket = revenue_bucket or "до 2,5 млн"
        sheet_names = [s for s in wb.sheetnames if s != COMMENTS_SHEET_NAME]
        sheet = bucket if bucket in wb.sheetnames else (sheet_names[0] if sheet_names else None)
        if sheet:
            windows = parse_windows(wb, sheet)
            wb.close()
            return _dedup_coverage(windows)
        wb.close()

    # Minimal fallback: two windows (morning + evening)
    from baking_plan.templates import Window
    windows = [
        Window(label="08:00-13:00", column=3, start_hour=8,  end_hour=13),
        Window(label="14:00-17:00", column=4, start_hour=14, end_hour=17),
    ]
    return _dedup_coverage(windows)


def _dedup_coverage(windows) -> dict[str, list[int]]:
    seen: set[str] = set()
    unique = []
    for w in windows:
        if w.label not in seen:
            unique.append(w)
            seen.add(w.label)
    cov = coverage_hours(unique)
    return {lbl: hrs for lbl, hrs in cov.items() if hrs}


# ── main computation ──────────────────────────────────────────────────────────

def build_corrections(
    client,
    *,
    profile_version: str,
    date_from: str,
    date_to: str,
) -> pd.DataFrame:
    """Return a DataFrame with correction factors per (bakery_id, product_id, hour)."""

    # 1. Revenue buckets for template sheet selection
    revenue_df = client.query_df("""
        SELECT bakery_id, argMax(revenue_bucket, month_start) AS revenue_bucket
        FROM bakery_month_revenue_embedded
        GROUP BY bakery_id
    """)
    revenue_by_bid = dict(zip(revenue_df["bakery_id"].astype(int), revenue_df["revenue_bucket"]))

    # 2. Daily production per bakery-product
    # bakery_id/product_id in fct_production_release are Strings — filter with padded strings,
    # convert to Int64 in SELECT using a different alias to avoid CH alias-shadowing in WHERE.
    bids_str_prod = [str(b).zfill(9) for b in PILOT_BAKERY_IDS]
    prod_df = client.query_df(
        """
        SELECT release_date AS date,
               toInt64OrZero(bakery_id)  AS bid,
               toInt64OrZero(product_id) AS pid,
               sum(toFloat64(quantity))  AS vypusk
        FROM fct_production_release
        WHERE release_date BETWEEN %(d0)s AND %(d1)s
          AND is_deleted != '1'
          AND bakery_id IN %(bids)s
        GROUP BY date, bid, pid
        """,
        parameters={"d0": date_from, "d1": date_to, "bids": bids_str_prod},
    )
    prod_df = prod_df.rename(columns={"bid": "bakery_id", "pid": "product_id"})
    prod_df["date"] = pd.to_datetime(prod_df["date"])

    # 3. Hourly sales for pilot bakeries
    # Use different alias names (bid/pid) to avoid ClickHouse alias-shadowing in WHERE
    bids_str = [str(b).zfill(9) for b in PILOT_BAKERY_IDS]
    sales_df = client.query_df(
        """
        SELECT toDate(check_datetime)       AS date,
               toHour(check_datetime)       AS hour,
               toInt64OrZero(bakery_id)     AS bid,
               toInt64OrZero(product_id)    AS pid,
               any(category_name)           AS category_name,
               sum(quantity)                AS sold
        FROM mart_sales_60d
        WHERE bakery_id IN %(bids)s
          AND check_date BETWEEN %(d0)s AND %(d1)s
          AND cash_event_type = %(sale_type)s
        GROUP BY date, hour, bid, pid
        """,
        parameters={"bids": bids_str, "d0": date_from, "d1": date_to, "sale_type": "Продажа"},
    )
    sales_df = sales_df.rename(columns={"bid": "bakery_id", "pid": "product_id"})
    sales_df["date"] = pd.to_datetime(sales_df["date"])
    sales_df = sales_df[sales_df["category_name"].isin(BAKEABLE_CATEGORIES)].copy()

    # Bakery-hour activity
    bakery_active = (
        sales_df.groupby(["date", "bakery_id", "hour"])["sold"]
        .sum()
        .rename("bakery_total")
        .reset_index()
    )
    bakery_lookup = {
        (r.date, r.bakery_id, r.hour): r.bakery_total
        for r in bakery_active.itertuples()
    }

    # Daily sold per (bakery, product)
    daily_sold = (
        sales_df.groupby(["date", "bakery_id", "product_id"])["sold"]
        .sum()
        .reset_index()
        .rename(columns={"sold": "daily_sold"})
    )

    # 4. Join with production -> detect stockout days
    merged = daily_sold.merge(prod_df, on=["date", "bakery_id", "product_id"], how="inner")
    merged["ratio"] = merged["daily_sold"] / merged["vypusk"].replace(0.0, float("nan"))
    stockout_mask = merged["ratio"] >= STOCKOUT_RATIO
    stockout_days = merged[stockout_mask].copy()

    # Hourly lookup for fast access
    hourly_lookup = {
        (r.date, r.bakery_id, r.product_id, r.hour): r.sold
        for r in sales_df.itertuples()
    }

    all_rows: list[dict] = []
    generated_at = datetime.now(tz=timezone.utc)

    for bakery_id_int in PILOT_BAKERY_IDS:
        bucket = revenue_by_bid.get(bakery_id_int)
        coverage_map = _load_windows_for_bakery(bakery_id_int, bucket)
        if not coverage_map:
            continue

        b_stockout = stockout_days[stockout_days["bakery_id"] == bakery_id_int]
        b_sales    = sales_df[sales_df["bakery_id"] == bakery_id_int]

        if b_stockout.empty or b_sales.empty:
            continue

        # Per-product stats
        product_ids = b_sales["product_id"].unique()

        # Avg daily sold across ALL days (not just stockout)
        avg_daily_sold = (
            daily_sold[daily_sold["bakery_id"] == bakery_id_int]
            .groupby("product_id")["daily_sold"]
            .mean()
            .to_dict()
        )
        total_days = b_sales["date"].nunique()

        for pid in product_ids:
            p_stockout = b_stockout[b_stockout["product_id"] == pid]
            if p_stockout.empty:
                continue

            n_stockout = len(p_stockout)
            stockout_rate = n_stockout / max(total_days, 1)
            avg_sold = avg_daily_sold.get(pid, 0.0)
            if avg_sold <= 0:
                continue

            # Per coverage window: estimate missed demand on stockout days
            window_correction: dict[str, float] = {}

            for win_label, hours in coverage_map.items():
                if not hours:
                    continue

                total_missed = 0.0
                counted = 0

                for _, row in p_stockout.iterrows():
                    date_val = row["date"]
                    win_sales = {
                        h: hourly_lookup.get((date_val, bakery_id_int, pid, h), 0.0)
                        for h in hours
                    }
                    sold_hours = [h for h, s in win_sales.items() if s > 0]
                    if not sold_hours:
                        continue

                    last_sale_h = max(sold_hours)
                    if last_sale_h >= max(hours):
                        continue

                    stockout_hrs = [
                        h for h in hours
                        if h > last_sale_h
                        and bakery_lookup.get((date_val, bakery_id_int, h), 0.0) >= MIN_BAKERY_ACTIVE
                    ]
                    if not stockout_hrs:
                        continue

                    selling_qty = [win_sales[h] for h in hours if h <= last_sale_h and win_sales[h] > 0]
                    avg_rate = sum(selling_qty) / len(selling_qty) if selling_qty else 0.0
                    total_missed += avg_rate * len(stockout_hrs)
                    counted += 1

                if counted < MIN_STOCKOUT_DAYS or stockout_rate < MIN_STOCKOUT_RATE:
                    continue

                avg_missed = total_missed / counted
                # E[correction] = (avg_sold + stockout_rate * avg_missed) / avg_sold
                correction = 1.0 + stockout_rate * avg_missed / avg_sold
                correction = min(correction, MAX_CORRECTION)
                if correction > 1.001:
                    window_correction[win_label] = correction

            if not window_correction:
                continue

            # Map coverage windows -> hours -> write correction_factor
            for win_label, hours in coverage_map.items():
                factor = window_correction.get(win_label, 1.0)
                for h in hours:
                    all_rows.append({
                        "bakery_id":        bakery_id_int,
                        "product_id":       int(pid),
                        "hour":             h,
                        "correction_factor": factor,
                        "profile_version":  profile_version,
                        "generated_at":     generated_at,
                    })

    if not all_rows:
        return pd.DataFrame(columns=[
            "bakery_id", "product_id", "hour",
            "correction_factor", "profile_version", "generated_at",
        ])

    result = pd.DataFrame(all_rows)
    # One row per (bakery, product, hour) — take max correction if duplicates
    result = (
        result.groupby(["bakery_id", "product_id", "hour", "profile_version", "generated_at"])
        ["correction_factor"].max().reset_index()
    )
    return result


# ── CLI ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Build stockout hour correction table")
    parser.add_argument("--env-file", default=str(ROOT / ".env.dev"))
    parser.add_argument("--date-from", default=None, help="YYYY-MM-DD, default 60 days ago")
    parser.add_argument("--date-to",   default=None, help="YYYY-MM-DD, default yesterday")
    parser.add_argument("--profile-version", default=None, help="version tag, default stockout_YYYYMMDD")
    parser.add_argument("--write-table", action="store_true", help="write to ClickHouse")
    parser.add_argument("--dry-run",  action="store_true", help="print stats only, no write")
    args = parser.parse_args()

    today = date.today()
    date_to   = args.date_to   or str(today - __import__("datetime").timedelta(days=1))
    date_from = args.date_from or str(today - __import__("datetime").timedelta(days=LOOKBACK_DAYS))
    version   = args.profile_version or f"stockout_{today.strftime('%Y%m%d')}"

    env = dotenv_values(args.env_file)
    _secure = str(env.get("CLICKHOUSE_SECURE", "true")).lower() not in ("0", "false", "no")
    _verify = str(env.get("CLICKHOUSE_VERIFY", "true")).lower() not in ("0", "false", "no")
    client = clickhouse_connect.get_client(
        host=env["CLICKHOUSE_HOST"],
        port=int(env["CLICKHOUSE_PORT"]),
        username=env["CLICKHOUSE_USER"],
        password=env["CLICKHOUSE_PASSWORD"],
        database=env["CLICKHOUSE_DATABASE"],
        secure=_secure,
        verify=_verify,
    )

    print(f"Building stockout corrections  {date_from} -> {date_to}  version={version}")
    print(f"Pilot bakeries: {sorted(PILOT_BAKERY_IDS)}")

    df = build_corrections(client, profile_version=version, date_from=date_from, date_to=date_to)

    if df.empty:
        print("No corrections computed — check data / thresholds.")
        return

    print(f"\nComputed {len(df)} (bakery, product, hour) rows")
    print(f"  Bakeries: {sorted(df['bakery_id'].unique())}")
    print(f"  Products: {df['product_id'].nunique()}")
    print(f"  Hours with correction>1: {(df['correction_factor'] > 1.001).sum()}")
    print()

    # Stats by hour
    hour_stats = (
        df[df["correction_factor"] > 1.001]
        .groupby("hour")
        .agg(skus=("product_id", "nunique"), mean_cf=("correction_factor", "mean"))
        .reset_index()
        .sort_values("hour")
    )
    print("Correction factors by hour (where >1):")
    print(hour_stats.to_string(index=False))

    if args.dry_run or not args.write_table:
        print("\n[dry-run] not writing to ClickHouse")
        return

    # Create table + insert
    print(f"\nCreating table {CORRECTION_TABLE} if not exists...")
    client.command(CREATE_TABLE_DDL)

    print(f"Inserting {len(df)} rows...")
    client.insert_df(CORRECTION_TABLE, df)
    print("Done.")


if __name__ == "__main__":
    main()
