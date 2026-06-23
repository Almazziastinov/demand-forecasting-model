from __future__ import annotations

# ruff: noqa: E501

import json
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs" / "pie_audit_new_predictions_20260616_20260622"
OUTPUT_XLSX = OUTPUT_DIR / "pie_audit_new_predictions_dev_uplifted_2026-06-16_2026-06-22.xlsx"
OUTPUT_JSON = OUTPUT_DIR / "pie_audit_new_predictions_data.json"
PREVIOUS_FACT_JSON = ROOT / "outputs" / "pie_audit_20260616_20260622" / "pie_audit_data_fixed_product_id.json"
FORECAST_CSV = ROOT / "data" / "processed" / "sku_day_forecast_prod_uplifted_bakery_norm_uplift_sku.csv"

RUN_ID = "dev_uplifted_bakery_norm_uplift_sku_20260616_h7"
START_DATE = "2026-06-16"
END_DATE = "2026-06-22"
PILOT_IDS = [20, 21, 22, 28, 80, 89, 107, 221, 222, 257]
CATEGORIES = ["\u041f\u0438\u0440\u043e\u0433\u0438 \u0441\u044b\u0442\u043d\u044b\u0435", "\u041f\u0438\u0440\u043e\u0433\u0438 \u0441\u043b\u0430\u0434\u043a\u0438\u0435"]


def normalize_product_id(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if not text:
        return ""
    text = re.sub(r"\.0$", "", text)
    stripped = text.lstrip("0")
    return stripped or "0"


def fix_mojibake(value: object) -> object:
    if not isinstance(value, str):
        return value
    if not any(marker in value for marker in ("Р", "С")):
        return value
    try:
        return value.encode("cp1251").decode("utf-8")
    except UnicodeError:
        return value


def safe_sheet_name(value: str) -> str:
    clean = re.sub(r"[\[\]\:\*\?\/\\]", " ", value).strip()
    return clean[:31] or "sheet"


def load_previous_facts() -> pd.DataFrame:
    payload = json.loads(PREVIOUS_FACT_JSON.read_text(encoding="utf-8"))
    facts = pd.DataFrame(payload["rows"])
    facts["forecast_date"] = pd.to_datetime(facts["date"]).dt.date
    facts["bakery_id"] = pd.to_numeric(facts["bakery_id"], errors="coerce").astype("Int64")
    facts["bakery_name"] = facts["bakery_name"].map(fix_mojibake)
    facts["product_id_norm"] = facts["product_id"].map(normalize_product_id)
    facts["product_name"] = facts["product_name"].map(fix_mojibake)
    facts["category_name"] = facts["category_name"].map(fix_mojibake)
    facts["fact_qty"] = pd.to_numeric(facts["fact_qty"], errors="coerce").fillna(0.0)
    facts = facts[facts["bakery_id"].isin(PILOT_IDS) & facts["category_name"].isin(CATEGORIES)].copy()
    return facts[
        ["forecast_date", "bakery_id", "bakery_name", "product_id_norm", "product_name", "category_name", "fact_qty"]
    ]


def load_lookup(forecast: pd.DataFrame) -> pd.DataFrame:
    wanted_products = set(forecast["product_id_norm"].dropna().astype(str).unique())
    chunks: list[pd.DataFrame] = []
    profile_path = ROOT / "data" / "processed" / "sku_hour_share_profile_smoothed.clickhouse.csv"
    if not profile_path.exists():
        profile_path = ROOT / "data" / "processed" / "sku_hour_share_profile_smoothed.csv"

    for chunk in pd.read_csv(
        profile_path,
        encoding="utf-8-sig",
        usecols=["bakery_id", "product_id", "product_name", "category_name"],
        chunksize=500_000,
    ):
        chunk["bakery_id"] = pd.to_numeric(chunk["bakery_id"], errors="coerce").astype("Int64")
        chunk["product_id_norm"] = chunk["product_id"].map(normalize_product_id)
        chunk["product_name"] = chunk["product_name"].map(fix_mojibake)
        chunk["category_name"] = chunk["category_name"].map(fix_mojibake)
        chunk = chunk[
            chunk["bakery_id"].isin(PILOT_IDS)
            & chunk["product_id_norm"].isin(wanted_products)
            & chunk["category_name"].isin(CATEGORIES)
        ]
        if not chunk.empty:
            chunks.append(chunk[["bakery_id", "product_id_norm", "product_name", "category_name"]])
    if not chunks:
        return pd.DataFrame(columns=["bakery_id", "product_id_norm", "product_name", "category_name"])
    return pd.concat(chunks, ignore_index=True).drop_duplicates(["bakery_id", "product_id_norm"])


def load_new_forecast() -> pd.DataFrame:
    forecast = pd.read_csv(FORECAST_CSV, encoding="utf-8-sig")
    forecast["forecast_date"] = pd.to_datetime(forecast["date"]).dt.date
    forecast["bakery_id"] = pd.to_numeric(forecast["bakery_id"], errors="coerce").astype("Int64")
    forecast["product_id_norm"] = forecast["product_id"].map(normalize_product_id)
    forecast["forecast_qty"] = pd.to_numeric(forecast["sku_day_forecast"], errors="coerce").fillna(0.0)
    forecast = forecast[
        forecast["forecast_date"].between(pd.Timestamp(START_DATE).date(), pd.Timestamp(END_DATE).date())
        & forecast["bakery_id"].isin(PILOT_IDS)
    ].copy()
    lookup = load_lookup(forecast)
    forecast = forecast.merge(lookup, on=["bakery_id", "product_id_norm"], how="left")
    forecast = forecast[forecast["category_name"].isin(CATEGORIES)].copy()
    return forecast[["forecast_date", "bakery_id", "product_id_norm", "product_name", "category_name", "forecast_qty"]]


def build_dataset() -> tuple[pd.DataFrame, dict[int, str]]:
    facts = load_previous_facts()
    forecast = load_new_forecast()
    bakery_names = (
        facts[["bakery_id", "bakery_name"]]
        .dropna()
        .drop_duplicates("bakery_id")
        .set_index("bakery_id")["bakery_name"]
        .to_dict()
    )

    merged = facts.merge(
        forecast,
        on=["forecast_date", "bakery_id", "product_id_norm"],
        how="outer",
        suffixes=("_fact", "_forecast"),
    )
    merged["bakery_name"] = merged["bakery_id"].map(bakery_names)
    merged["product_name"] = merged["product_name_forecast"].combine_first(merged["product_name_fact"])
    merged["category_name"] = merged["category_name_forecast"].combine_first(merged["category_name_fact"])
    merged["fact_qty"] = pd.to_numeric(merged["fact_qty"], errors="coerce").fillna(0.0)
    merged["forecast_qty"] = pd.to_numeric(merged["forecast_qty"], errors="coerce").fillna(0.0)
    merged = merged[merged["category_name"].isin(CATEGORIES)].copy()

    grouped = (
        merged.groupby(["forecast_date", "bakery_id", "bakery_name", "category_name", "product_name"], as_index=False)[
            ["fact_qty", "forecast_qty"]
        ]
        .sum()
        .sort_values(["bakery_id", "category_name", "product_name", "forecast_date"])
    )
    return grouped, bakery_names


def add_title(ws, title: str, subtitle: str, max_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.cell(1, 1, title)
    ws.cell(2, 1, subtitle)
    ws.cell(1, 1).font = Font(bold=True, size=14, color="1F2937")
    ws.cell(2, 1).font = Font(size=10, color="4B5563")


def build_workbook(data: pd.DataFrame, bakery_names: dict[int, str]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    dates = [item.date() for item in pd.date_range(START_DATE, END_DATE)]
    max_col = 2 + len(dates) * 2 + 4
    wb = Workbook()
    wb.remove(wb.active)

    date_fills = ["DBEAFE", "DCFCE7", "FEF3C7", "FCE7F3", "E0E7FF", "CCFBF1", "FFEDD5"]
    sku_fills = ["FFFFFF", "F8FAFC", "F1F5F9", "EFF6FF"]
    group_fill = PatternFill("solid", fgColor="111827")
    header_fill = PatternFill("solid", fgColor="E5E7EB")
    fact_fill = PatternFill("solid", fgColor="F9FAFB")
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    summary_rows = []

    for bakery_id in PILOT_IDS:
        bakery_name = bakery_names.get(bakery_id, str(bakery_id))
        ws = wb.create_sheet(safe_sheet_name(f"{bakery_id} {bakery_name}"))
        add_title(
            ws,
            f"{bakery_id} {bakery_name}",
            f"{RUN_ID}; {START_DATE} - {END_DATE}; \u0444\u0430\u043a\u0442 \u0438\u0437 \u043f\u0440\u0435\u0434\u044b\u0434\u0443\u0449\u0435\u0433\u043e audit-\u0444\u0430\u0439\u043b\u0430",
            max_col,
        )
        ws.cell(4, 1, "\u0413\u0440\u0443\u043f\u043f\u0430")
        ws.cell(4, 2, "\u041d\u043e\u043c\u0435\u043d\u043a\u043b\u0430\u0442\u0443\u0440\u0430")
        col = 3
        for idx, day in enumerate(dates):
            fill = PatternFill("solid", fgColor=date_fills[idx % len(date_fills)])
            ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
            ws.cell(4, col, day)
            ws.cell(4, col).fill = fill
            ws.cell(4, col + 1).fill = fill
            ws.cell(5, col, "\u0424\u0430\u043a\u0442")
            ws.cell(5, col + 1, "\u041f\u0440\u043e\u0433\u043d\u043e\u0437")
            col += 2
        total_fact_col = col
        ws.cell(4, col, "\u0424\u0430\u043a\u0442 \u0438\u0442\u043e\u0433\u043e")
        ws.cell(4, col + 1, "\u041f\u0440\u043e\u0433\u043d\u043e\u0437 \u0438\u0442\u043e\u0433\u043e")
        ws.cell(4, col + 2, "\u041e\u0442\u043a\u043b.")
        ws.cell(4, col + 3, "\u041e\u0442\u043a\u043b., %")

        for row in ws.iter_rows(min_row=4, max_row=5, min_col=1, max_col=max_col):
            for cell in row:
                cell.font = Font(bold=True, color="111827")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
                if cell.fill.fill_type is None:
                    cell.fill = header_fill

        row_idx = 6
        bakery_data = data[data["bakery_id"] == bakery_id]
        for category in CATEGORIES:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_col)
            ws.cell(row_idx, 1, category)
            ws.cell(row_idx, 1).fill = group_fill
            ws.cell(row_idx, 1).font = Font(bold=True, color="FFFFFF")
            row_idx += 1
            cat_data = bakery_data[bakery_data["category_name"] == category]
            products = sorted(cat_data["product_name"].dropna().unique().tolist())
            for sku_idx, product_name in enumerate(products):
                row_fill = PatternFill("solid", fgColor=sku_fills[sku_idx % len(sku_fills)])
                ws.cell(row_idx, 1, category)
                ws.cell(row_idx, 2, product_name)
                by_date = cat_data[cat_data["product_name"] == product_name].set_index("forecast_date")
                col = 3
                for date_idx, day in enumerate(dates):
                    fact = float(by_date["fact_qty"].get(day, 0.0)) if day in by_date.index else 0.0
                    forecast = float(by_date["forecast_qty"].get(day, 0.0)) if day in by_date.index else 0.0
                    ws.cell(row_idx, col, fact)
                    ws.cell(row_idx, col + 1, forecast)
                    ws.cell(row_idx, col).fill = fact_fill
                    ws.cell(row_idx, col + 1).fill = PatternFill("solid", fgColor=date_fills[date_idx % len(date_fills)])
                    col += 2
                fact_cols = [get_column_letter(3 + idx * 2) for idx in range(len(dates))]
                forecast_cols = [get_column_letter(4 + idx * 2) for idx in range(len(dates))]
                ws.cell(row_idx, total_fact_col, f"=SUM({','.join(col + str(row_idx) for col in fact_cols)})")
                ws.cell(row_idx, total_fact_col + 1, f"=SUM({','.join(col + str(row_idx) for col in forecast_cols)})")
                ws.cell(row_idx, total_fact_col + 2, f"={get_column_letter(total_fact_col + 1)}{row_idx}-{get_column_letter(total_fact_col)}{row_idx}")
                ws.cell(row_idx, total_fact_col + 3, f"=IF({get_column_letter(total_fact_col)}{row_idx}=0,\"\",{get_column_letter(total_fact_col + 2)}{row_idx}/{get_column_letter(total_fact_col)}{row_idx})")
                for cell in ws[row_idx]:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="right" if cell.column >= 3 else "left")
                    if cell.column <= 2:
                        cell.fill = row_fill
                    if cell.column >= 3:
                        cell.number_format = "#,##0.0"
                ws.cell(row_idx, total_fact_col + 3).number_format = "0.0%"
                row_idx += 1

        total_fact = float(bakery_data["fact_qty"].sum())
        total_forecast = float(bakery_data["forecast_qty"].sum())
        summary_rows.append([bakery_id, bakery_name, total_fact, total_forecast, total_forecast - total_fact, (total_forecast - total_fact) / total_fact if total_fact else None])
        ws.freeze_panes = "C6"
        ws.auto_filter.ref = f"A5:{get_column_letter(max_col)}{max(row_idx - 1, 5)}"
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 34
        for col_idx in range(3, max_col + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 11

    summary = wb.create_sheet("\u0418\u0442\u043e\u0433\u0438", 0)
    add_title(summary, "\u0418\u0442\u043e\u0433\u0438 \u043f\u043e \u043d\u043e\u0432\u044b\u043c dev-\u043f\u0440\u043e\u0433\u043d\u043e\u0437\u0430\u043c", f"{RUN_ID}; {START_DATE} - {END_DATE}", 6)
    headers = ["bakery_id", "\u041f\u0435\u043a\u0430\u0440\u043d\u044f", "\u0424\u0430\u043a\u0442", "\u041f\u0440\u043e\u0433\u043d\u043e\u0437", "\u041e\u0442\u043a\u043b.", "\u041e\u0442\u043a\u043b., %"]
    for col_idx, header in enumerate(headers, 1):
        summary.cell(4, col_idx, header)
    for row_idx, row in enumerate(summary_rows, 5):
        for col_idx, value in enumerate(row, 1):
            summary.cell(row_idx, col_idx, value)
    for row in summary.iter_rows(min_row=4, max_row=4 + len(summary_rows), min_col=1, max_col=6):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if cell.column >= 3 else "left")
            if cell.row == 4:
                cell.fill = header_fill
                cell.font = Font(bold=True)
            if cell.column in {3, 4, 5} and cell.row > 4:
                cell.number_format = "#,##0.0"
            if cell.column == 6 and cell.row > 4:
                cell.number_format = "0.0%"
    summary.column_dimensions["A"].width = 10
    summary.column_dimensions["B"].width = 36
    for col_idx in range(3, 7):
        summary.column_dimensions[get_column_letter(col_idx)].width = 14
    summary.freeze_panes = "A5"
    wb.save(OUTPUT_XLSX)


def verify(data: pd.DataFrame) -> dict[str, object]:
    duplicates = (
        data[["bakery_id", "category_name", "product_name"]]
        .drop_duplicates()
        .groupby(["bakery_id", "category_name", "product_name"])
        .size()
        .reset_index(name="rows")
    )
    duplicates = duplicates[duplicates["rows"] > 1]
    return {
        "run_id": RUN_ID,
        "start_date": START_DATE,
        "end_date": END_DATE,
        "rows": int(len(data)),
        "bakeries": int(data["bakery_id"].nunique()),
        "categories": sorted(data["category_name"].dropna().unique().tolist()),
        "total_fact_qty": float(data["fact_qty"].sum()),
        "total_forecast_qty": float(data["forecast_qty"].sum()),
        "duplicate_product_names": duplicates.to_dict("records"),
        "output_xlsx": str(OUTPUT_XLSX),
        "fact_source": str(PREVIOUS_FACT_JSON),
        "forecast_source": str(FORECAST_CSV),
    }


def main() -> None:
    data, bakery_names = build_dataset()
    if data.empty:
        raise RuntimeError("No rows for pie audit.")
    build_workbook(data, bakery_names)
    summary = verify(data)
    OUTPUT_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
