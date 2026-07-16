"""Classify buyer review sentiment through the existing VibeCode API setup.

The script reads review text from selected worksheets and writes a JSON file
with row-level sentiment labels. It keeps a checkpoint so interrupted runs can
be resumed without reclassifying completed rows.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import time
from pathlib import Path
from typing import Any

import openpyxl
import requests


BASE_URL = "https://vibecode.bitrix24.tech/v1"
MODEL = "bitrix/bitrixgpt-5.5"
BATCH_SIZE = 25

DEFAULT_INPUT = Path(r"C:\Users\dns\Downloads\Отзывы покупателей.xlsx")
PROJECT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_CHECKPOINT = PROJECT_DIR / "outputs" / "buyer_reviews_sentiment_checkpoint.json"
DEFAULT_RESULTS = PROJECT_DIR / "outputs" / "buyer_reviews_sentiment_results.json"
LEGACY_SCRIPT = PROJECT_DIR / "analyze_reviews.py"

SHEET_CONFIGS = [
    {"sheet": "Отзывы 2025", "text_header": "Отзыв"},
    {"sheet": "2024-25 Жалобы", "text_header": "Суть жалобы"},
    {"sheet": "2022 Полож отзывы", "text_header": "Суть жалобы"},
]

SENTIMENTS = {"позитивный", "нейтральный", "негативный"}


def load_api_key() -> str:
    env_key = os.getenv("VIBECODE_API_KEY")
    if env_key:
        return env_key

    if LEGACY_SCRIPT.exists():
        text = LEGACY_SCRIPT.read_text(encoding="utf-8", errors="ignore")
        match = re.search(r'API_KEY\s*=\s*["\']([^"\']+)["\']', text)
        if match:
            return match.group(1)

    raise RuntimeError(
        "VIBECODE_API_KEY is not set and API_KEY was not found in analyze_reviews.py"
    )


def ai_complete(prompt: str, api_key: str, max_tokens: int = 1200) -> str:
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": max_tokens,
    }

    for attempt in range(5):
        response = requests.post(
            f"{BASE_URL}/chat/completions",
            headers=headers,
            json=payload,
            timeout=60,
        )
        if response.status_code == 429:
            retry = int(response.headers.get("Retry-After", 5))
            print(f"  rate limit, waiting {retry}s")
            time.sleep(retry)
            continue
        if 500 <= response.status_code < 600 and attempt < 4:
            time.sleep(3 + attempt * 2)
            continue

        response.raise_for_status()
        return response.json()["choices"][0]["message"]["content"].strip()

    raise RuntimeError("API request failed after retries")


def normalize_sentiment(value: str) -> str:
    value = value.strip().lower().replace("ё", "е")
    if "негат" in value or "плох" in value or "отриц" in value:
        return "негативный"
    if "нейтр" in value:
        return "нейтральный"
    if "позит" in value or "полож" in value or "хорош" in value:
        return "позитивный"
    return "нейтральный"


def parse_numbered_response(raw: str, expected: int) -> list[str]:
    results: list[str] = []
    for line in raw.splitlines():
        line = line.strip()
        if not line:
            continue
        match = re.search(r"^\s*\d+[\).\s-]+(.+?)\s*$", line)
        value = match.group(1) if match else line
        results.append(normalize_sentiment(value))

    if len(results) < expected:
        results.extend(["нейтральный"] * (expected - len(results)))
    return results[:expected]


def classify_batch(records: list[dict[str, Any]], api_key: str) -> list[str]:
    lines = "\n".join(
        f"{idx + 1}. {str(record['text']).replace(chr(10), ' ')[:600]}"
        for idx, record in enumerate(records)
    )
    prompt = f"""Ты классификатор отзывов покупателей на русском языке.
Для каждого отзыва определи тональность.
Ответь строго одним из трех слов: позитивный / нейтральный / негативный.
Верни только пронумерованный список, без пояснений.

Отзывы:
{lines}

Ответ:"""
    raw = ai_complete(prompt, api_key, max_tokens=max(300, len(records) * 18))
    return parse_numbered_response(raw, len(records))


def find_header_index(headers: list[Any], target: str) -> int | None:
    for index, value in enumerate(headers, start=1):
        if str(value).strip() == target:
            return index
    return None


def collect_reviews(input_path: Path) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    records: list[dict[str, Any]] = []

    for config in SHEET_CONFIGS:
        sheet_name = config["sheet"]
        if sheet_name not in workbook.sheetnames:
            print(f"Sheet not found, skipping: {sheet_name}")
            continue

        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        text_col = find_header_index(headers, config["text_header"])
        if text_col is None:
            print(f"Text column not found, skipping: {sheet_name}")
            continue

        for row_idx, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                max_row=sheet.max_row,
                min_col=text_col,
                max_col=text_col,
                values_only=True,
            ),
            start=2,
        ):
            value = row[0]
            text = "" if value is None else str(value).strip()
            if not text:
                continue
            records.append(
                {
                    "id": f"{sheet_name}|{row_idx}",
                    "sheet": sheet_name,
                    "row": row_idx,
                    "text_header": config["text_header"],
                    "text": text,
                }
            )

    return records


def load_checkpoint(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def run(input_path: Path, checkpoint_path: Path, results_path: Path) -> None:
    api_key = load_api_key()
    records = collect_reviews(input_path)
    checkpoint = load_checkpoint(checkpoint_path)
    total = len(records)
    todo = [record for record in records if record["id"] not in checkpoint]

    print(f"Reviews found: {total}")
    print(f"Already classified: {len(checkpoint)}")
    print(f"Remaining: {len(todo)}")

    for start in range(0, len(todo), BATCH_SIZE):
        batch = todo[start : start + BATCH_SIZE]
        sentiments = classify_batch(batch, api_key)
        for record, sentiment in zip(batch, sentiments):
            checkpoint[record["id"]] = sentiment

        save_json(checkpoint_path, checkpoint)
        done = min(start + len(batch), len(todo))
        print(f"  classified {done}/{len(todo)} new rows")
        time.sleep(0.2)

    results = []
    for record in records:
        result = dict(record)
        result["sentiment"] = checkpoint.get(record["id"], "нейтральный")
        result.pop("text", None)
        results.append(result)

    save_json(results_path, results)
    counts = {
        sentiment: sum(1 for result in results if result["sentiment"] == sentiment)
        for sentiment in sorted(SENTIMENTS)
    }
    print(f"Saved results: {results_path}")
    print(f"Counts: {counts}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Classify buyer review sentiment")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--checkpoint", type=Path, default=DEFAULT_CHECKPOINT)
    parser.add_argument("--results", type=Path, default=DEFAULT_RESULTS)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    run(args.input, args.checkpoint, args.results)


if __name__ == "__main__":
    main()
