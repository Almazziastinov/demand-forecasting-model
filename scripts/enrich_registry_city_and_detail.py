"""Fill missing cities and add detailed complaint categories using VibeCode API."""

from __future__ import annotations

import argparse
import json
import os
import re
import time
from pathlib import Path
from typing import Any

import openpyxl
import requests


BASE_URL = "https://vibecode.bitrix24.tech/v1"
MODEL = "bitrix/bitrixgpt-5.5"
PROJECT_DIR = Path(__file__).resolve().parents[1]
LEGACY_SCRIPT = PROJECT_DIR / "analyze_reviews.py"

INPUT_FILE = Path(
    r"C:\Users\dns\Documents\Codex\2026-06-30\new-chat\outputs\Отзывы покупателей единый реестр аналитика.xlsx"
)
CHECKPOINT_FILE = PROJECT_DIR / "outputs" / "registry_city_detail_checkpoint.json"
OUTPUT_FILE = Path(
    r"C:\Users\dns\Documents\Codex\2026-06-30\new-chat\outputs\Отзывы покупателей единый реестр аналитика города детали.xlsx"
)

DETAIL_CATEGORIES = [
    "Низкое качество еды",
    "Невежливое обслуживание",
    "Несвежие продукты",
    "Проблемы с чистотой",
    "Недостаток ассортимента",
    "Ошибки при расчетах",
    "Наличие вредителей",
    "Плохая упаковка товаров",
    "Другое",
]

BATCH_SIZE = 20


def load_api_key() -> str:
    env_key = os.getenv("VIBECODE_API_KEY")
    if env_key:
        return env_key
    text = LEGACY_SCRIPT.read_text(encoding="utf-8", errors="ignore")
    match = re.search(r'API_KEY\s*=\s*["\']([^"\']+)["\']', text)
    if not match:
        raise RuntimeError("API key was not found")
    return match.group(1)


def ai_complete(prompt: str, api_key: str, max_tokens: int = 2500) -> str:
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    payload = {
        "model": MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": max_tokens,
    }
    attempt = 0
    while True:
        try:
            response = requests.post(
                f"{BASE_URL}/chat/completions",
                headers=headers,
                json=payload,
                timeout=120,
            )
            if response.status_code == 429:
                wait = int(response.headers.get("Retry-After", 20))
                print(f"  rate limit, waiting {wait}s")
                time.sleep(wait)
                continue
            if response.status_code >= 500:
                wait = min(300, 15 + attempt * 15)
                print(f"  server error {response.status_code}, waiting {wait}s")
                time.sleep(wait)
                attempt += 1
                continue
            response.raise_for_status()
            return response.json()["choices"][0]["message"]["content"].strip()
        except requests.RequestException as exc:
            wait = min(300, 15 + attempt * 15)
            print(f"  request error, waiting {wait}s: {exc}")
            time.sleep(wait)
            attempt += 1


def strip_json(raw: str) -> str:
    text = raw.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    start = text.find("[")
    end = text.rfind("]")
    if start >= 0 and end > start:
        return text[start : end + 1]
    return text


def load_checkpoint(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def save_checkpoint(path: Path, data: dict[str, dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def headers_map(sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {str(c.value).strip(): i for i, c in enumerate(sheet[1], start=1) if c.value}


def collect_records(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Единый реестр отзывов"]
    h = headers_map(ws)
    records = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        city = str(row[h["город"] - 1] or "").strip()
        category = str(row[h["категория_нормализованная"] - 1] or "").strip()
        review_type = str(row[h["тип_отзыва"] - 1] or "").strip()
        source_sheet = str(row[h["исходный_лист"] - 1] or "").strip()
        source_row = str(row[h["исходная_строка"] - 1] or "").strip()
        text = str(row[h["текст_отзыва"] - 1] or "").strip()
        address = str(row[h["адрес"] - 1] or "").strip()
        needs_city = city in {"", "Неизвестно"}
        needs_detail = review_type == "Жалоба" and category in {"Сервис", "Качество продукции", "Чистота", "Ассортимент", "Просрочка", "Без категории"}
        if not needs_city and not needs_detail:
            continue
        records.append(
            {
                "id": f"{source_sheet}|{source_row}",
                "row": row_idx,
                "needs_city": needs_city,
                "needs_detail": needs_detail,
                "city": city,
                "category": category,
                "address": address,
                "text": text[:900],
            }
        )
    return records


def normalize_detail(value: str) -> str:
    value = str(value or "").strip()
    return value if value in DETAIL_CATEGORIES else "Другое"


def classify_batch(records: list[dict[str, Any]], api_key: str) -> list[dict[str, str]]:
    categories = "\n".join(f"- {cat}" for cat in DETAIL_CATEGORIES)
    lines = []
    for i, rec in enumerate(records, start=1):
        lines.append(
            f"{i}. id={rec['id']}; текущий_город={rec['city']}; адрес={rec['address']}; "
            f"категория={rec['category']}; текст={rec['text']}"
        )

    prompt = f"""Ты обрабатываешь отзывы покупателей сети пекарен.
Для каждой строки верни JSON-массив объектов в том же порядке.

Поля:
- "город_уточненный": город из адреса или текста. Если город невозможно определить, верни "Неизвестно".
- "детальная_категория": выбери РОВНО одно значение из справочника ниже. Если строка не жалоба или категория не требует детализации, верни "".

Справочник детальных категорий:
{categories}

Правила:
- Город извлекай только если он явно указан в адресе или тексте: Казань, Набережные Челны, Альметьевск и т.д.
- Не выдумывай город по улице.
- Для сервисных жалоб чаще всего подходит "Невежливое обслуживание" или "Ошибки при расчетах".
- Для качества еды выбирай "Низкое качество еды" или "Несвежие продукты".
- Верни только валидный JSON-массив без пояснений.

Строки:
{chr(10).join(lines)}
"""
    raw = ai_complete(prompt, api_key, max_tokens=max(1200, len(records) * 80))
    try:
        data = json.loads(strip_json(raw))
        if not isinstance(data, list):
            raise ValueError("not a list")
    except Exception as exc:
        print(f"  parse failed: {exc}")
        data = []

    result = []
    for i, rec in enumerate(records):
        item = data[i] if i < len(data) and isinstance(data[i], dict) else {}
        city = str(item.get("город_уточненный") or "").strip()
        detail = str(item.get("детальная_категория") or "").strip()
        result.append(
            {
                "город_уточненный": city if city else "Неизвестно",
                "детальная_категория": normalize_detail(detail) if rec["needs_detail"] else "",
            }
        )
    return result


def apply_to_workbook(input_path: Path, output_path: Path, checkpoint: dict[str, dict[str, str]]) -> None:
    wb = openpyxl.load_workbook(input_path)
    ws = wb["Единый реестр отзывов"]
    h = headers_map(ws)

    city_col = h["город"]
    start_col = ws.max_column + 1
    detail_col = start_col
    city_source_col = start_col + 1
    ws.cell(1, detail_col).value = "детальная_категория"
    ws.cell(1, city_source_col).value = "город_источник"
    for col in [detail_col, city_source_col]:
        c = ws.cell(1, col)
        c.fill = openpyxl.styles.PatternFill("solid", fgColor="7030A0")
        c.font = openpyxl.styles.Font(bold=True, color="FFFFFF")

    filled_cities = 0
    detailed = 0
    for row_idx in range(2, ws.max_row + 1):
        source_sheet = str(ws.cell(row_idx, h["исходный_лист"]).value or "").strip()
        source_row = str(ws.cell(row_idx, h["исходная_строка"]).value or "").strip()
        key = f"{source_sheet}|{source_row}"
        item = checkpoint.get(key)
        if not item:
            continue
        current_city = str(ws.cell(row_idx, city_col).value or "").strip()
        city = item.get("город_уточненный", "")
        if current_city in {"", "Неизвестно"} and city and city != "Неизвестно":
            ws.cell(row_idx, city_col).value = city
            ws.cell(row_idx, city_source_col).value = "API"
            filled_cities += 1
        detail = item.get("детальная_категория", "")
        if detail:
            ws.cell(row_idx, detail_col).value = detail
            detailed += 1

    ws.column_dimensions[openpyxl.utils.get_column_letter(detail_col)].width = 28
    ws.column_dimensions[openpyxl.utils.get_column_letter(city_source_col)].width = 16

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"filled_cities={filled_cities}")
    print(f"detailed={detailed}")
    print(f"output={output_path}")


def run(limit: int | None, apply_only: bool) -> None:
    checkpoint = load_checkpoint(CHECKPOINT_FILE)
    if not apply_only:
        api_key = load_api_key()
        records = collect_records(INPUT_FILE)
        if limit is not None:
            records = records[:limit]
        todo = [rec for rec in records if rec["id"] not in checkpoint]
        print(f"records={len(records)} already={len(checkpoint)} remaining={len(todo)}")
        for start in range(0, len(todo), BATCH_SIZE):
            batch = todo[start : start + BATCH_SIZE]
            result = classify_batch(batch, api_key)
            for rec, item in zip(batch, result):
                checkpoint[rec["id"]] = item
            save_checkpoint(CHECKPOINT_FILE, checkpoint)
            print(f"  processed {min(start + len(batch), len(todo))}/{len(todo)}")
            time.sleep(0.2)
    apply_to_workbook(INPUT_FILE, OUTPUT_FILE, checkpoint)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--apply-only", action="store_true")
    args = parser.parse_args()
    run(args.limit, args.apply_only)


if __name__ == "__main__":
    main()
