"""Build a unified buyer reviews workbook from all source review sheets."""

from __future__ import annotations

from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_FILE = Path(r"C:\Users\dns\Downloads\Отзывы покупателей.xlsx")
OUTPUT_FILE = Path(
    r"C:\Users\dns\Documents\Codex\2026-06-30\new-chat\outputs\Отзывы покупателей единый реестр.xlsx"
)

SOURCES = [
    {
        "sheet": "2024-25 Жалобы",
        "date": "Дата",
        "source": "Жалоба поступила",
        "category": "Тема сообщения",
        "text": "Суть жалобы",
        "city": "Город",
        "street": "Улица, дом",
        "address": "Адрес",
        "rd": "РД",
        "manager": "Управляющий ",
        "network": "Франшиза/Сеть",
    },
    {
        "sheet": "Отзывы 2025",
        "date": "Дата",
        "source": "Источник",
        "category": "Категория",
        "text": "Отзыв",
        "city": "Город",
        "address": "Адрес",
        "rd": "РД",
        "manager": "Управляющий",
        "network": "Наименование",
    },
    {
        "sheet": "2022 Полож отзывы",
        "date": "Дата",
        "source": "Жалоба поступила",
        "category": "Тема сообщения",
        "text": "Суть жалобы",
        "city": "Город",
        "street": "Улица, дом",
        "address": "Адрес",
        "rd": "РД",
        "manager": "Управляющий ",
        "network": "Франшиза/Сеть",
    },
]

OUTPUT_HEADERS = [
    "дата",
    "год",
    "месяц",
    "тип_отзыва",
    "категория_исходная",
    "категория_нормализованная",
    "текст_отзыва",
    "город",
    "адрес",
    "РД",
    "управляющий",
    "источник",
    "формат_точки",
    "исходный_лист",
    "исходная_строка",
]


def normalize_category(value: Any) -> str:
    raw = "" if value is None else str(value).strip()
    if not raw:
        return ""

    key = raw.lower().replace("ё", "е")
    key = " ".join(key.split())
    dictionary = {
        "благодарность": "Благодарность",
        "благодарсность": "Благодарность",
        "сервис": "Сервис",
        "качество продукции": "Качество продукции",
        "качество": "Качество продукции",
        "качество продукта": "Качество продукции",
        "качество сырья": "Качество продукции",
        "качкство": "Качество продукции",
        "состав": "Качество продукции",
        "вкус": "Качество продукции",
        "упаковка": "Качество продукции",
        "хот доги": "Качество продукции",
        "кбжу": "Качество продукции",
        "чистота": "Чистота",
        "просрок": "Просрочка",
        "просрочка": "Просрочка",
        "звонок на горячую линию": "Звонок на горячую линию",
        "отзыв бывшего сотрудника": "Отзыв сотрудника",
        "отзыв сотрудника": "Отзыв сотрудника",
        "жалоба персонала": "Жалоба персонала",
        "помещение": "Помещение",
        "ассортимент": "Ассортимент",
    }
    return dictionary.get(key, raw)


def infer_review_type(normalized_category: str) -> str:
    if normalized_category == "Благодарность":
        return "Благодарность"
    return "Жалоба"


def headers_map(sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): index
        for index, cell in enumerate(sheet[1], start=1)
        if cell.value is not None
    }


def get_value(row: tuple[Any, ...], header_indexes: dict[str, int], header: str | None) -> Any:
    if not header:
        return None
    index = header_indexes.get(header)
    if index is None or index > len(row):
        return None
    return row[index - 1]


def as_date(value: Any) -> Any:
    if isinstance(value, datetime):
        return value
    return value


def build_address(row: tuple[Any, ...], header_indexes: dict[str, int], source: dict[str, str]) -> str:
    address = get_value(row, header_indexes, source.get("address"))
    if address not in (None, "", 0, "0"):
        return str(address).strip()

    city = get_value(row, header_indexes, source.get("city"))
    street = get_value(row, header_indexes, source.get("street"))
    parts = [str(part).strip() for part in [city, street] if part not in (None, "", 0, "0")]
    return ", ".join(parts)


def collect_rows(input_file: Path) -> list[list[Any]]:
    workbook = openpyxl.load_workbook(input_file, read_only=True, data_only=True)
    output_rows: list[list[Any]] = []

    for source in SOURCES:
        sheet = workbook[source["sheet"]]
        header_indexes = headers_map(sheet)
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            text = get_value(row, header_indexes, source["text"])
            text = "" if text is None else str(text).strip()
            if not text:
                continue

            date_value = as_date(get_value(row, header_indexes, source["date"]))
            year = date_value.year if isinstance(date_value, datetime) else ""
            month = date_value.month if isinstance(date_value, datetime) else get_value(row, header_indexes, "Месяца")
            source_category = get_value(row, header_indexes, source["category"])
            normalized_category = normalize_category(source_category)

            output_rows.append(
                [
                    date_value,
                    year,
                    month,
                    infer_review_type(normalized_category),
                    "" if source_category is None else str(source_category).strip(),
                    normalized_category,
                    text,
                    "" if get_value(row, header_indexes, source.get("city")) is None else str(get_value(row, header_indexes, source.get("city"))).strip(),
                    build_address(row, header_indexes, source),
                    "" if get_value(row, header_indexes, source.get("rd")) is None else str(get_value(row, header_indexes, source.get("rd"))).strip(),
                    "" if get_value(row, header_indexes, source.get("manager")) is None else str(get_value(row, header_indexes, source.get("manager"))).strip(),
                    "" if get_value(row, header_indexes, source.get("source")) is None else str(get_value(row, header_indexes, source.get("source"))).strip(),
                    "" if get_value(row, header_indexes, source.get("network")) is None else str(get_value(row, header_indexes, source.get("network"))).strip(),
                    source["sheet"],
                    row_number,
                ]
            )

    output_rows.sort(
        key=lambda item: (
            item[0] if isinstance(item[0], datetime) else datetime.max,
            str(item[13]),
            item[14],
        )
    )
    return output_rows


def write_summary(workbook: Workbook, rows: list[list[Any]]) -> None:
    sheet = workbook.create_sheet("Сводка")
    total = len(rows)
    by_year = Counter(row[1] for row in rows if row[1] != "")
    by_type = Counter(row[3] for row in rows)
    by_category = Counter(row[5] for row in rows if row[5])

    data = [["Показатель", "Значение"], ["Всего отзывов", total]]
    if rows:
        dates = [row[0] for row in rows if isinstance(row[0], datetime)]
        data.extend(
            [
                ["Минимальная дата", min(dates) if dates else ""],
                ["Максимальная дата", max(dates) if dates else ""],
                ["", ""],
                ["По типу", ""],
            ]
        )
    for key, value in by_type.most_common():
        data.append([key, value])
    data.append(["", ""])
    data.append(["По годам", ""])
    for key, value in sorted(by_year.items()):
        data.append([key, value])
    data.append(["", ""])
    data.append(["Топ категорий", ""])
    for key, value in by_category.most_common(20):
        data.append([key, value])

    for row in data:
        sheet.append(row)

    sheet["A1"].font = Font(bold=True, color="FFFFFF")
    sheet["B1"].font = Font(bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet["B1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 18
    for cell in sheet["B"]:
        if isinstance(cell.value, datetime):
            cell.number_format = "yyyy-mm-dd"


def write_workbook(rows: list[list[Any]], output_file: Path) -> None:
    workbook = Workbook()
    registry = workbook.active
    registry.title = "Единый реестр отзывов"
    registry.append(OUTPUT_HEADERS)

    for row in rows:
        registry.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in registry[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [14, 10, 10, 16, 24, 28, 80, 22, 32, 18, 20, 24, 16, 22, 14]
    for index, width in enumerate(widths, start=1):
        registry.column_dimensions[get_column_letter(index)].width = width

    registry.freeze_panes = "A2"
    registry.auto_filter.ref = registry.dimensions
    registry.column_dimensions["G"].width = 90
    for cell in registry["A"]:
        if isinstance(cell.value, datetime):
            cell.number_format = "yyyy-mm-dd"
    for row in registry.iter_rows(min_row=2):
        row[6].alignment = Alignment(wrap_text=True, vertical="top")

    write_summary(workbook, rows)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)


def main() -> None:
    rows = collect_rows(INPUT_FILE)
    write_workbook(rows, OUTPUT_FILE)
    print(f"rows={len(rows)}")
    print(f"output={OUTPUT_FILE}")


if __name__ == "__main__":
    main()
