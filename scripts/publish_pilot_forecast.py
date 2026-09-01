"""Generate and publish the daily pilot forecast summary to Bitrix24 chat.

For the given date (default: today), queries ClickHouse for per-SKU
daily forecasts and previous-day closing stock for all pilot bakeries.
The stock is subtracted from forecast demand before kratnost rounding.
If VIBECODE_API_KEY is set (via .env or environment), uploads the generated
Excel file to the Bitrix24 pilot chat automatically.

Usage (local / on VM):
    python scripts/publish_pilot_forecast.py --env-file .env
    python scripts/publish_pilot_forecast.py --env-file .env --date 2026-07-24
    python scripts/publish_pilot_forecast.py --env-file .env --dry-run
"""

from __future__ import annotations

import argparse
import math
import os
import sys
from datetime import date as date_type
from datetime import timedelta
from io import BytesIO
from pathlib import Path

import pandas as pd
import numpy as np

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "apps"))
sys.path.insert(0, str(ROOT / "apps" / "forecast_embedded"))

from src.experiments_v2.sku_systematic_correction import (  # noqa: E402
    CorrectionConfig,
    apply_category_neutral_corrections,
    build_correction_registry,
)
from src.experiments_v2.sku_cold_start import (  # noqa: E402
    apply_independent_cold_start,
    build_cold_start_registry,
)
from src.pilot_scope import BASE_PILOT_10  # noqa: E402

# Compatibility/exported analytical scope. Runtime publishing uses the dynamic
# pilot_scope_events membership loaded by _load_pilot_bakery_ids().
PILOT_BAKERY_IDS = sorted(BASE_PILOT_10.bakery_ids)

# Управляемый scope в ClickHouse.  Если таблица пуста — используется fallback.
_PILOT_SCOPE_NAME = "expanded_pilot_38"

# Seed-список: отражает состав пилота на момент запуска (38 пекарен).
# Используется только как fallback, если в pilot_scope_events нет ни одной записи.
_SEED_PILOT_IDS: list[int] = [
    1, 3, 12, 13, 14, 20, 21, 22, 26, 27, 28, 39, 41, 56, 57, 66, 67, 69,
    80, 89, 99, 107, 113, 125, 149, 153, 155, 160, 191, 221, 222, 229, 230,
    246, 257, 260, 268, 270,
]

WEEKDAY_RU = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]

BAKEABLE_CATEGORIES = {
    "Пироги сытные",
    "Пироги сладкие",
    "Выпечка сытная",
    "Выпечка сладкая",
    "Фастфуд",
}

PRODUCT_NAME_OVERRIDES = {
    11615: "Плетенка кленовая",
    11616: "Плетенка с черникой",
    11617: "Плетенка с земляникой",
}

MISSING_KRATNOST_LABEL = "нет данных по кратности"


def _load_env(env_file: str) -> None:
    with open(env_file) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def _round_up_kratnost(value: float, kratnost: int) -> int:
    """Always round UP to the nearest kratnost multiple (same as MILP)."""
    if value <= 0 or kratnost <= 0:
        return 0
    return int(math.ceil(value / kratnost - 1e-9) * kratnost)


def _production_plan_with_optional_kratnost(
    net_need: float,
    kratnost: int | None,
) -> tuple[int, int | str]:
    """Keep an SKU in the plan when its production multiple is unknown."""
    if kratnost is None:
        return max(0, int(math.ceil(net_need - 1e-9))), MISSING_KRATNOST_LABEL
    return _round_up_kratnost(net_need, kratnost), kratnost


def _enrich_forecast_product_metadata(
    forecast: pd.DataFrame,
    product_meta: pd.DataFrame,
) -> pd.DataFrame:
    """Fill missing forecast names/categories from the product dimension."""
    if forecast.empty or product_meta.empty:
        return forecast.copy()
    meta = product_meta[["product_id", "product_name", "category_name"]].copy()
    meta = meta.drop_duplicates("product_id").rename(
        columns={
            "product_name": "dimension_product_name",
            "category_name": "dimension_category_name",
        }
    )
    enriched = forecast.merge(meta, on="product_id", how="left", validate="many_to_one")
    for column in ("product_name", "category_name"):
        dimension_column = f"dimension_{column}"
        current = enriched[column].astype("string").str.strip()
        missing = current.isna() | current.eq("")
        enriched.loc[missing, column] = enriched.loc[missing, dimension_column]
        enriched = enriched.drop(columns=dimension_column)
    return enriched


def _add_missing_cold_start_candidates(
    forecast: pd.DataFrame,
    candidates: pd.DataFrame,
) -> pd.DataFrame:
    """Add zero-forecast candidate rows before applying cold-start floors."""
    keys = ["bakery_id", "product_id"]
    existing = set(map(tuple, forecast[keys].values.tolist()))
    additions = candidates[
        ~candidates.apply(
            lambda row: (row["bakery_id"], row["product_id"]) in existing,
            axis=1,
        )
    ].copy()
    if additions.empty:
        return forecast.copy()
    additions["forecast_qty"] = 0.0
    if "date" in forecast.columns:
        additions["date"] = forecast["date"].iloc[0] if not forecast.empty else pd.NaT
    additions = additions.reindex(columns=forecast.columns, fill_value=pd.NA)
    return pd.concat([forecast, additions], ignore_index=True)


def _load_pilot_bakery_ids(client) -> list[int]:
    """Загружает актуальный список пилотных пекарен из pilot_scope_events.

    Последнее событие на пекарню определяет её статус (add/exclude).
    Пекарни без событий берутся из seed-списка (действовали до введения лога).
    При любой ошибке (таблица не создана, ClickHouse недоступен) — seed.
    """
    try:
        df = client.query_df(
            """
            SELECT bakery_id, argMax(action, changed_at) AS last_action
            FROM Svezhar.pilot_scope_events
            WHERE scope_name = %(scope_name)s
            GROUP BY bakery_id
            """,
            parameters={"scope_name": _PILOT_SCOPE_NAME},
        )
    except Exception as exc:
        print(f"  WARNING: не удалось загрузить pilot_scope_events, используем seed: {exc}")
        return list(_SEED_PILOT_IDS)

    if df.empty:
        print("  INFO: pilot_scope_events пуст, используем seed-список пекарен")
        return list(_SEED_PILOT_IDS)

    event_active: set[int] = set()
    event_all: set[int] = set()
    for row in df.to_dict("records"):
        try:
            bid = int(row["bakery_id"])
        except (TypeError, ValueError):
            continue
        event_all.add(bid)
        if row["last_action"] == "add":
            event_active.add(bid)

    seed_active = {bid for bid in _SEED_PILOT_IDS if bid not in event_all}
    result = sorted(event_active | seed_active)
    print(f"  INFO: пилот — {len(result)} пекарен ({len(event_active)} из ClickHouse, {len(seed_active)} из seed)")
    return result


def _build_report(
    forecast_date: str,
    *,
    sku_correction_registry: str | Path | None = None,
    enable_new_sku_cold_start: bool = True,
    forecast_override: str | Path | None = None,
) -> list[dict]:
    """Build daily pilot rows with previous-day stock and production plan."""
    from app.db import get_client
    from app.table_names import table_name

    client = get_client()
    pilot_bakery_ids = _load_pilot_bakery_ids(client)

    run_df = client.query_df(
        f"select run_id, model_version from "
        f"{table_name('forecast_runs_embedded')} "
        "where status = 'active' limit 1"
    )
    if run_df.empty:
        raise RuntimeError("No active forecast run")
    run_id = str(run_df.iloc[0]["run_id"])
    active_model_version = str(run_df.iloc[0].get("model_version") or "")
    direct_handles_cold_start = (
        active_model_version == "direct_alpha_025_v1"
        and forecast_override is None
    )

    # Bakery names for pilot bakeries
    # dim_bakeries uses zero-padded string bakery_id ("000000016"), same as baking_sku_meta
    pilot_bids_str = [f"{b:09d}" for b in pilot_bakery_ids]
    bakery_df = client.query_df(
        "select bakery_id as bid, any(bakery_name) as name, any(city) as city "
        "from dim_bakeries "
        "where bakery_id in %(bids)s "
        "group by bakery_id",
        parameters={"bids": pilot_bids_str},
    )
    bakery_info: dict[int, dict] = {}
    for row in bakery_df.to_dict("records"):
        try:
            bid = int(row["bid"])
            bakery_info[bid] = {"name": row["name"], "city": row["city"]}
        except (TypeError, ValueError):
            pass

    # Daily SKU forecasts for all pilot bakeries
    forecast_df = client.query_df(
        f"""
        select
            bakery_id,
            product_id,
            any(product_name) as product_name,
            any(category_name) as category_name,
            sum(forecast_qty) as forecast_qty
        from {table_name('sku_forecast_day_embedded')}
        where run_id = %(run_id)s
          and forecast_date = %(forecast_date)s
          and bakery_id in %(bids)s
        group by bakery_id, product_id
        """,
        parameters={
            "run_id": run_id,
            "forecast_date": forecast_date,
            "bids": pilot_bakery_ids,
        },
    )

    if forecast_override is not None:
        override_path = Path(forecast_override)
        forecast_df = pd.read_csv(override_path)
        required = {
            "bakery_id",
            "product_id",
            "product_name",
            "category_name",
            "forecast_qty",
        }
        missing = sorted(required.difference(forecast_df.columns))
        if missing:
            raise ValueError(f"Forecast override is missing columns: {missing}")
        forecast_df = forecast_df[list(required)].copy()
        forecast_df["bakery_id"] = pd.to_numeric(forecast_df["bakery_id"], errors="raise").astype("int64")
        forecast_df["product_id"] = pd.to_numeric(forecast_df["product_id"], errors="raise").astype("int64")
        forecast_df["forecast_qty"] = pd.to_numeric(forecast_df["forecast_qty"], errors="raise").clip(lower=0.0)
        forecast_df = forecast_df[forecast_df["bakery_id"].isin(pilot_bakery_ids)]

    if forecast_df.empty:
        print(f"  WARNING: no forecast rows for {forecast_date}, run {run_id}")
        return []

    forecast_df["product_id"] = pd.to_numeric(
        forecast_df["product_id"], errors="coerce"
    ).astype("Int64")
    dimension_df = client.query_df(
        """
        select
            toInt64OrZero(toString(product_id)) as product_id,
            any(product_name) as product_name,
            any(category_name) as category_name
        from Svezhar.dim_products
        where toInt64OrZero(toString(product_id)) in %(product_ids)s
        group by product_id
        """,
        parameters={
            "product_ids": [
                int(product_id)
                for product_id in forecast_df["product_id"].dropna().unique()
            ]
        },
    )
    forecast_df = _enrich_forecast_product_metadata(forecast_df, dimension_df)

    previous_date = str(date_type.fromisoformat(forecast_date) - timedelta(days=1))
    # Stock: qty_produced - qty_sold for previous day, computed from fct tables.
    # fct_check_lines dedup: DISTINCT on business keys (matches nightly pipeline).
    # fct_production_release dedup: GROUP BY (release_id, line_id) only, argMax all
    # mutable fields — prevents double-counting when ETL rewrites product_id between
    # versions (observed 2026-08-14: ~689 lines had product_id null→correct, causing
    # old GROUP BY on (release_id,line_id,bakery_id,product_id) to count each line twice).
    _fct_sold_yesterday = client.query_df(
        """
        select
            toInt64OrZero(toString(bakery_id)) as bakery_id,
            toInt64OrZero(toString(product_id)) as product_id,
            sum(toFloat64(quantity)) as qty_sold
        from (
            select distinct
                check_datetime, check_date, bakery_id, product_id,
                quantity, price, line_amount, cash_event_type
            from Svezhar.fct_check_lines
            where hex(cash_event_type) = 'D09FD180D0BED0B4D0B0D0B6D0B0'
              and check_date = toDate(%(previous_date)s)
              and toInt64OrZero(toString(bakery_id)) in %(bids)s
        )
        group by bakery_id, product_id
        """,
        parameters={"previous_date": previous_date, "bids": pilot_bakery_ids},
    )
    _fct_produced_yesterday = client.query_df(
        """
        select
            toInt64OrZero(toString(bid)) as bakery_id,
            toInt64OrZero(toString(pid)) as product_id,
            sum(qty) as qty_produced
        from (
            select
                argMax(bakery_id, _updated_at) as bid,
                argMax(product_id, _updated_at) as pid,
                toFloat64(argMax(quantity, _updated_at)) as qty
            from Svezhar.fct_production_release
            where toDate(release_date) = toDate(%(previous_date)s)
              and toInt64OrZero(toString(bakery_id)) in %(bids)s
            group by release_id, line_id
            having argMax(is_deleted, _updated_at) not in ('1', 'true', 'Да')
        )
        group by bakery_id, product_id
        """,
        parameters={"previous_date": previous_date, "bids": pilot_bakery_ids},
    )
    _stock_cols = ["bakery_id", "product_id"]
    if _fct_sold_yesterday.empty or "bakery_id" not in _fct_sold_yesterday.columns:
        _fct_sold_yesterday = pd.DataFrame(columns=_stock_cols + ["qty_sold"])
    if _fct_produced_yesterday.empty or "bakery_id" not in _fct_produced_yesterday.columns:
        _fct_produced_yesterday = pd.DataFrame(columns=_stock_cols + ["qty_produced"])
    _stock_merged = _fct_produced_yesterday.merge(_fct_sold_yesterday, on=_stock_cols, how="outer")
    _stock_merged["qty_produced"] = pd.to_numeric(_stock_merged.get("qty_produced", 0), errors="coerce").fillna(0.0)
    _stock_merged["qty_sold"] = pd.to_numeric(_stock_merged.get("qty_sold", 0), errors="coerce").fillna(0.0)
    _stock_merged["stock_balance"] = (_stock_merged["qty_produced"] - _stock_merged["qty_sold"]).clip(lower=0.0)
    stock_df = _stock_merged[_stock_merged["stock_balance"] > 0][_stock_cols + ["stock_balance"]]
    yesterday_stock: dict[tuple[int, int], float] = {}
    for row in stock_df.to_dict("records"):
        try:
            key = (int(row["bakery_id"]), int(row["product_id"]))
            yesterday_stock[key] = max(float(row.get("stock_balance") or 0), 0.0)
        except (TypeError, ValueError):
            continue

    # SKU meta (kratnost) — base + bakery overrides
    # baking_sku_meta.product_id is zero-padded string ("000001234")
    # sku_forecast_day_embedded.product_id is int64
    all_pids_int = [int(r) for r in forecast_df["product_id"].dropna().unique()]
    all_pids_str = [f"{p:09d}" for p in all_pids_int]
    meta_df = client.query_df(
        f"""
        select product_id, bakery_id, dough_group, kratnost, scope
        from {table_name('baking_sku_meta')} final
        where is_active = 1 and product_id in %(pids)s
        """,
        parameters={"pids": all_pids_str},
    )

    # Build kratnost lookup keyed by int product_id
    base_kratnost: dict[int, int] = {}
    bakery_kratnost: dict[tuple[int, int], int] = {}
    frozen_pids: set[int] = set()
    for row in meta_df.to_dict("records"):
        try:
            pid_int = int(row["product_id"])
        except (TypeError, ValueError):
            continue
        dg = str(row.get("dough_group") or "").lower()
        if "замороженные полуфабрикаты" in dg:
            frozen_pids.add(pid_int)
            continue
        kr = int(row.get("kratnost") or 1) or 1
        if row["scope"] == "bakery" and row["bakery_id"] is not None:
            try:
                bakery_kratnost[(pid_int, int(row["bakery_id"]))] = kr
            except (TypeError, ValueError):
                pass
        else:
            base_kratnost[pid_int] = kr

    corrected_forecast: dict[tuple[int, int], float] = {}
    if sku_correction_registry or enable_new_sku_cold_start:
        eligible = forecast_df.copy()
        eligible["bakery_id"] = pd.to_numeric(
            eligible["bakery_id"],
            errors="coerce",
        )
        eligible["product_id"] = pd.to_numeric(
            eligible["product_id"],
            errors="coerce",
        )
        eligible = eligible[
            eligible["bakery_id"].isin(pilot_bakery_ids)
            & eligible["category_name"].isin(BAKEABLE_CATEGORIES)
            & ~eligible["product_id"].isin(frozen_pids)
        ].copy()
        eligible["date"] = pd.Timestamp(forecast_date)

    if enable_new_sku_cold_start and not direct_handles_cold_start:
        history_from = str(
            date_type.fromisoformat(forecast_date) - timedelta(days=60)
        )
        sales_history = client.query_df(
            """
            select
                check_date as date,
                toInt64OrZero(toString(bakery_id)) as bakery_id,
                toInt64OrZero(toString(product_id)) as product_id,
                sum(toFloat64(quantity)) as sold_qty
            from (
                select distinct
                    check_datetime, check_date, bakery_id, product_id,
                    quantity, price, line_amount, cash_event_type
                from Svezhar.fct_check_lines
                where hex(cash_event_type) = 'D09FD180D0BED0B4D0B0D0B6D0B0'
                  and check_date >= toDate(%(history_from)s)
                  and check_date < toDate(%(forecast_date)s)
                  and toInt64OrZero(toString(bakery_id)) in %(bids)s
            )
            group by date, bakery_id, product_id
            """,
            parameters={
                "history_from": history_from,
                "forecast_date": forecast_date,
                "bids": pilot_bakery_ids,
            },
        )
        forecast_history = client.query_df(
            """
            select
                forecast_date as date,
                toInt64(bakery_id) as bakery_id,
                toInt64(product_id) as product_id,
                argMax(forecast_qty, generated_at) as forecast_qty
            from Svezhar.sku_forecast_day_snapshots
            where forecast_date >= toDate(%(history_from)s)
              and forecast_date < toDate(%(forecast_date)s)
              and lead_days = 1
              and toInt64(bakery_id) in %(bids)s
            group by date, bakery_id, product_id
            """,
            parameters={
                "history_from": history_from,
                "forecast_date": forecast_date,
                "bids": pilot_bakery_ids,
            },
        )
        cold_history = sales_history.merge(
            forecast_history,
            on=["date", "bakery_id", "product_id"],
            how="left",
            validate="one_to_one",
        )
        cold_history["forecast_qty"] = cold_history["forecast_qty"].fillna(0.0)
        cold_registry = build_cold_start_registry(
            cold_history,
            as_of_date=forecast_date,
        )
        if not cold_registry.empty:
            cold_ids = sorted(cold_registry["product_id"].astype(int).unique())
            cold_ids_padded = [f"{product_id:09d}" for product_id in cold_ids]
            cold_meta = client.query_df(
                f"""
                select product_id, bakery_id, dough_group, kratnost, scope
                from {table_name('baking_sku_meta')} final
                where is_active = 1 and product_id in %(pids)s
                """,
                parameters={"pids": cold_ids_padded},
            )
            for row in cold_meta.to_dict("records"):
                product_id = int(row["product_id"])
                if "замороженные полуфабрикаты" in str(
                    row.get("dough_group") or ""
                ).lower():
                    frozen_pids.add(product_id)
                    continue
                kratnost = int(row.get("kratnost") or 1) or 1
                if row["scope"] == "bakery" and row["bakery_id"] is not None:
                    bakery_kratnost[(product_id, int(row["bakery_id"]))] = kratnost
                else:
                    base_kratnost[product_id] = kratnost
            product_meta = client.query_df(
                """
                select
                    toInt64OrZero(toString(product_id)) as product_id,
                    any(product_name) as product_name,
                    any(category_name) as category_name
                from Svezhar.dim_products
                where toInt64OrZero(toString(product_id)) in %(product_ids)s
                group by product_id
                """,
                parameters={"product_ids": cold_ids},
            )
            candidates = cold_registry[["bakery_id", "product_id"]].merge(
                product_meta,
                on="product_id",
                how="inner",
                validate="many_to_one",
            )
            candidates = candidates[
                candidates["category_name"].isin(BAKEABLE_CATEGORIES)
                & ~candidates["product_id"].isin(frozen_pids)
            ]
            candidates = candidates[
                candidates.apply(
                    lambda row: int(row["product_id"]) in base_kratnost
                    or (int(row["product_id"]), int(row["bakery_id"]))
                    in bakery_kratnost,
                    axis=1,
                )
            ]
            eligible = _add_missing_cold_start_candidates(eligible, candidates)
        eligible = apply_independent_cold_start(
            eligible,
            cold_registry,
        )
        eligible["forecast_qty"] = eligible["independent_forecast_qty"]
        eligible = eligible.drop(columns="independent_forecast_qty")
        print(
            "  new-SKU cold start: "
            f"{len(cold_registry)} bakery/SKU floors, "
            "independent volume added above mature allocation"
        )

    mature_registry = pd.DataFrame()
    if sku_correction_registry:
        registry_path = Path(sku_correction_registry)
        if not registry_path.exists():
            raise FileNotFoundError(
                f"SKU correction registry not found: {registry_path}"
            )
        mature_registry = pd.read_csv(registry_path, encoding="utf-8-sig")
    else:
        mature_history_from = str(
            date_type.fromisoformat(forecast_date)
            - timedelta(days=CorrectionConfig().history_days)
        )
        _mature_sold = client.query_df(
            """
            select
                check_date as date,
                toInt64OrZero(toString(bakery_id)) as bakery_id,
                toInt64OrZero(toString(product_id)) as product_id,
                sum(toFloat64(quantity)) as sold_qty,
                max(check_datetime) as last_sale_time
            from (
                select distinct
                    check_datetime, check_date, bakery_id, product_id,
                    quantity, price, line_amount, cash_event_type
                from Svezhar.fct_check_lines
                where hex(cash_event_type) = 'D09FD180D0BED0B4D0B0D0B6D0B0'
                  and check_date >= toDate(%(history_from)s)
                  and check_date < toDate(%(forecast_date)s)
                  and toInt64OrZero(toString(bakery_id)) in %(bids)s
            )
            group by date, bakery_id, product_id
            """,
            parameters={
                "history_from": mature_history_from,
                "forecast_date": forecast_date,
                "bids": pilot_bakery_ids,
            },
        )
        _mature_produced = client.query_df(
            """
            select
                date,
                toInt64OrZero(toString(bid)) as bakery_id,
                toInt64OrZero(toString(pid)) as product_id,
                sum(qty) as produced_qty
            from (
                select
                    toDate(argMax(release_date, _updated_at)) as date,
                    argMax(bakery_id, _updated_at) as bid,
                    argMax(product_id, _updated_at) as pid,
                    toFloat64(argMax(quantity, _updated_at)) as qty
                from Svezhar.fct_production_release
                where toDate(release_date) >= toDate(%(history_from)s)
                  and toDate(release_date) < toDate(%(forecast_date)s)
                  and toInt64OrZero(toString(bakery_id)) in %(bids)s
                group by release_id, line_id
                having argMax(is_deleted, _updated_at) not in ('1', 'true', 'Да')
            )
            group by date, bakery_id, product_id
            """,
            parameters={
                "history_from": mature_history_from,
                "forecast_date": forecast_date,
                "bids": pilot_bakery_ids,
            },
        )
        if _mature_sold.empty or "bakery_id" not in _mature_sold.columns:
            _mature_sold = pd.DataFrame(columns=["date", "bakery_id", "product_id", "sold_qty", "last_sale_time"])
        if _mature_produced.empty or "bakery_id" not in _mature_produced.columns:
            _mature_produced = pd.DataFrame(columns=["date", "bakery_id", "product_id", "produced_qty"])
        mature_fact = _mature_sold.merge(
            _mature_produced, on=["date", "bakery_id", "product_id"], how="outer"
        )
        mature_fact["sold_qty"] = pd.to_numeric(mature_fact.get("sold_qty", 0), errors="coerce").fillna(0.0)
        mature_fact["produced_qty"] = pd.to_numeric(mature_fact.get("produced_qty", 0), errors="coerce").fillna(0.0)
        # Join product_name and category_name from forecast_df
        _product_meta = (
            forecast_df[["product_id", "product_name", "category_name"]]
            .drop_duplicates("product_id")
            .assign(product_id=lambda d: pd.to_numeric(d["product_id"], errors="coerce").astype("Int64"))
        )
        mature_fact["product_id"] = pd.to_numeric(mature_fact["product_id"], errors="coerce").astype("Int64")
        mature_fact = mature_fact.merge(_product_meta, on="product_id", how="left")
        mature_fact["product_name"] = mature_fact.get("product_name", "").fillna("")
        mature_fact["category_name"] = mature_fact.get("category_name", "").fillna("")
        mature_forecast = client.query_df(
            """
            select
                forecast_date as date,
                toInt64(bakery_id) as bakery_id,
                toInt64(product_id) as product_id,
                argMax(forecast_qty, generated_at) as forecast_qty
            from Svezhar.sku_forecast_day_snapshots
            where forecast_date >= toDate(%(history_from)s)
              and forecast_date < toDate(%(forecast_date)s)
              and lead_days = 1
              and toInt64(bakery_id) in %(bids)s
            group by date, bakery_id, product_id
            """,
            parameters={
                "history_from": mature_history_from,
                "forecast_date": forecast_date,
                "bids": pilot_bakery_ids,
            },
        )
        mature_history = mature_fact.merge(
            mature_forecast,
            on=["date", "bakery_id", "product_id"],
            how="left",
            validate="one_to_one",
        )
        mature_history["forecast_qty"] = mature_history[
            "forecast_qty"
        ].fillna(0.0)
        last_sale = pd.to_datetime(
            mature_history["last_sale_time"],
            errors="coerce",
        )
        day_start = last_sale.dt.normalize() + pd.Timedelta(hours=7)
        day_end = last_sale.dt.normalize() + pd.Timedelta(hours=19)
        elapsed_hours = (
            (last_sale - day_start).dt.total_seconds() / 3600.0
        )
        remaining_hours = (
            (day_end - last_sale).dt.total_seconds() / 3600.0
        ).clip(lower=0.0)
        full_realization = (
            mature_history["produced_qty"].gt(0)
            & mature_history["sold_qty"].ge(
                mature_history["produced_qty"] - 0.01
            )
            & elapsed_hours.ge(2.0)
            & last_sale.lt(day_end)
        )
        raw_lost = np.where(
            full_realization,
            mature_history["sold_qty"] / elapsed_hours * remaining_hours,
            0.0,
        )
        conservative_cap = np.maximum(
            mature_history["sold_qty"] * 1.5,
            15.0,
        )
        mature_history["lost_demand_qty"] = np.minimum(
            raw_lost,
            conservative_cap,
        )
        mature_history["demand_qty"] = (
            mature_history["sold_qty"]
            + mature_history["lost_demand_qty"]
        )
        mature_registry = build_correction_registry(
            mature_history,
            as_of_date=forecast_date,
        )

    if not mature_registry.empty:
        cold_mask = eligible.get(
            "is_cold_start",
            pd.Series(False, index=eligible.index),
        )
        mature_eligible = eligible[~cold_mask].copy()
        corrected = apply_category_neutral_corrections(
            mature_eligible,
            mature_registry,
        )
        corrected_forecast = {
            (int(row["bakery_id"]), int(row["product_id"])): float(
                row["forecast_qty"]
            )
            for row in eligible.to_dict("records")
        }
        corrected_forecast.update({
            (int(row["bakery_id"]), int(row["product_id"])): float(
                row["corrected_forecast_qty"]
            )
            for row in corrected.to_dict("records")
        })
        changed = (
            corrected["corrected_forecast_qty"]
            - corrected["forecast_qty"]
        ).abs() > 1e-9
        print(
            "  systematic SKU correction: "
            f"{int(changed.sum())} changed rows, "
            "bakery/category totals preserved"
        )
    elif enable_new_sku_cold_start:
        corrected_forecast = {
            (int(row["bakery_id"]), int(row["product_id"])): float(
                row["forecast_qty"]
            )
            for row in eligible.to_dict("records")
        }

    report_df = (
        eligible
        if sku_correction_registry or enable_new_sku_cold_start
        else forecast_df
    )
    rows = []
    for rec in report_df.to_dict("records"):
        try:
            bid = int(rec["bakery_id"])
        except (TypeError, ValueError):
            continue
        try:
            pid_int = int(rec["product_id"])
        except (TypeError, ValueError):
            continue
        if bid not in pilot_bakery_ids:
            continue
        category = str(rec.get("category_name") or "")
        if category not in BAKEABLE_CATEGORIES:
            continue
        if pid_int in frozen_pids:
            continue
        kratnost = bakery_kratnost.get((pid_int, bid)) or base_kratnost.get(pid_int)
        forecast_qty = corrected_forecast.get(
            (bid, pid_int),
            float(rec.get("forecast_qty") or 0),
        )
        stock_qty = yesterday_stock.get((bid, pid_int), 0.0)
        net_need = max(forecast_qty - stock_qty, 0.0)
        production_plan, kratnost_display = _production_plan_with_optional_kratnost(
            net_need,
            kratnost,
        )

        bname = bakery_info.get(bid, {}).get("name") or str(bid)
        rows.append({
            "bakery_id": bid,
            "bakery_name": bname,
            "category": category,
            "product_name": PRODUCT_NAME_OVERRIDES.get(
                pid_int, str(rec.get("product_name") or "")
            ),
            "forecast": round(forecast_qty, 1),
            "yesterday_stock": round(stock_qty, 1),
            "net_need": round(net_need, 1),
            "production_plan": production_plan,
            "total_for_sale": round(production_plan + stock_qty, 1),
            "kratnost": kratnost_display,
        })

    missing_kratnost_rows = sum(
        row["kratnost"] == MISSING_KRATNOST_LABEL for row in rows
    )
    if missing_kratnost_rows:
        print(
            "  WARNING: "
            f"{missing_kratnost_rows} forecast rows have no baking_sku_meta; "
            "kept with unit rounding"
        )

    rows.sort(key=lambda r: (r["bakery_id"], r["category"], r["product_name"]))
    return rows


def _build_excel(rows: list[dict], forecast_date: str) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    d = date_type.fromisoformat(forecast_date)
    weekday_name = WEEKDAY_RU[d.weekday()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Прогноз"

    # Header row
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_font = Font(bold=True, size=12)
    ws["A1"] = f"Прогноз выпечки — {d.strftime('%d.%m.%Y')} ({weekday_name})"
    ws["A1"].font = title_font
    ws.row_dimensions[1].height = 20

    headers = [
        "Пекарня",
        "Категория",
        "Номенклатура",
        "Прогноз",
        "Остаток со вчерашнего дня",
        "Чистая потребность",
        "План выпуска",
        "Итого на продажу",
        "Кратность",
    ]
    col_widths = [35, 20, 40, 12, 24, 20, 16, 18, 12]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[2].height = 30

    # Alternating bakery fill
    bakery_fills = [
        PatternFill("solid", fgColor="EBF3FB"),
        PatternFill("solid", fgColor="FFFFFF"),
    ]
    number_fmt = "#,##0.0"
    int_fmt = "#,##0"

    prev_bid = None
    fill_idx = 0
    for data_row in rows:
        if data_row["bakery_id"] != prev_bid:
            fill_idx = 1 - fill_idx
            prev_bid = data_row["bakery_id"]
        fill = bakery_fills[fill_idx]

        row_num = ws.max_row + 1
        values = [
            data_row["bakery_name"],
            data_row["category"],
            data_row["product_name"],
            data_row["forecast"],
            data_row["yesterday_stock"],
            data_row["net_need"],
            data_row["production_plan"],
            data_row["total_for_sale"],
            data_row["kratnost"],
        ]
        kratnost_fmt = int_fmt if isinstance(data_row["kratnost"], int) else None
        fmts = [
            None,
            None,
            None,
            number_fmt,
            number_fmt,
            number_fmt,
            int_fmt,
            number_fmt,
            kratnost_fmt,
        ]
        for col_idx, (val, fmt) in enumerate(zip(values, fmts), start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill = fill
            cell.border = cell_border
            cell.font = Font(size=10)
            if fmt:
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right")

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{ws.max_row}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


VIBECODE_API_BASE = "https://vibecode.bitrix24.tech/v1"
# Chat "Пилот выставления планов выпекания ИИ" — diskFolderId 1473995, chatId 179919
PILOT_CHAT_DIALOG_ID = "chat179919"
PILOT_CHAT_ID = 179919
PILOT_CHAT_DISK_FOLDER_ID = 1473995
# Native Bitrix24 webhook base URL — set via B24_WEBHOOK_URL env var (not hardcoded to avoid token leak)
# Example: https://franshizasvezhar.bitrix24.ru/rest/27979/<token>
B24_WEBHOOK_URL_ENV = "B24_WEBHOOK_URL"


def _send_to_chat(file_bytes: bytes, filename: str, forecast_date: str) -> None:
    """Upload Excel to the chat's Disk folder and send it as a file message.

    Flow:
      1. Native B24 disk.folder.uploadfile → get uploadUrl (supports Cyrillic filenames)
      2. POST file bytes to uploadUrl via multipart → get disk object id
      3. im.disk.file.commit → sends file as a proper attachment in chat
      4. Send short text message via VibeCode chats API
    """
    import json
    import time as _time
    import urllib.request

    api_key = os.environ.get("VIBECODE_API_KEY") or ""
    if not api_key:
        raise RuntimeError("VIBECODE_API_KEY not set")
    b24_webhook_base = os.environ.get(B24_WEBHOOK_URL_ENV, "").rstrip("/")
    if not b24_webhook_base:
        raise RuntimeError(f"{B24_WEBHOOK_URL_ENV} not set in environment")

    d = date_type.fromisoformat(forecast_date)
    weekday_name = WEEKDAY_RU[d.weekday()]
    # Russian filename shown in chat
    ru_filename = f"Прогноз_{d.strftime('%d.%m.%Y')}_{weekday_name}.xlsx"

    # Step 1: get uploadUrl from native B24 REST
    step1_body = json.dumps({
        "id": PILOT_CHAT_DISK_FOLDER_ID,
        "data": {"NAME": ru_filename},
        "generateUniqueName": "Y",
    }).encode("utf-8")
    req = urllib.request.Request(
        f"{b24_webhook_base}/disk.folder.uploadfile",
        data=step1_body,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        step1 = json.loads(resp.read())
    if "error" in step1:
        raise RuntimeError(f"disk.folder.uploadfile failed: {step1}")
    upload_url = step1["result"]["uploadUrl"]
    print("  [b24] uploadUrl obtained")

    # Step 2: POST file bytes to uploadUrl as multipart/form-data
    boundary = f"----FormBoundary{int(_time.time())}"
    body_parts = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="file"; filename="{ru_filename}"\r\n'
        f"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n"
        f"\r\n"
    ).encode("utf-8") + file_bytes + f"\r\n--{boundary}--\r\n".encode("utf-8")
    req = urllib.request.Request(
        upload_url,
        data=body_parts,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=120) as resp:
        step2 = json.loads(resp.read())
    if "error" in step2:
        raise RuntimeError(f"File upload to uploadUrl failed: {step2}")
    disk_id = step2["result"]["ID"]
    print(f"  [b24] file uploaded, disk_id={disk_id}")

    # Step 3: send short text message first via VibeCode
    msg_text = f"Прогноз — {d.strftime('%d.%m.%Y')} ({weekday_name})"
    msg_body = json.dumps({"message": msg_text}).encode("utf-8")
    req = urllib.request.Request(
        f"{VIBECODE_API_BASE}/chats/{PILOT_CHAT_DIALOG_ID}/messages",
        data=msg_body,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        msg_result = json.loads(resp.read())
    if not msg_result.get("success"):
        raise RuntimeError(f"Message send failed: {msg_result}")
    print(f"  [vibecode] text message sent, id={msg_result['data']}")

    # Step 4: commit file to chat — sends it as a proper attachment message
    commit_body = json.dumps({
        "CHAT_ID": PILOT_CHAT_ID,
        "DISK_ID": disk_id,
    }).encode("utf-8")
    req = urllib.request.Request(
        f"{b24_webhook_base}/im.disk.file.commit",
        data=commit_body,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        commit_result = json.loads(resp.read())
    if "error" in commit_result:
        raise RuntimeError(f"im.disk.file.commit failed: {commit_result}")
    file_msg_id = commit_result.get("result", {}).get("MESSAGE_ID")
    print(f"  [b24] file message sent, message_id={file_msg_id}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--env-file", default=".env")
    parser.add_argument("--date", default=None, help="Forecast date (YYYY-MM-DD); default: today")
    parser.add_argument("--dry-run", action="store_true", help="Build Excel but do not send to Bitrix24")
    parser.add_argument("--out-dir", default="output/pilot_forecast")
    parser.add_argument(
        "--forecast-override",
        default=None,
        help="Optional read-only CSV replacement for forecast rows; intended for controlled dry-runs.",
    )
    parser.add_argument(
        "--sku-correction-registry",
        default=None,
        help=(
            "Optional mature-SKU correction registry CSV. "
            "If omitted, PILOT_SKU_CORRECTION_REGISTRY is used."
        ),
    )
    args = parser.parse_args()

    if args.env_file and Path(args.env_file).exists():
        _load_env(args.env_file)

    forecast_date = args.date or str(date_type.today())
    d = date_type.fromisoformat(forecast_date)
    weekday_abbr = WEEKDAY_RU[d.weekday()][:2]

    print(f"Pilot forecast summary | date: {forecast_date} ({weekday_abbr})")

    registry_path = (
        args.sku_correction_registry
        or os.environ.get("PILOT_SKU_CORRECTION_REGISTRY")
        or None
    )
    rows = _build_report(
        forecast_date,
        sku_correction_registry=registry_path,
        forecast_override=args.forecast_override,
    )
    if not rows:
        print("No data found — aborting.")
        return

    print(f"  {len(rows)} SKU rows across {len({r['bakery_id'] for r in rows})} bakeries")

    file_bytes = _build_excel(rows, forecast_date)

    filename = f"Прогноз_выпечки_{d.strftime('%d.%m.%Y')}_{weekday_abbr}.xlsx"
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / filename
    out_path.write_bytes(file_bytes)
    print(f"  saved: {out_path}")

    if args.dry_run:
        print("  --dry-run: skipping Bitrix24 send")
        return

    api_key = (
        os.environ.get("VIBECODE_API_KEY")
        or os.environ.get("VIBECODE_API_KEY".lower())
        or ""
    )
    if not api_key:
        print("  VIBECODE_API_KEY not set — skipping Bitrix24 send")
        return

    print(f"  sending to {PILOT_CHAT_DIALOG_ID}...")
    _send_to_chat(file_bytes, filename, forecast_date)
    print("  done.")


if __name__ == "__main__":
    main()
