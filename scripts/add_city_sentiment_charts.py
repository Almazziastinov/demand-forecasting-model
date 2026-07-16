"""Add monthly city thank-you vs complaint stacked area charts to the workbook."""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
import sys
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent))

import build_city_detail_analytics as base
from openpyxl.chart import AreaChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation


INPUT_FILE = Path(
    r"C:\Users\dns\Documents\Codex\2026-06-30\new-chat\outputs\Отзывы покупателей единый реестр аналитика API категории без кофеен адреса и города нормализованы.xlsx"
)
OUTPUT_FILE = Path(
    r"C:\Users\dns\Documents\Codex\2026-06-30\new-chat\outputs\Отзывы покупателей единый реестр аналитика API категории без кофеен адреса и города нормализованы с графиками и выбором пекарни.xlsx"
)

TARGET_CITIES = [
    "Казань",
    "Иркутск",
    "Набережные Челны",
    "Чебоксары",
    "Новокузнецк",
    "Нижнекамск",
    "Зеленодольск",
    "Альметьевск",
]

BLUE = "8DB4E2"
RED = "E6A0A0"
BLUE_LINE = "5B9BD5"
RED_LINE = "C0504D"
LIGHT_BLUE = "D9EAF7"
LIGHT_RED = "FCE4D6"
CHOICE_FILL = "FFF2CC"


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def month_start(value: datetime) -> datetime:
    return datetime(value.year, value.month, 1)


def month_range(start: datetime, end: datetime) -> list[datetime]:
    current = month_start(start)
    last = month_start(end)
    months = []
    while current <= last:
        months.append(current)
        year = current.year + (1 if current.month == 12 else 0)
        month = 1 if current.month == 12 else current.month + 1
        current = datetime(year, month, 1)
    return months


def read_registry_rows(workbook) -> list[dict[str, Any]]:
    sheet = workbook["Единый реестр отзывов"]
    headers = [clean(cell.value) for cell in sheet[1]]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, values)))
    return rows


def build_counts(rows: list[dict[str, Any]]) -> tuple[list[datetime], dict[str, Counter[tuple[datetime, str]]]]:
    dates = [row.get("дата") for row in rows if isinstance(row.get("дата"), datetime)]
    months = month_range(min(dates), max(dates))
    counts: dict[str, Counter[tuple[datetime, str]]] = defaultdict(Counter)
    for row in rows:
        date = row.get("дата")
        if not isinstance(date, datetime):
            continue
        month = month_start(date)
        review_type = clean(row.get("тип_отзыва"))
        city = clean(row.get("город")) or "Неизвестно"
        counts["Все города"][(month, review_type)] += 1
        if city in TARGET_CITIES:
            counts[city][(month, review_type)] += 1
    return months, counts


def write_city_block(sheet, title: str, start_row: int, months: list[datetime], counts: Counter[tuple[datetime, str]]) -> int:
    sheet.cell(start_row, 1, title)
    sheet.cell(start_row, 1).font = base.Font(bold=True, size=13)
    header_row = start_row + 1
    headers = [
        "Месяц",
        "Благодарности",
        "Жалобы",
        "Всего",
        "Доля благодарностей",
        "Доля жалоб",
        "Благодарности / Жалобы",
    ]
    for column, header in enumerate(headers, 1):
        sheet.cell(header_row, column, header)

    for offset, month in enumerate(months, 1):
        row_index = header_row + offset
        thanks = counts[(month, "Благодарность")]
        complaints = counts[(month, "Жалоба")]
        total = thanks + complaints
        sheet.cell(row_index, 1, month.strftime("%Y-%m"))
        sheet.cell(row_index, 2, thanks)
        sheet.cell(row_index, 3, complaints)
        sheet.cell(row_index, 4, total)
        sheet.cell(row_index, 5, thanks / total if total else 0)
        sheet.cell(row_index, 6, complaints / total if total else 0)
        sheet.cell(row_index, 7, thanks / complaints if complaints else None)
        sheet.cell(row_index, 5).number_format = "0.0%"
        sheet.cell(row_index, 6).number_format = "0.0%"
        sheet.cell(row_index, 7).number_format = "0.00"

    base.style_header(sheet)

    last_row = header_row + len(months)
    chart = AreaChart()
    chart.grouping = "percentStacked"
    chart.title = f"{title}: доля благодарностей и жалоб"
    chart.style = 13
    chart.y_axis.title = "Доля отзывов"
    chart.x_axis.title = "Месяц"
    chart.height = 8
    chart.width = 24
    chart.add_data(Reference(sheet, min_col=5, max_col=6, min_row=header_row, max_row=last_row), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=header_row + 1, max_row=last_row))
    chart.y_axis.numFmt = "0%"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.tickLblSkip = 1
    chart.x_axis.noMultiLvlLbl = True
    apply_chart_colors(chart)
    sheet.add_chart(chart, f"I{start_row}")
    return last_row + 3


def apply_chart_colors(chart: AreaChart) -> None:
    colors = [BLUE, RED]
    line_colors = [BLUE_LINE, RED_LINE]
    for index, series in enumerate(chart.series):
        color = colors[index % len(colors)]
        line_color = line_colors[index % len(line_colors)]
        series.graphicalProperties.solidFill = color
        series.graphicalProperties.line.solidFill = line_color
        series.graphicalProperties.line.width = 18000
        series.dLbls = None


def add_charts_sheet(workbook, rows: list[dict[str, Any]]) -> None:
    sheet_name = "Динамика городов"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name, 2)
    months, counts = build_counts(rows)

    current_row = 1
    for title in ["Все города", *TARGET_CITIES]:
        current_row = write_city_block(sheet, title, current_row, months, counts[title])

    widths = [12, 16, 12, 12, 20, 14, 22]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[base.get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"


def unique_bakeries(rows: list[dict[str, Any]]) -> list[str]:
    counts = Counter(clean(row.get("адрес")) for row in rows)
    addresses = [
        address
        for address, count in counts.items()
        if address and address != "Неизвестно" and count > 0
    ]
    return ["Все пекарни", *sorted(addresses, key=lambda address: (-counts[address], address))]


def add_bakery_reference_sheet(workbook, rows: list[dict[str, Any]]) -> None:
    sheet_name = "Справочник пекарен"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name)
    sheet.sheet_state = "hidden"
    sheet["A1"] = "Пекарня"
    for index, address in enumerate(unique_bakeries(rows), start=2):
        sheet.cell(index, 1, address)


def add_interactive_bakery_chart(workbook, rows: list[dict[str, Any]]) -> None:
    sheet_name = "График по пекарне"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name, 3)

    dates = [row.get("дата") for row in rows if isinstance(row.get("дата"), datetime)]
    months = month_range(min(dates), max(dates))
    bakery_count = len(unique_bakeries(rows))

    sheet["A1"] = "Пекарня"
    sheet["B1"] = "Все пекарни"
    sheet["D1"] = "Выберите пекарню в желтой ячейке B1"
    sheet["A1"].font = base.Font(bold=True)
    sheet["B1"].fill = base.PatternFill("solid", fgColor=CHOICE_FILL)
    sheet["B1"].font = base.Font(bold=True)
    sheet["D1"].font = base.Font(italic=True, color="666666")

    validation = DataValidation(
        type="list",
        formula1=f"'Справочник пекарен'!$A$2:$A${bakery_count + 1}",
        allow_blank=False,
    )
    validation.promptTitle = "Выбор пекарни"
    validation.prompt = "Нажмите на стрелку справа в ячейке и выберите адрес пекарни."
    validation.errorTitle = "Пекарня не найдена"
    validation.error = "Выберите значение из выпадающего списка."
    sheet.add_data_validation(validation)
    validation.add(sheet["B1"])

    headers = [
        "Месяц",
        "Благодарности",
        "Жалобы",
        "Всего",
        "Доля благодарностей",
        "Доля жалоб",
        "Благодарности / Жалобы",
        "Дата месяца",
    ]
    header_row = 3
    for column, header in enumerate(headers, 1):
        sheet.cell(header_row, column, header)

    for offset, month in enumerate(months, 1):
        row_index = header_row + offset
        sheet.cell(row_index, 1, month.strftime("%Y-%m"))
        sheet.cell(row_index, 2, f'=IF($B$1="Все пекарни",COUNTIFS(\'Единый реестр отзывов\'!$A:$A,">="&$H{row_index},\'Единый реестр отзывов\'!$A:$A,"<"&DATE(YEAR($H{row_index}),MONTH($H{row_index})+1,1),\'Единый реестр отзывов\'!$D:$D,"Благодарность"),COUNTIFS(\'Единый реестр отзывов\'!$K:$K,$B$1,\'Единый реестр отзывов\'!$A:$A,">="&$H{row_index},\'Единый реестр отзывов\'!$A:$A,"<"&DATE(YEAR($H{row_index}),MONTH($H{row_index})+1,1),\'Единый реестр отзывов\'!$D:$D,"Благодарность"))')
        sheet.cell(row_index, 3, f'=IF($B$1="Все пекарни",COUNTIFS(\'Единый реестр отзывов\'!$A:$A,">="&$H{row_index},\'Единый реестр отзывов\'!$A:$A,"<"&DATE(YEAR($H{row_index}),MONTH($H{row_index})+1,1),\'Единый реестр отзывов\'!$D:$D,"Жалоба"),COUNTIFS(\'Единый реестр отзывов\'!$K:$K,$B$1,\'Единый реестр отзывов\'!$A:$A,">="&$H{row_index},\'Единый реестр отзывов\'!$A:$A,"<"&DATE(YEAR($H{row_index}),MONTH($H{row_index})+1,1),\'Единый реестр отзывов\'!$D:$D,"Жалоба"))')
        sheet.cell(row_index, 4, f"=B{row_index}+C{row_index}")
        sheet.cell(row_index, 5, f"=IF(D{row_index}=0,0,B{row_index}/D{row_index})")
        sheet.cell(row_index, 6, f"=IF(D{row_index}=0,0,C{row_index}/D{row_index})")
        sheet.cell(row_index, 7, f"=IF(C{row_index}=0,\"\",B{row_index}/C{row_index})")
        sheet.cell(row_index, 8, month)
        sheet.cell(row_index, 8).number_format = "yyyy-mm"
        sheet.cell(row_index, 5).number_format = "0.0%"
        sheet.cell(row_index, 6).number_format = "0.0%"
        sheet.cell(row_index, 7).number_format = "0.00"

    base.style_header(sheet)
    sheet["B1"].fill = base.PatternFill("solid", fgColor=CHOICE_FILL)
    sheet["B1"].font = base.Font(bold=True)
    sheet["D1"].font = base.Font(italic=True, color="666666")
    sheet.freeze_panes = "A4"
    widths = [12, 16, 12, 12, 20, 14, 22, 12, 24]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[base.get_column_letter(index)].width = width
    sheet.column_dimensions["H"].hidden = True

    last_row = header_row + len(months)
    chart = AreaChart()
    chart.grouping = "percentStacked"
    chart.title = "Выбранная пекарня: доля благодарностей и жалоб"
    chart.y_axis.title = "Доля отзывов"
    chart.x_axis.title = "Месяц"
    chart.height = 14
    chart.width = 32
    chart.add_data(Reference(sheet, min_col=5, max_col=6, min_row=header_row, max_row=last_row), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=header_row + 1, max_row=last_row))
    chart.y_axis.numFmt = "0%"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.tickLblSkip = 1
    chart.x_axis.noMultiLvlLbl = True
    apply_chart_colors(chart)
    sheet.add_chart(chart, "I3")

    sheet.conditional_formatting.add(
        f"E4:E{last_row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["0.5"], fill=base.PatternFill("solid", fgColor=LIGHT_BLUE)),
    )
    sheet.conditional_formatting.add(
        f"F4:F{last_row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["0.5"], fill=base.PatternFill("solid", fgColor=LIGHT_RED)),
    )


def main() -> None:
    workbook = base.openpyxl.load_workbook(INPUT_FILE)
    rows = read_registry_rows(workbook)
    add_charts_sheet(workbook, rows)
    add_bakery_reference_sheet(workbook, rows)
    add_interactive_bakery_chart(workbook, rows)
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_FILE)
    print(f"rows={len(rows)}")
    print(f"charts={len(['Все города', *TARGET_CITIES]) + 1}")
    print(f"bakeries={len(unique_bakeries(rows)) - 1}")
    print(f"output={OUTPUT_FILE}")


if __name__ == "__main__":
    main()
