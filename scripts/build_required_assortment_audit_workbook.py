"""Build a readable Excel audit workbook for required assortment issues."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT_DIR = ROOT / "reports" / "required_assortment"
DEFAULT_OUTPUT_PATH = DEFAULT_INPUT_DIR / "required_assortment_audit.xlsx"


def join_unique(series: pd.Series) -> str:
    values = {str(value) for value in series.dropna() if str(value) != "nan"}
    return " | ".join(sorted(values))


def select_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    return df[[column for column in columns if column in df.columns]]


def add_problem_groups(
    *,
    rows: list[dict[str, object]],
    df: pd.DataFrame,
    problem_type: str,
    severity: str,
    problem: str,
    recommended_action: str,
    extra_columns: list[str] | None = None,
) -> None:
    if df.empty:
        return

    group_columns = ["product_name", "category", *(extra_columns or [])]
    for keys, group in df.groupby(group_columns, dropna=False):
        if not isinstance(keys, tuple):
            keys = (keys,)
        grouped = dict(zip(group_columns, keys, strict=False))
        rows.append(
            {
                "severity": severity,
                "problem_type": problem_type,
                "product_name": grouped.get("product_name", ""),
                "required_category": join_unique(group["category"]),
                "dim_categories": grouped.get("dim_categories", ""),
                "excel_categories": grouped.get("full_tops_category_any_category", ""),
                "cities": join_unique(group["city"]),
                "city_count": group["city"].nunique(),
                "top_required_any": int(
                    pd.to_numeric(
                        group.get("is_top", pd.Series([0])),
                        errors="coerce",
                    )
                    .fillna(0)
                    .max()
                ),
                "problem": problem,
                "recommended_action": recommended_action,
                "source_scope": join_unique(group["market_scope"])
                if "market_scope" in group
                else "",
            }
        )


def build_problem_sheet(
    *,
    missing_dim: pd.DataFrame,
    category_mismatch: pd.DataFrame,
    missing_full_actionable: pd.DataFrame,
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    add_problem_groups(
        rows=rows,
        df=missing_dim,
        problem_type="Нет в dim_products",
        severity="HIGH",
        problem=(
            "Обязательная позиция из OCR не найдена в справочнике "
            "Svezhar.dim_products по нормализованному названию."
        ),
        recommended_action=(
            "Проверить, нужно ли добавить позицию в справочник, заменить на "
            "актуальное название/синоним или исключить из обязательного ассортимента."
        ),
    )
    add_problem_groups(
        rows=rows,
        df=category_mismatch,
        problem_type="Категория отличается в dim_products",
        severity="MEDIUM",
        problem=(
            "Позиция найдена в dim_products, но категория из OCR отличается "
            "от категории справочника."
        ),
        recommended_action=(
            "Подтвердить целевую категорию; если справочник верный, зафиксировать "
            "alias категории в контракте ассортимента."
        ),
        extra_columns=["dim_categories"],
    )
    add_problem_groups(
        rows=rows,
        df=missing_full_actionable,
        problem_type="Нет в Excel топах/продажах",
        severity="MEDIUM",
        problem=(
            "Позиция есть в dim_products, но не найдена в полном Excel-топе/"
            "детализации продаж для города."
        ),
        recommended_action=(
            "Проверить городовую применимость: обязательна ли позиция для "
            "этого города, либо scope OCR слишком широкий."
        ),
    )

    problems = pd.DataFrame(rows)
    if problems.empty:
        return problems

    severity_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    problems["_severity_order"] = problems["severity"].map(severity_order).fillna(9)
    return problems.sort_values(
        ["_severity_order", "problem_type", "product_name"]
    ).drop(columns=["_severity_order"])


def build_workbook(
    input_dir: Path = DEFAULT_INPUT_DIR,
    output_path: Path = DEFAULT_OUTPUT_PATH,
) -> Path:
    required_vs_dim = pd.read_csv(input_dir / "required_vs_dim_products.csv")
    missing_dim = pd.read_csv(input_dir / "required_missing_from_dim_products.csv")
    category_mismatch = pd.read_csv(input_dir / "required_dim_category_mismatches.csv")
    inactive_dim = pd.read_csv(input_dir / "required_inactive_in_dim_products.csv")
    missing_full = pd.read_csv(input_dir / "required_missing_from_full_tops.csv")
    summary = pd.read_csv(input_dir / "required_assortment_summary.csv")
    director_comparison_path = (
        input_dir / "director_tatarstan_assortment_comparison.csv"
    )
    director_comparison = (
        pd.read_csv(director_comparison_path)
        if director_comparison_path.exists()
        else pd.DataFrame()
    )

    missing_full_actionable = missing_full.merge(
        required_vs_dim[
            [
                "city",
                "product_key",
                "dim_product_status",
                "present_in_dim_products",
            ]
        ],
        on=["city", "product_key"],
        how="left",
    )
    missing_full_actionable = missing_full_actionable[
        missing_full_actionable["dim_product_status"].eq("active_found")
    ].copy()

    problems = build_problem_sheet(
        missing_dim=missing_dim,
        category_mismatch=category_mismatch,
        missing_full_actionable=missing_full_actionable,
    )

    summary_cards = pd.DataFrame(
        [
            {
                "metric": "Обязательных строк город-товар",
                "value": len(required_vs_dim),
                "comment": "Развернуто по городам из OCR scope",
            },
            {
                "metric": "Не найдено в dim_products, строк",
                "value": len(missing_dim),
                "comment": (
                    f"Уникальных позиций: {missing_dim['product_name'].nunique()}"
                ),
            },
            {
                "metric": "Выведено в dim_products, строк",
                "value": len(inactive_dim),
                "comment": (
                    f"Уникальных позиций: {inactive_dim['product_name'].nunique()}"
                ),
            },
            {
                "metric": "Категория отличается в dim_products, строк",
                "value": len(category_mismatch),
                "comment": (
                    f"Уникальных позиций: "
                    f"{category_mismatch['product_name'].nunique()}"
                ),
            },
            {
                "metric": "Нет в полном Excel-топе/продажах, строк",
                "value": len(missing_full),
                "comment": (
                    "После исключения отсутствующих/выведенных в dim_products: "
                    f"{len(missing_full_actionable)}"
                ),
            },
            {
                "metric": "OCR Татарстан: нет в файле директора, позиций",
                "value": (
                    int((~director_comparison["present_in_director_tatarstan"]).sum())
                    if not director_comparison.empty
                    else ""
                ),
                "comment": (
                    f"Из {len(director_comparison)} уникальных OCR-позиций"
                    if not director_comparison.empty
                    else "Файл сравнения директора не найден"
                ),
            },
            {
                "metric": "Дата сборки",
                "value": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"),
                "comment": "Локальное время",
            },
        ]
    )

    sheets = {
        "Резюме": summary_cards,
        "Проблемы": problems,
        "Нет в справочнике": select_columns(
            missing_dim,
            [
                "market_scope",
                "city",
                "category",
                "product_name",
                "is_top",
                "top_rank",
                "present_in_full_tops",
                "present_in_sales_detail",
                "source_note",
            ],
        ),
        "Выведено": select_columns(
            inactive_dim,
            [
                "market_scope",
                "city",
                "category",
                "product_name",
                "is_top",
                "top_rank",
                "inactive_dim_product_ids",
                "inactive_dim_product_names",
                "inactive_dim_categories",
                "present_in_full_tops",
                "present_in_sales_detail",
            ],
        ),
        "Категории": select_columns(
            category_mismatch,
            [
                "market_scope",
                "city",
                "category",
                "product_name",
                "is_top",
                "top_rank",
                "dim_product_ids",
                "dim_product_names",
                "dim_categories",
                "present_in_full_tops",
                "full_tops_category_any_category",
            ],
        ),
        "Нет в Excel": select_columns(
            missing_full_actionable,
            [
                "market_scope",
                "city",
                "category",
                "product_name",
                "is_top",
                "top_rank",
                "dim_product_status",
                "present_in_dim_products",
                "present_in_sales_detail",
                "dim_product_names",
                "dim_categories",
            ],
        ),
        "Директор Татарстан": select_columns(
            director_comparison,
            [
                "director_status",
                "category",
                "product_name",
                "is_top",
                "top_rank",
                "director_product_name",
                "director_supercategory",
                "director_category",
                "director_qty",
                "director_bakery_count",
                "director_current_price",
            ],
        ),
        "Свод города": summary,
        "Полный контракт": select_columns(
            required_vs_dim,
            [
                "market_scope",
                "city",
                "category",
                "product_name",
                "is_required",
                "is_top",
                "top_rank",
                "dim_product_status",
                "present_in_dim_products",
                "dim_category_mismatch",
                "present_in_full_tops",
                "full_tops_category_mismatch",
                "present_in_sales_detail",
                "dim_product_ids",
                "dim_product_names",
                "dim_categories",
                "inactive_dim_product_ids",
                "inactive_dim_product_names",
            ],
        ),
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    target_path = output_path
    try:
        writer = pd.ExcelWriter(target_path, engine="openpyxl")
    except PermissionError:
        target_path = output_path.with_name("required_assortment_audit_updated.xlsx")
        writer = pd.ExcelWriter(target_path, engine="openpyxl")

    with writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)

    format_workbook(target_path)
    return target_path


def format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    high_fill = PatternFill("solid", fgColor="FEE2E2")
    medium_fill = PatternFill("solid", fgColor="FEF3C7")
    green_fill = PatternFill("solid", fgColor="DCFCE7")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = border

        for row in worksheet.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

        if max_row >= 2 and max_col >= 1:
            reference = f"A1:{get_column_letter(max_col)}{max_row}"
            table_name = "".join(
                character for character in worksheet.title if character.isalnum()
            )[:20]
            table = Table(displayName=f"{table_name or 'Audit'}Table", ref=reference)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)

        for column_idx in range(1, max_col + 1):
            letter = get_column_letter(column_idx)
            values = [
                str(worksheet.cell(row=row_idx, column=column_idx).value or "")
                for row_idx in range(1, min(max_row, 80) + 1)
            ]
            width = min(max(max(len(value) for value in values) + 2, 10), 55)
            if worksheet.cell(row=1, column=column_idx).value in {
                "problem",
                "recommended_action",
                "cities",
                "dim_product_names",
            }:
                width = 45
            worksheet.column_dimensions[letter].width = width

        for row_idx in range(1, max_row + 1):
            worksheet.row_dimensions[row_idx].height = 30 if row_idx == 1 else 42

    summary_sheet = workbook["Резюме"]
    summary_sheet.column_dimensions["A"].width = 38
    summary_sheet.column_dimensions["B"].width = 18
    summary_sheet.column_dimensions["C"].width = 70
    for row in summary_sheet.iter_rows(min_row=2, max_row=summary_sheet.max_row):
        row[0].font = Font(bold=True)
        if "Не найдено" in str(row[0].value):
            for cell in row:
                cell.fill = high_fill
        elif "Категория" in str(row[0].value):
            for cell in row:
                cell.fill = medium_fill

    problem_sheet = workbook["Проблемы"]
    headers = [cell.value for cell in problem_sheet[1]]
    if "severity" in headers and problem_sheet.max_row >= 2:
        severity_letter = get_column_letter(headers.index("severity") + 1)
        data_range = (
            f"A2:{get_column_letter(problem_sheet.max_column)}"
            f"{problem_sheet.max_row}"
        )
        problem_sheet.conditional_formatting.add(
            data_range,
            FormulaRule(formula=[f'${severity_letter}2="HIGH"'], fill=high_fill),
        )
        problem_sheet.conditional_formatting.add(
            data_range,
            FormulaRule(formula=[f'${severity_letter}2="MEDIUM"'], fill=medium_fill),
        )

    contract_sheet = workbook["Полный контракт"]
    for row in contract_sheet.iter_rows(min_row=2, max_row=contract_sheet.max_row):
        for cell in row:
            if cell.value is True:
                cell.fill = green_fill
            elif cell.value is False:
                cell.fill = high_fill

    workbook.save(path)


def main() -> None:
    output_path = build_workbook()
    print(output_path)


if __name__ == "__main__":
    main()
