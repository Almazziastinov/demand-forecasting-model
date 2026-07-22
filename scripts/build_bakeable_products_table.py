"""Build city-specific bakeable-product allowlists for the baking plan.

Two modes:

1. **Category filter** (default, new behaviour):
   Reads the active assortment CSV and keeps only products whose category_name
   matches one of the bakeable category patterns (пироги / выпечка / фастфуд).
   No manual markup files required.

2. **Markup xlsx** (legacy, kept for backward compatibility):
   Pass ``--markup-xlsx <path>`` to use the old partner red-font markup logic.
   Products rendered with red font are subtracted from the assortment.

Category patterns are case-insensitive substring matches against category_name.
Adjust with ``--bakeable-categories`` if the ClickHouse category names differ
from the defaults.

Usage
-----
# Default: category filter (prod)
.venv\\Scripts\\python.exe scripts\\build_bakeable_products_table.py

# Dev tables
.venv\\Scripts\\python.exe scripts\\build_bakeable_products_table.py \\
    --assortment-csv reports\\required_assortment\\assortment_city_products.csv

# Override category patterns
.venv\\Scripts\\python.exe scripts\\build_bakeable_products_table.py \\
    --bakeable-categories "пирог" "выпечка" "фастфуд" "слойка"

# Legacy xlsx-markup mode
.venv\\Scripts\\python.exe scripts\\build_bakeable_products_table.py \\
    --markup-xlsx \\
    reports\\baking_plan_templates\\preview_sibirskiy_25_screenshot_assortment_v2.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from scripts.build_required_assortment_contract import (  # noqa: E402
    DEFAULT_OUTPUT_DIR,
    normalize_text,
)
from scripts.compare_required_assortment_to_dim_products import (  # noqa: E402
    normalize_product_for_dim_lookup,
)

DEFAULT_ASSORTMENT_CSV = DEFAULT_OUTPUT_DIR / "assortment_city_products.csv"
DEFAULT_DIM_PRODUCTS_PATH = DEFAULT_OUTPUT_DIR / "dim_products_lookup.csv"
DEFAULT_OUTPUT_PATH = DEFAULT_OUTPUT_DIR / "bakeable_products.csv"
DEFAULT_VALID_FROM = pd.Timestamp.now().date().isoformat()
SOURCE_NAME = "forecast_category_filter"

# Default category patterns (case-insensitive substring match on category_name).
# A product is bakeable if its category_name contains ANY of these patterns.
DEFAULT_BAKEABLE_CATEGORY_PATTERNS: list[str] = [
    "пирог",     # covers: Пироги сытные, Пироги сладкие, Пирог, etc.
    "выпечка",   # covers: Выпечка, Выпечка ночная, etc.
    "фастфуд",   # covers: Фастфуд, Фастфуд горячий, etc.
]

# Markup xlsx legacy constants (used only in --markup-xlsx mode)
MARKUP_START_ROW = 6
MARKUP_NAME_COLUMN = 2
RED_FONT_RGBS = {"FFFF0000", "FF0000"}
MARKUP_NAME_ALIASES = {
    normalize_text("Торт Меренговый с абрикосом"): "Меренговый с абрикосом",
    normalize_text("Торт Меренговый с вишней"): "Меренговый с вишней",
}


# ---------------------------------------------------------------------------
# Shared: assortment loading
# ---------------------------------------------------------------------------


def load_active_assortment(assortment_csv: Path) -> pd.DataFrame:
    assortment = pd.read_csv(assortment_csv, dtype={"product_id": str})
    active = assortment[
        pd.to_numeric(assortment["is_active"], errors="coerce").fillna(0).eq(1)
    ].copy()
    required_columns = {"city", "product_id", "product_name", "category_name"}
    missing = required_columns - set(active.columns)
    if missing:
        raise ValueError(
            "Assortment CSV is missing columns: " + ", ".join(sorted(missing))
        )
    universe = active.dropna(subset=["city", "product_id"]).copy()
    universe["city"] = universe["city"].astype(str).str.strip()
    universe["product_id"] = universe["product_id"].astype(str).str.strip()
    universe = universe[
        universe["city"].ne("") & universe["product_id"].ne("")
    ].drop_duplicates(["city", "product_id"])
    return universe


def _finalise_table(
    df: pd.DataFrame,
    *,
    source: str,
    source_file: str,
    valid_from: str,
) -> pd.DataFrame:
    table = df[["city", "product_id", "product_name", "category_name"]].copy()
    loaded_at = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
    table["is_bakeable"] = 1
    table["source"] = source
    table["source_file"] = source_file
    table["valid_from"] = pd.to_datetime(valid_from).date().isoformat()
    table["valid_to"] = pd.NA
    table["is_active"] = 1
    table["loaded_at"] = loaded_at
    table["comment"] = ""
    return table.sort_values(["city", "product_id"]).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Mode 1: category filter (new default)
# ---------------------------------------------------------------------------


def build_bakeable_by_category(
    *,
    assortment_csv: Path,
    category_patterns: list[str],
    valid_from: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Filter assortment to bakeable categories.

    Returns (bakeable_table, excluded_table).
    """
    universe = load_active_assortment(assortment_csv)

    category_col = universe["category_name"].fillna("").str.lower()
    pattern_re = "|".join(p.lower() for p in category_patterns)
    is_bakeable = category_col.str.contains(pattern_re, regex=True)

    included = universe[is_bakeable].copy()
    excluded = universe[~is_bakeable].copy()

    table = _finalise_table(
        included,
        source=SOURCE_NAME,
        source_file="category_filter:" + ",".join(category_patterns),
        valid_from=valid_from,
    )
    return table, excluded


# ---------------------------------------------------------------------------
# Mode 2: markup xlsx (legacy)
# ---------------------------------------------------------------------------


def _read_red_markup_names(markup_xlsx: Path) -> list[str]:
    import openpyxl
    workbook = openpyxl.load_workbook(markup_xlsx)
    sheet = workbook.active
    names: list[str] = []
    for row in range(MARKUP_START_ROW, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=MARKUP_NAME_COLUMN)
        name = cell.value
        if not name:
            continue
        font = cell.font
        color = font.color if font else None
        if (
            color is not None
            and color.type == "rgb"
            and str(color.rgb).upper() in RED_FONT_RGBS
        ):
            names.append(str(name).strip())
    return names


def _build_dim_id_lookups(
    dim_products_path: Path,
) -> tuple[dict[str, set[str]], dict[str, set[str]]]:
    dim = pd.read_csv(dim_products_path, dtype={"product_id": str})
    dim["product_key_exact"] = dim["product_name"].map(normalize_text)
    dim["product_key_lookup"] = dim["product_name"].map(
        normalize_product_for_dim_lookup
    )
    exact = dim.groupby("product_key_exact")["product_id"].apply(set).to_dict()
    alias = dim.groupby("product_key_lookup")["product_id"].apply(set).to_dict()
    return exact, alias


def _resolve_markup_ids(
    names: list[str],
    exact_lookup: dict[str, set[str]],
    alias_lookup: dict[str, set[str]],
) -> tuple[set[str], list[str]]:
    matched: set[str] = set()
    not_found: list[str] = []
    for name in names:
        lookup_name = MARKUP_NAME_ALIASES.get(normalize_text(name), name)
        ids = exact_lookup.get(normalize_text(lookup_name)) or alias_lookup.get(
            normalize_product_for_dim_lookup(lookup_name)
        )
        if ids:
            matched |= ids
        else:
            not_found.append(name)
    return matched, not_found


def build_bakeable_by_markup(
    *,
    assortment_csv: Path,
    markup_xlsx: Path,
    dim_products_path: Path,
    valid_from: str,
) -> tuple[pd.DataFrame, set[str], list[str]]:
    """Legacy mode: subtract red-font products from assortment."""
    universe = load_active_assortment(assortment_csv)
    red_names = _read_red_markup_names(markup_xlsx)
    exact_lookup, alias_lookup = _build_dim_id_lookups(dim_products_path)
    excluded_ids, not_found = _resolve_markup_ids(red_names, exact_lookup, alias_lookup)
    if not_found:
        names = ", ".join(sorted(not_found))
        raise ValueError(f"Red markup products not found in dim_products: {names}")
    included = universe[~universe["product_id"].isin(excluded_ids)].copy()
    table = _finalise_table(
        included,
        source="partner_baking_markup",
        source_file=markup_xlsx.name,
        valid_from=valid_from,
    )
    return table, excluded_ids, not_found


# Public legacy API retained for callers that used the script as a module
# before the category-filter mode introduced explicit builder names.
read_red_markup_names = _read_red_markup_names
resolve_markup_ids = _resolve_markup_ids
build_bakeable_table = build_bakeable_by_markup


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build bakeable product allowlist from assortment CSV."
    )
    parser.add_argument(
        "--assortment-csv",
        default=DEFAULT_ASSORTMENT_CSV,
        type=Path,
        help=(
            "Path to assortment_city_products CSV "
            "(output of build_city_assortment_from_forecast.py)."
        ),
    )
    parser.add_argument(
        "--output-path",
        default=DEFAULT_OUTPUT_PATH,
        type=Path,
    )
    parser.add_argument(
        "--valid-from",
        default=DEFAULT_VALID_FROM,
        help="valid_from date for output rows (ISO format).",
    )

    # Category filter mode (default)
    parser.add_argument(
        "--bakeable-categories",
        nargs="+",
        default=DEFAULT_BAKEABLE_CATEGORY_PATTERNS,
        metavar="PATTERN",
        help=(
            "Case-insensitive substrings to match against category_name. "
            "A product is included if any pattern matches. "
            f"Default: {DEFAULT_BAKEABLE_CATEGORY_PATTERNS}"
        ),
    )

    # Legacy markup mode
    parser.add_argument(
        "--markup-xlsx",
        default=None,
        type=Path,
        help=(
            "Legacy mode: path to partner baking-plan preview xlsx with "
            "red-font markup. "
            "When set, category filter is NOT used."
        ),
    )
    parser.add_argument(
        "--dim-products-path",
        default=DEFAULT_DIM_PRODUCTS_PATH,
        type=Path,
        help="Required only in --markup-xlsx mode.",
    )
    args = parser.parse_args()

    args.output_path.parent.mkdir(parents=True, exist_ok=True)

    if args.markup_xlsx is not None:
        # ---- legacy mode ----
        print("mode: markup xlsx (legacy)")
        print(f"markup: {args.markup_xlsx}")
        table, excluded_ids, not_found = build_bakeable_by_markup(
            assortment_csv=args.assortment_csv,
            markup_xlsx=args.markup_xlsx,
            dim_products_path=args.dim_products_path,
            valid_from=args.valid_from,
        )
        table.to_csv(args.output_path, index=False, encoding="utf-8-sig")
        print(f"bakeable rows         : {len(table)}")
        print(f"excluded (red) ids    : {len(excluded_ids)}")
        print(f"red names not matched : {len(not_found)}")
    else:
        # ---- category filter mode (default) ----
        print("mode: category filter")
        print(f"patterns : {args.bakeable_categories}")
        table, excluded = build_bakeable_by_category(
            assortment_csv=args.assortment_csv,
            category_patterns=args.bakeable_categories,
            valid_from=args.valid_from,
        )
        table.to_csv(args.output_path, index=False, encoding="utf-8-sig")
        print(f"bakeable rows    : {len(table)}")
        print(f"excluded rows    : {len(excluded)}")

        if not excluded.empty and "category_name" in excluded.columns:
            print("\nexcluded category breakdown:")
            print(
                excluded.groupby(excluded["category_name"].fillna("<null>"))["product_id"]
                .nunique()
                .sort_values(ascending=False)
                .to_string()
            )

    print(f"\noutput: {args.output_path}")


if __name__ == "__main__":
    main()
