"""
Выгрузка отзывов из чата Bitrix24 (chat40781) за 2026 год в Excel.
Использует VibeCode API с курсорной пагинацией (lastId).
"""

import requests
import json
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

API_KEY = "vibe_api_o9eDjakzMxux55JyoJS6GINHhfqYeuc5_dc336e"
BASE_URL = "https://vibecode.bitrix24.tech/v1"
CHAT_ID = "chat40781"
OUTPUT_PATH = r"C:\Users\dns\Desktop\reviews_2026.xlsx"

HEADERS = {
    "Authorization": f"Bearer {API_KEY}",
}


def fetch_all_messages():
    """Получает все сообщения за 2026 год из чата."""
    all_messages = []
    last_id = None
    page = 0
    stop = False

    print("Начинаю выгрузку сообщений из чата...")

    while not stop:
        params = {"limit": 200}
        if last_id:
            params["lastId"] = last_id

        url = f"{BASE_URL}/chats/{CHAT_ID}/messages"
        resp = requests.get(url, headers=HEADERS, params=params)
        resp.raise_for_status()
        data = resp.json()

        if not data.get("success"):
            print(f"Ошибка API: {data}")
            break

        messages = data["data"]["messages"]
        if not messages:
            print("Больше нет сообщений.")
            break

        page += 1
        added = 0
        oldest_date = None

        for msg in messages:
            msg_date = datetime.fromisoformat(msg["date"])
            oldest_date = msg_date

            if msg_date.year < 2026:
                stop = True
                break

            if msg_date.year == 2026:
                all_messages.append(msg)
                added += 1

        last_id = messages[-1]["id"]
        print(f"Страница {page}: получено {len(messages)} сообщений, добавлено {added}, "
              f"последняя дата: {oldest_date.strftime('%Y-%m-%d') if oldest_date else '?'}")

        if stop:
            print("Достигли сообщений до 2026 года. Остановка.")
            break

    print(f"\nВсего получено сообщений за 2026 год: {len(all_messages)}")
    return all_messages


def parse_review(text):
    """Парсит текст сообщения и извлекает поля отзыва."""
    # Формат: "Отзыв 2ГИС DD.MM.YYYY Город, Адрес Текст отзыва"
    date_str = ""
    source = ""
    address = ""
    review_text = text

    # Извлечь источник (2ГИС, Яндекс и т.д.)
    source_match = re.search(r"Отзыв\s+([^\d]+?)\s+\d{2}\.\d{2}\.\d{4}", text)
    if source_match:
        source = source_match.group(1).strip()

    # Извлечь дату
    date_match = re.search(r"(\d{2}\.\d{2}\.\d{4})", text)
    if date_match:
        date_str = date_match.group(1)

    # Текст после даты
    after_date = re.sub(r"Отзыв\s+\S+\s+\d{2}\.\d{2}\.\d{4}\s*", "", text).strip()

    # Попробуем выделить адрес (первая строка/предложение до текста отзыва)
    # Адрес обычно выглядит как "Город, Улица, Номер"
    addr_match = re.match(
        r"^([А-ЯЁа-яёA-Za-z\s\-]+,\s*[А-ЯЁа-яёA-Za-z\s\-]+(?:улица|проспект|бульвар|переулок|площадь|шоссе|пр-кт|пр\.|ул\.)[^,\n]*(?:,\s*[\d\w]+)?)\s+(.+)",
        after_date,
        re.DOTALL | re.IGNORECASE
    )
    if addr_match:
        address = addr_match.group(1).strip()
        review_text = addr_match.group(2).strip()
    else:
        # Простой вариант: первое предложение как адрес если оно короткое
        parts = after_date.split("\n", 1)
        if len(parts) == 2 and len(parts[0]) < 100:
            address = parts[0].strip()
            review_text = parts[1].strip()
        else:
            # Разделить по первым 80 символам
            space_parts = re.split(r"(?<=\S)\s{2,}", after_date, maxsplit=1)
            if len(space_parts) == 2:
                address = space_parts[0].strip()
                review_text = space_parts[1].strip()
            else:
                review_text = after_date

    return {
        "source": source,
        "date_str": date_str,
        "address": address,
        "review_text": review_text,
    }


def save_to_excel(messages, output_path):
    """Сохраняет отзывы в Excel файл."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отзывы 2026"

    # Заголовки
    headers = [
        "ID сообщения",
        "Дата сообщения",
        "Источник",
        "Дата отзыва",
        "Адрес пекарни",
        "Текст отзыва",
        "Полный текст",
    ]

    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Данные — сортируем от старых к новым
    sorted_msgs = sorted(messages, key=lambda m: m["date"])

    alt_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")

    for row_idx, msg in enumerate(sorted_msgs, 2):
        msg_date = datetime.fromisoformat(msg["date"])
        parsed = parse_review(msg["text"])

        row_data = [
            msg["id"],
            msg_date.strftime("%Y-%m-%d %H:%M"),
            parsed["source"],
            parsed["date_str"],
            parsed["address"],
            parsed["review_text"],
            msg["text"],
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Ширина столбцов
    col_widths = [14, 18, 12, 14, 40, 60, 80]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"\nФайл сохранён: {output_path}")
    print(f"Строк данных: {len(sorted_msgs)}")


def main():
    messages = fetch_all_messages()
    if not messages:
        print("Нет сообщений для сохранения.")
        return
    save_to_excel(messages, OUTPUT_PATH)


if __name__ == "__main__":
    main()
