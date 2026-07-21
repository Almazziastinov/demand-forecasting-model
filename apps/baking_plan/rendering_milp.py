"""Build the "План выпекания" xlsx sheet from scratch for the MILP allocator.

Restored from git 3b18eac, import updated: demand -> demand_milp.
"""

from __future__ import annotations

# ruff: noqa: E501
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .demand_milp import SkuDemand
from .templates import Window

SHEET_NAME = "План выпекания"
CATEGORY_ORDER = [
    "Выпечка сытная",
    "Выпечка сладкая",
    "Пироги сытные",
    "Пироги сладкие",
    "Фастфуд",
]
HEADER_FILL = PatternFill("solid", fgColor="FFD8D8D8")
CATEGORY_FILL = PatternFill("solid", fgColor="FFD9EAF7")
DEFROST_FILL = PatternFill("solid", fgColor="FFFCE4D6")
TWO_DAY_FILL = PatternFill("solid", fgColor="FFE6D9F2")
MANDATORY_FILL = PatternFill("solid", fgColor="FFFFFF00")
SHORTFALL_FULL_FILL = PatternFill("solid", fgColor="FFFFC7CE")    # red — zero produced, demand > 0
SHORTFALL_PARTIAL_FILL = PatternFill("solid", fgColor="FFFFEB9C") # yellow — produced < demand
CAPACITY_NOTE_FILL = PatternFill("solid", fgColor="FFFFEB9C")

DEFROST_SUFFIX = " (доп. партия на завтра)"
SHORTFALL_TOLERANCE = 1e-6

MANDATORY_ASSORTMENT = {
    "Треугольник курица безд",
    "Треугольник говядина безд",
    "Треугольник острый",
    "Вак-бэлиш",
    "Жар пицца с курицей",
    "Сосиска в тесте",
    "Элеш с курицей",
    "Беккен капуста",
    "Сосиска под шубой",
    "Кыстыбый П",
}


def render_workbook(
    *,
    bakery_name: str,
    forecast_date: str,
    windows: list[Window],
    skus: list[SkuDemand],
    regular_alloc: dict[tuple[str, str], float],
    defrost_alloc: dict[tuple[str, str], float],
    two_day_alloc: dict[tuple[str, str], float],
    shortfall_by_sku: dict[str, float] | None = None,
    defrost_shortfall_by_sku: dict[str, float] | None = None,
    capacity_note: str | None = None,
) -> Workbook:
    shortfall_by_sku = shortfall_by_sku or {}
    defrost_shortfall_by_sku = defrost_shortfall_by_sku or {}
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME

    total_col = 3 + len(windows)

    sheet["A1"] = f"План выпекания: {bakery_name} на {forecast_date}."
    sheet["A1"].font = Font(bold=True)
    sheet["A2"] = (
        "Количество по SKU-hour прогнозу активного run, разнесено по окнам "
        "выпекания через MILP-оптимизацию; доп. партия на завтра (дефрост/"
        "двухдневка) размещается в любом окне со свободной мощностью."
    )
    sheet["A2"].font = Font(bold=True)

    header_row = 3
    if capacity_note:
        sheet.cell(row=3, column=1, value=capacity_note)
        sheet.cell(row=3, column=1).font = Font(bold=True)
        for col in range(1, total_col + 1):
            sheet.cell(row=3, column=col).fill = CAPACITY_NOTE_FILL
        header_row = 4

    sheet.cell(row=header_row, column=1, value="Стол")
    sheet.cell(row=header_row, column=2, value="Наименование")
    for idx, window in enumerate(windows):
        sheet.cell(row=header_row, column=3 + idx, value=window.label)
    sheet.cell(row=header_row, column=total_col, value="Итого")
    for col in range(1, total_col + 1):
        cell = sheet.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    row = header_row
    for category in CATEGORY_ORDER:
        members = sorted(
            (sku for sku in skus if sku.category_name == category),
            key=lambda sku: sku.avg_daily_sales,
            reverse=True,
        )
        if not members:
            continue

        row += 1
        category_cell = sheet.cell(row=row, column=1, value=category)
        category_cell.font = Font(bold=True)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_col)
        for col in range(1, total_col + 1):
            sheet.cell(row=row, column=col).fill = CATEGORY_FILL

        for sku in members:
            row += 1
            station_cell = sheet.cell(row=row, column=1, value=sku.station or "")
            name_cell = sheet.cell(row=row, column=2, value=sku.product_name)
            if sku.product_name in MANDATORY_ASSORTMENT:
                station_cell.fill = MANDATORY_FILL
                name_cell.fill = MANDATORY_FILL
            row_total = 0.0
            for idx, window in enumerate(windows):
                key = (sku.product_id, window.label)
                two_day_qty = two_day_alloc.get(key, 0)
                if two_day_qty:
                    window_cell = sheet.cell(row=row, column=3 + idx, value=two_day_qty)
                    window_cell.fill = TWO_DAY_FILL
                    row_total += two_day_qty
                    continue
                qty = regular_alloc.get(key, 0)
                defrost_qty = defrost_alloc.get(key, 0)
                cell_value = _format_cell(qty, defrost_qty)
                if cell_value != "":
                    window_cell = sheet.cell(row=row, column=3 + idx, value=cell_value)
                    if defrost_qty:
                        window_cell.fill = DEFROST_FILL
                    row_total += qty

            shortfall = shortfall_by_sku.get(sku.product_id, 0.0) - defrost_shortfall_by_sku.get(
                sku.product_id, 0.0
            )
            total_cell = sheet.cell(row=row, column=total_col)
            forecast_total = row_total + shortfall
            if shortfall > SHORTFALL_TOLERANCE and round(row_total) != round(forecast_total):
                if row_total <= SHORTFALL_TOLERANCE:
                    total_cell.value = round(forecast_total)
                    total_cell.fill = SHORTFALL_FULL_FILL
                else:
                    total_cell.value = f"{round(row_total)}/{round(forecast_total)}"
                    total_cell.fill = SHORTFALL_PARTIAL_FILL
            else:
                total_cell.value = row_total

    _set_column_widths(sheet, total_col)
    return workbook


def _format_cell(qty: float, defrost_qty: float) -> object:
    if defrost_qty:
        return f"{qty + defrost_qty:g}{DEFROST_SUFFIX}"
    if qty:
        return qty
    return ""


def _set_column_widths(sheet: Worksheet, total_col: int) -> None:
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 34
    for col in range(3, total_col + 1):
        sheet.column_dimensions[sheet.cell(row=3, column=col).column_letter].width = 14
