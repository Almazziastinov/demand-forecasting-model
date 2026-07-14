"""Write the computed plan rows back into the loaded template sheet.

Mutates the sheet in place (row snapshot/restore, unmerge, delete, rewrite)
instead of building a fresh Workbook — the template's own row structure and
cell styling (including the SKU rows' original C:L formatting) is exactly
what "use the template" means for this feature, so it's preserved rather
than rebuilt from scratch.
"""

from __future__ import annotations

# ruff: noqa: E501
from copy import copy
from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .templates import PLAN_START_ROW, WINDOWS_HEADER_ROW, Window

SHEET_TITLE = "План выпекания"
CATEGORY_FILL = PatternFill("solid", fgColor="D9EAF7")
GROUP_SORT_ORDER = {
    "Выпечка сытная": 0,
    "Выпечка сладкая": 1,
    "Пироги сытные": 2,
    "Пироги сладкие": 3,
    "Фастфуд": 4,
}


def snapshot_row(sheet: Worksheet, row_index: int, total_column: int) -> dict[str, Any]:
    return {
        "cells": [
            {
                "value": sheet.cell(row=row_index, column=column).value,
                "style": copy(sheet.cell(row=row_index, column=column)._style),
            }
            for column in range(1, total_column)
        ],
        "height": sheet.row_dimensions[row_index].height,
    }


def _restore_row(
    sheet: Worksheet,
    row_index: int,
    snapshot: dict[str, Any] | None,
    prototype: dict[str, Any] | None,
) -> None:
    source = snapshot or prototype
    if not source:
        return
    for column, cell_snapshot in enumerate(source["cells"], start=1):
        cell = sheet.cell(row=row_index, column=column)
        cell.value = cell_snapshot["value"] if snapshot else None
        cell._style = copy(cell_snapshot["style"])
    sheet.row_dimensions[row_index].height = source["height"]


def write_plan(
    *,
    sheet: Worksheet,
    windows: list[Window],
    plan_rows: list[dict[str, Any]],
    bakery_name: str,
    forecast_date: str,
    selected_sheet_name: str,
) -> None:
    """Rewrite `sheet`'s data rows (from `PLAN_START_ROW`) with `plan_rows`.

    Each plan row: {snapshot, product_id, product_name, category_name,
    allocated: {column: value}, total: float, source_order: int}.
    `snapshot` is `None` for assortment SKUs with no template row (see
    `service.py`) — those get plain formatting, no window breakdown.
    """
    total_column = 3 + len(windows)  # A=Стол, B=Наименование, C..=windows, then Итого

    sheet["A1"] = f"План выпекания: {bakery_name} на {forecast_date}. Шаблон: {selected_sheet_name}"
    sheet["A2"] = (
        "Количество по SKU-hour прогнозу активного run, разнесено по расписанию "
        "выпекания шаблона; ночная дефростация — по утреннему прогнозу след. дня."
    )

    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.max_row >= PLAN_START_ROW:
            sheet.unmerge_cells(str(merged_range))
    if sheet.max_row >= PLAN_START_ROW:
        sheet.delete_rows(PLAN_START_ROW, sheet.max_row - PLAN_START_ROW + 1)

    sheet.cell(row=WINDOWS_HEADER_ROW, column=total_column).value = "Итого"
    sheet.cell(row=WINDOWS_HEADER_ROW, column=total_column)._style = copy(
        sheet.cell(row=WINDOWS_HEADER_ROW, column=total_column - 1)._style
    )
    if sheet.max_column > total_column:
        sheet.delete_cols(total_column + 1, sheet.max_column - total_column)
    sheet.column_dimensions[sheet.cell(row=1, column=total_column).column_letter].width = 14
    sheet.column_dimensions[sheet.cell(row=1, column=total_column).column_letter].hidden = False

    prototype: dict[str, Any] | None = next(
        (row["snapshot"] for row in plan_rows if row["snapshot"] is not None), None
    )

    plan_rows = sorted(
        plan_rows,
        key=lambda row: (
            GROUP_SORT_ORDER.get(str(row["category_name"]), 100),
            str(row["category_name"]).casefold(),
            int(row["source_order"]),
            str(row["product_name"]).casefold(),
        ),
    )

    current_row = PLAN_START_ROW
    current_category: str | None = None
    for plan_row in plan_rows:
        category = str(plan_row["category_name"] or "Без группы")
        if category != current_category:
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_column)
            category_cell = sheet.cell(row=current_row, column=1)
            category_cell.value = category
            category_cell.fill = CATEGORY_FILL
            category_cell.font = Font(bold=True)
            category_cell.alignment = Alignment(horizontal="left")
            sheet.row_dimensions[current_row].hidden = False
            current_category = category
            current_row += 1

        _restore_row(sheet, current_row, plan_row["snapshot"], prototype)
        if plan_row["snapshot"] is None:
            for column in range(1, total_column + 1):
                sheet.cell(row=current_row, column=column).fill = PatternFill()
        sheet.cell(row=current_row, column=2).value = plan_row["product_name"]
        for window in windows:
            target_cell = sheet.cell(row=current_row, column=window.column)
            target_cell.value = plan_row["allocated"].get(window.column)
        total_cell = sheet.cell(row=current_row, column=total_column)
        total_cell.value = plan_row["total"]
        total_cell.number_format = "0"
        total_cell.alignment = Alignment(horizontal="center")
        sheet.row_dimensions[current_row].hidden = False
        current_row += 1

    for row_index in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row_index].hidden = False
    for column in range(1, total_column + 1):
        column_letter = sheet.cell(row=1, column=column).column_letter
        sheet.column_dimensions[column_letter].hidden = False
