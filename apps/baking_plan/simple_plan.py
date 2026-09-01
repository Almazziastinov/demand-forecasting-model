"""Single-bakery production plan matching the pilot chat workbook logic."""

from __future__ import annotations

import math
from datetime import date as date_type
from datetime import timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ._clickhouse import get_client, records, table_name

BAKEABLE_CATEGORIES = {
    "Пироги сытные",
    "Пироги сладкие",
    "Выпечка сытная",
    "Выпечка сладкая",
    "Фастфуд",
}
FROZEN_MARKER = "замороженные полуфабрикаты"
PRODUCT_NAME_OVERRIDES = {
    11615: "Плетенка кленовая",
    11616: "Плетенка с черникой",
    11617: "Плетенка с земляникой",
}
MISSING_KRATNOST_LABEL = "нет данных по кратности"
WEEKDAY_RU = [
    "Понедельник",
    "Вторник",
    "Среда",
    "Четверг",
    "Пятница",
    "Суббота",
    "Воскресенье",
]


def round_up_kratnost(value: float, kratnost: int) -> int:
    if value <= 0 or kratnost <= 0:
        return 0
    return int(math.ceil(value / kratnost - 1e-9) * kratnost)


def production_plan_with_optional_kratnost(
    net_need: float,
    kratnost: int | None,
) -> tuple[int, int | str]:
    if kratnost is None:
        return max(0, int(math.ceil(net_need - 1e-9))), MISSING_KRATNOST_LABEL
    return round_up_kratnost(net_need, kratnost), kratnost


def calculate_plan_rows(
    forecast_rows: list[dict],
    stock_by_product: dict[int, float],
    base_meta: dict[int, dict],
    bakery_meta: dict[int, dict],
    *,
    bakery_name: str,
) -> list[dict]:
    """Apply the same stock subtraction and kratnost rounding as the pilot publisher."""
    result: list[dict] = []
    for item in forecast_rows:
        category = str(item.get("category_name") or "")
        if category not in BAKEABLE_CATEGORIES:
            continue
        product_id = int(item["product_id"])
        meta = bakery_meta.get(product_id) or base_meta.get(product_id)
        if meta and FROZEN_MARKER in str(meta.get("dough_group") or "").lower():
            continue
        kratnost = max(1, int(meta.get("kratnost") or 1)) if meta else None
        forecast = max(0.0, float(item.get("forecast_qty") or 0.0))
        stock = max(0.0, float(stock_by_product.get(product_id, 0.0)))
        net_need = max(forecast - stock, 0.0)
        production_plan, kratnost_display = production_plan_with_optional_kratnost(
            net_need,
            kratnost,
        )
        result.append(
            {
                "bakery_name": bakery_name,
                "category": category,
                "product_name": PRODUCT_NAME_OVERRIDES.get(
                    product_id, str(item.get("product_name") or "")
                ),
                "forecast": round(forecast, 1),
                "yesterday_stock": round(stock, 1),
                "net_need": round(net_need, 1),
                "production_plan": production_plan,
                "total_for_sale": round(production_plan + stock, 1),
                "kratnost": kratnost_display,
            }
        )
    return sorted(result, key=lambda row: (row["category"], row["product_name"]))


def _load_forecast_rows(run_id: str, forecast_date: str, bakery_id: int) -> list[dict]:
    client = get_client()
    frame = client.query_df(
        f"""
        with
        day_forecast as (
            select
                toInt64(product_id) product_id,
                any(product_name) product_name,
                any(category_name) category_name,
                sum(forecast_qty) forecast_qty
            from {table_name('sku_forecast_day_embedded')}
            where run_id=%(run_id)s
              and forecast_date=toDate(%(forecast_date)s)
              and bakery_id=%(bakery_id)s
            group by product_id
        ),
        horizon_average as (
            select toInt64(product_id) product_id, avg(forecast_qty) forecast_qty
            from {table_name('sku_forecast_day_embedded')}
            where run_id=%(run_id)s and bakery_id=%(bakery_id)s
            group by product_id
        ),
        effective_assortment as (
            select distinct toInt64OrZero(product_id) product_id
            from {table_name('bakery_product_assortment_embedded')} final
            where bakery_id=%(bakery_id)s
              and valid_from=(
                  select max(valid_from)
                  from {table_name('bakery_product_assortment_embedded')} final
                  where bakery_id=%(bakery_id)s
                    and valid_from <= toDate(%(forecast_date)s)
              )
        )
        select
            a.product_id product_id,
            coalesce(any(d.product_name), any(p.product_name), '') product_name,
            coalesce(any(d.category_name), any(p.category_name), '') category_name,
            if(
                max(d.product_id)=0,
                coalesce(any(h.forecast_qty), 0.0),
                any(d.forecast_qty)
            ) forecast_qty
        from effective_assortment a
        left join day_forecast d on d.product_id=a.product_id
        left join horizon_average h on h.product_id=a.product_id
        left join (
            select
                toInt64OrZero(toString(product_id)) product_id,
                any(product_name) product_name,
                any(category_name) category_name
            from Svezhar.dim_products
            group by product_id
        ) p on p.product_id=a.product_id
        group by a.product_id
        order by category_name, product_name, product_id
        """,
        parameters={
            "run_id": run_id,
            "forecast_date": forecast_date,
            "bakery_id": bakery_id,
        },
    )
    return records(frame)


def _load_stock(forecast_date: str, bakery_id: int) -> dict[int, float]:
    client = get_client()
    previous_date = str(date_type.fromisoformat(forecast_date) - timedelta(days=1))
    sold = client.query_df(
        """
        select toInt64OrZero(toString(product_id)) product_id,
               sum(toFloat64(quantity)) qty_sold
        from (
            select distinct check_datetime, check_date, bakery_id, product_id,
                   quantity, price, line_amount, cash_event_type
            from Svezhar.fct_check_lines
            where hex(cash_event_type)='D09FD180D0BED0B4D0B0D0B6D0B0'
              and check_date=toDate(%(previous_date)s)
              and toInt64OrZero(toString(bakery_id))=%(bakery_id)s
        )
        group by product_id
        """,
        parameters={"previous_date": previous_date, "bakery_id": bakery_id},
    )
    produced = client.query_df(
        """
        select toInt64OrZero(toString(pid)) product_id, sum(qty) qty_produced
        from (
            select argMax(product_id, _updated_at) pid,
                   toFloat64(argMax(quantity, _updated_at)) qty
            from Svezhar.fct_production_release
            where toDate(release_date)=toDate(%(previous_date)s)
              and toInt64OrZero(toString(bakery_id))=%(bakery_id)s
            group by release_id, line_id
            having argMax(is_deleted, _updated_at) not in ('1', 'true', 'Да')
        )
        group by product_id
        """,
        parameters={"previous_date": previous_date, "bakery_id": bakery_id},
    )
    sold_by_id = {
        int(row["product_id"]): float(row["qty_sold"] or 0)
        for row in records(sold)
    }
    produced_by_id = {
        int(row["product_id"]): float(row["qty_produced"] or 0)
        for row in records(produced)
    }
    return {
        product_id: max(produced_qty - sold_by_id.get(product_id, 0.0), 0.0)
        for product_id, produced_qty in produced_by_id.items()
        if produced_qty > sold_by_id.get(product_id, 0.0)
    }


def _load_meta(
    product_ids: list[int], bakery_id: int
) -> tuple[dict[int, dict], dict[int, dict]]:
    if not product_ids:
        return {}, {}
    product_ids_padded = [f"{product_id:09d}" for product_id in product_ids]
    frame = get_client().query_df(
        f"""
        select product_id, bakery_id, dough_group, kratnost, scope
        from {table_name('baking_sku_meta')} final
        where is_active=1 and product_id in %(product_ids)s
          and (scope='base' or (scope='bakery' and bakery_id=%(bakery_id)s))
        """,
        parameters={"product_ids": product_ids_padded, "bakery_id": bakery_id},
    )
    base: dict[int, dict] = {}
    bakery: dict[int, dict] = {}
    for row in records(frame):
        target = bakery if row.get("scope") == "bakery" else base
        target[int(row["product_id"])] = row
    return base, bakery


def render_workbook(rows: list[dict], forecast_date: str, bakery_name: str) -> bytes:
    day = date_type.fromisoformat(forecast_date)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "План выпуска"
    sheet["A1"] = (
        f"План выпуска — {bakery_name} — {day.strftime('%d.%m.%Y')} "
        f"({WEEKDAY_RU[day.weekday()]})"
    )
    sheet["A1"].font = Font(bold=True, size=12)
    headers = [
        "Пекарня",
        "Категория",
        "Номенклатура",
        "Прогноз",
        "Остаток со вчерашнего дня",
        "Чистая потребность",
        "План выпуска",
        "Итого на продажу",
        "Кратность",
    ]
    widths = [35, 20, 40, 12, 24, 20, 16, 18, 12]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for column, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = sheet.cell(2, column, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in rows:
        values = [
            row["bakery_name"], row["category"], row["product_name"], row["forecast"],
            row["yesterday_stock"], row["net_need"], row["production_plan"],
            row["total_for_sale"], row["kratnost"],
        ]
        row_number = sheet.max_row + 1
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column, value)
            cell.border = border
            if column >= 4 and not isinstance(value, str):
                cell.number_format = "#,##0.0" if column not in (7, 9) else "#,##0"
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:I{sheet.max_row}"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_workbook(
    *, run_id: str, forecast_date: str, bakery_id: int, bakery_name: str
) -> bytes:
    forecast_rows = _load_forecast_rows(run_id, forecast_date, bakery_id)
    product_ids = [int(row["product_id"]) for row in forecast_rows]
    base_meta, bakery_meta = _load_meta(product_ids, bakery_id)
    rows = calculate_plan_rows(
        forecast_rows,
        _load_stock(forecast_date, bakery_id),
        base_meta,
        bakery_meta,
        bakery_name=bakery_name,
    )
    return render_workbook(rows, forecast_date, bakery_name)
