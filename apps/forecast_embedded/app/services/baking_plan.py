from __future__ import annotations

import logging
import math
import re
from copy import copy
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


logger = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE_PATH = ROOT / "assets" / "baking_plan_template.xlsx"
INDIVIDUAL_TEMPLATE_DIR = ROOT / "assets" / "baking_plan_individual"
INDIVIDUAL_TEMPLATE_PATHS = {
    20: INDIVIDUAL_TEMPLATE_DIR / "20_mira_45.xlsx",
    21: INDIVIDUAL_TEMPLATE_DIR / "21_parkovaya_7.xlsx",
    22: INDIVIDUAL_TEMPLATE_DIR / "22_sibirskiy_trakt_25.xlsx",
}
DEFAULT_BUCKET = "до 2,5 млн"
REVENUE_BUCKETS = (
    (1_500_000, "до 1,5 млн"),
    (2_500_000, "до 2,5 млн"),
    (3_000_000, "от 2,5 млн"),
)
WINDOW_COLUMNS = tuple(range(3, 13))
TOTAL_COLUMN = 13
PLAN_START_ROW = 6
GROUP_SORT_ORDER = {
    "Выпечка сытная": 0,
    "Выпечка сладкая": 1,
    "Пироги сытные": 2,
    "Фастфуд": 3,
}
FIRST_SALES_HOUR = 6
LAST_SALES_HOUR = 23
# Night-defrost columns are prep for tomorrow's morning batches; size them from
# tomorrow's forecast over hours [FIRST_SALES_HOUR .. DEFROST_EARLY_CUTOFF - 1].
DEFROST_EARLY_CUTOFF = 12
DEFROST_MARKERS = ("дефр", "ночн")
SKU_ALIAS_TO_CANONICAL = {
    "треугольник курица безд": "треугольник курица",
    "треугольник говядина безд": "треугольник говядина",
    "хуплу": "хуплу чебоксары",
    "элеш с курицей": "элеш",
    "конвертик курица": "конвертик с курицей",
    "ватрушка": "ватрушка в ассортменте",
    "жарпицца пикантная": "жар пицца пикантная",
    "жарпицца оригинальная": "жар пицца оригинальная",
    "кыстыбый п": "кыстыбый",
    "киш курица": "киш с курицей",
    "жар киш курица": "жар киш с курицей",
    "трехслойник новый": "трехслойник",
    "пирог ханский": "ханский",
    "капустный": "пирог капустный",
    "капуста и мясо": "пирог капуста мясо",
    "капуста и курица": "пирог капуста курица",
    "горбуша саго": "пирог горбуша саго",
    "пирожок яблоко": "пирожок булочка с яблоками",
    "клубника и банан новый": "клубника банан",
    "клубника и банан зкз": "клубника банан",
    "печенье детское 250": "печенье детское",
}


@dataclass(frozen=True)
class BakingWindow:
    column: int
    label: str
    start_hour: int
    end_hour: int


_PEAK_MERGE_HOURS = 3
_PEAK_MIN_SHARE = 0.05
_PEAK_MIN_PROMINENCE = 1.2
_FORECAST_HOURS = list(range(6, 23))


@dataclass(frozen=True)
class SkuMeta:
    """Metadata for a SKU parsed from the template 'комментарии' sheet."""

    dough_group: str
    kratnost: int
    is_two_day: bool
    is_on_demand: bool


@dataclass(frozen=True)
class ScheduledColumn:
    """A C:L cell that the template pre-filled for a SKU row.

    ``is_defrost`` marks night-defrost prep (sized from tomorrow's early forecast,
    excluded from today's coverage). ``note`` keeps the original cell text so the
    defrost annotation (e.g. ``"20 (ночная дефр)"``) can be re-emitted.
    """

    window: BakingWindow
    is_defrost: bool
    note: str | None
    quantity: int | None = None
    is_artificial: bool = False


def is_defrost_cell(value: object) -> bool:
    text = str(value or "").lower().replace("ё", "е")
    return any(marker in text for marker in DEFROST_MARKERS)


def normalize_sku_name(value: object) -> str:
    text = str(value or "").lower().replace("ё", "е")
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"[^0-9a-zа-я]+", " ", text)
    stop_words = {
        "тесто",
        "ночного",
        "ночное",
        "брожжения",
        "дефростация",
        "ночная",
        "ночн",
        "дефр",
    }
    tokens = [token for token in text.split() if token not in stop_words]
    return " ".join(tokens)


def sku_match_keys(value: object) -> list[str]:
    key = normalize_sku_name(value)
    if not key:
        return []

    keys = [key]
    alias = SKU_ALIAS_TO_CANONICAL.get(key)
    if alias:
        keys.append(alias)

    if "жарпицца" in key:
        keys.append(key.replace("жарпицца", "жар пицца"))

    result: list[str] = []
    seen: set[str] = set()
    for item in keys:
        if item and item not in seen:
            result.append(item)
            seen.add(item)
    return result


def revenue_bucket(revenue: float | int | None) -> str:
    if revenue is None:
        return DEFAULT_BUCKET
    value = float(revenue)
    for threshold, bucket in REVENUE_BUCKETS:
        if value < threshold:
            return bucket
    return "от 3млн"


def parse_window_label(value: object, column: int) -> BakingWindow | None:
    label = str(value or "").strip()
    match = re.search(r"(\d{1,2})[:.]\d{2}\s*-\s*(\d{1,2})[:.]\d{2}", label)
    if not match:
        return None
    return BakingWindow(
        column=column,
        label=label,
        start_hour=int(match.group(1)),
        end_hour=int(match.group(2)),
    )


def coverage_hours(
    windows: list[BakingWindow],
    *,
    first_sales_hour: int = FIRST_SALES_HOUR,
    last_sales_hour: int = LAST_SALES_HOUR,
) -> dict[int, list[int]]:
    """Map each bake window to the sales hours it covers.

    A batch becomes available at its window's end hour and covers sales until the
    next batch becomes available (the next window's end hour); the last window runs
    through ``last_sales_hour``. The earliest window always starts its coverage at
    ``first_sales_hour`` so a SKU baked only once (e.g. midday) still absorbs the
    whole day's forecast instead of silently dropping the morning.
    """
    if not windows:
        return {}

    ordered = sorted(
        windows,
        key=lambda item: (item.start_hour, item.end_hour, item.column),
    )
    result: dict[int, list[int]] = {}
    for index, window in enumerate(ordered):
        if index == 0:
            start_hour = first_sales_hour
        else:
            start_hour = max(window.end_hour, first_sales_hour)

        if index + 1 < len(ordered):
            end_hour = ordered[index + 1].end_hour - 1
        else:
            end_hour = last_sales_hour

        if end_hour >= start_hour:
            result[window.column] = list(range(start_hour, end_hour + 1))
        else:
            result[window.column] = []
    return result


def build_product_hour_lookup(
    sku_hour_rows: list[dict[str, Any]],
) -> dict[str, dict[int, float]]:
    lookup: dict[str, dict[int, float]] = {}
    for row in sku_hour_rows:
        name_keys = sku_match_keys(row.get("product_name"))
        if not name_keys or row.get("hour") is None:
            continue

        hour = int(row["hour"])
        qty = float(row.get("forecast_qty") or 0.0)
        for name_key in name_keys:
            lookup.setdefault(name_key, {})
            lookup[name_key][hour] = lookup[name_key].get(hour, 0.0) + qty
    return lookup


def build_product_total_lookup(
    sku_hour_rows: list[dict[str, Any]],
) -> dict[str, float]:
    lookup: dict[str, float] = {}
    for row in sku_hour_rows:
        qty = float(row.get("forecast_qty") or 0.0)
        for name_key in sku_match_keys(row.get("product_name")):
            lookup[name_key] = lookup.get(name_key, 0.0) + qty
    return lookup


def build_assortment_lookup(
    assortment_rows: list[dict[str, Any]],
) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    by_name: dict[str, dict[str, Any]] = {}
    unique: dict[str, dict[str, Any]] = {}
    for source in assortment_rows:
        product_id = str(source.get("product_id") or "").strip()
        product_name = str(source.get("product_name") or "").strip()
        if not product_id or not product_name:
            continue
        record = {
            "product_id": product_id,
            "product_name": product_name,
            "category_name": str(source.get("category_name") or "Без группы").strip()
            or "Без группы",
        }
        unique.setdefault(product_id, record)
        for name_key in sku_match_keys(product_name):
            current = by_name.get(name_key)
            if current is None or _assortment_match_priority(record) < (
                _assortment_match_priority(current)
            ):
                by_name[name_key] = record
    return by_name, list(unique.values())


def _assortment_match_priority(product: dict[str, Any]) -> tuple[int, str]:
    text = (
        f"{product.get('product_name', '')} {product.get('category_name', '')}"
    ).casefold()
    is_order_product = any(marker in text for marker in ("зкз", "заказ"))
    return int(is_order_product), str(product.get("product_id") or "")


def resolve_assortment_product(
    template_sku_name: object,
    assortment_lookup: dict[str, dict[str, Any]],
) -> dict[str, Any] | None:
    for name_key in sku_match_keys(template_sku_name):
        product = assortment_lookup.get(name_key)
        if product:
            return product
    return None


def _resolve_product_total(
    product_name: object,
    product_total_lookup: dict[str, float],
) -> float:
    for name_key in sku_match_keys(product_name):
        if name_key in product_total_lookup:
            return product_total_lookup[name_key]
    return 0.0


def _snapshot_row(sheet: Worksheet, row_index: int) -> dict[str, Any]:
    return {
        "cells": [
            {
                "value": sheet.cell(row=row_index, column=column).value,
                "style": copy(sheet.cell(row=row_index, column=column)._style),
            }
            for column in range(1, TOTAL_COLUMN)
        ],
        "height": sheet.row_dimensions[row_index].height,
    }


def _restore_row(
    sheet: Worksheet,
    row_index: int,
    snapshot: dict[str, Any] | None,
    prototype: dict[str, Any] | None,
) -> None:
    source = snapshot or prototype
    if not source:
        return
    for column, cell_snapshot in enumerate(source["cells"], start=1):
        cell = sheet.cell(row=row_index, column=column)
        cell.value = cell_snapshot["value"] if snapshot else None
        cell._style = copy(cell_snapshot["style"])
    sheet.row_dimensions[row_index].height = source["height"]


def _numeric_plan_total(values: dict[int, Any]) -> float:
    return float(
        sum(value for value in values.values() if isinstance(value, (int, float)))
    )


def _resolve_hourly_forecast(
    template_sku_name: object,
    product_hour_lookup: dict[str, dict[int, float]],
) -> dict[int, float] | None:
    for name_key in sku_match_keys(template_sku_name):
        hourly = product_hour_lookup.get(name_key)
        if hourly:
            return hourly
    return None


def read_row_schedule(
    sheet: Worksheet,
    row_index: int,
    sheet_windows: dict[int, BakingWindow],
) -> list[ScheduledColumn]:
    """Read which C:L columns the template pre-filled for this SKU row.

    A non-empty cell marks a scheduled baking window. Cells whose text carries a
    night-defrost marker (e.g. ``"20 (ночная дефр)"``) are classified as defrost
    prep. Classification keys off the *cell value*, not the SKU name (many SKU
    names contain "ночного брожжения" yet have plain integer bake cells).
    """
    schedule: list[ScheduledColumn] = []
    for column, window in sheet_windows.items():
        value = sheet.cell(row=row_index, column=column).value
        if value is None or str(value).strip() == "":
            continue
        defrost = is_defrost_cell(value)
        schedule.append(
            ScheduledColumn(
                window=window,
                is_defrost=defrost,
                note=str(value) if defrost else None,
                quantity=_extract_positive_int(value),
            )
        )
    return schedule


def _round_up(qty: float, round_to: int) -> int:
    if round_to <= 1:
        return int(math.ceil(qty))
    return int(math.ceil(qty / round_to) * round_to)


def _extract_positive_int(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value if value > 0 else None
    if isinstance(value, float):
        rounded = int(round(value))
        return rounded if rounded > 0 and math.isclose(value, rounded) else None

    match = re.search(r"\d+", str(value))
    if not match:
        return None
    parsed = int(match.group(0))
    return parsed if parsed > 0 else None


def schedule_round_to(schedule: list[ScheduledColumn]) -> int:
    quantities = [
        item.quantity
        for item in schedule
        if not item.is_defrost and item.quantity is not None and item.quantity > 0
    ]
    if not quantities:
        quantities = [
            item.quantity
            for item in schedule
            if item.quantity is not None and item.quantity > 0
        ]
    if not quantities:
        return 1

    round_to = quantities[0]
    for quantity in quantities[1:]:
        round_to = math.gcd(round_to, quantity)
    return max(round_to, 1)


def _early_window_sum(
    hourly: dict[int, float],
    first_sales_hour: int = FIRST_SALES_HOUR,
    cutoff: int = DEFROST_EARLY_CUTOFF,
) -> float:
    return sum(
        qty for hour, qty in hourly.items() if first_sales_hour <= hour < cutoff
    )


def _with_midday_split_window(
    bake_windows: list[BakingWindow],
    *,
    available_windows: dict[int, BakingWindow] | None,
    hourly: dict[int, float],
    first_sales_hour: int,
    last_sales_hour: int,
) -> list[BakingWindow]:
    if len(bake_windows) != 1 or not available_windows:
        return bake_windows

    early_window = bake_windows[0]
    if early_window.end_hour > DEFROST_EARLY_CUTOFF:
        return bake_windows

    later_demand = sum(
        qty
        for hour, qty in hourly.items()
        if DEFROST_EARLY_CUTOFF <= hour <= last_sales_hour
    )
    if later_demand <= 0:
        return bake_windows

    candidates = [
        window
        for window in available_windows.values()
        if window.column != early_window.column
        and window.end_hour >= DEFROST_EARLY_CUTOFF
        and window.start_hour >= first_sales_hour
    ]
    if not candidates:
        return bake_windows

    split_window = sorted(
        candidates,
        key=lambda item: (
            abs(item.end_hour - DEFROST_EARLY_CUTOFF),
            item.end_hour,
            item.column,
        ),
    )[0]
    return [early_window, split_window]


def parse_comments_sheet(workbook: Any) -> dict[str, SkuMeta]:
    """Parse 'комментарии' sheet → normalized_sku_name → SkuMeta."""
    if "комментарии" not in workbook.sheetnames:
        return {}
    ws = workbook["комментарии"]

    def _is_qty(v: object) -> bool:
        if v is None:
            return False
        try:
            float(str(v).replace(",", "."))
            return True
        except ValueError:
            pass
        return str(v).strip().lower() == "по запросу"

    result: dict[str, SkuMeta] = {}
    current_group: str | None = None
    for row in ws.iter_rows(values_only=True):
        b = row[1] if len(row) > 1 else None
        c = row[2] if len(row) > 2 else None
        d = row[3] if len(row) > 3 else None
        b_str = str(b).strip() if b is not None else ""
        c_str = str(c).strip() if c is not None else ""

        if _is_qty(c):
            if b_str and b_str not in ("кратность", "кратность выпуска") and current_group:
                is_on_demand = str(c).strip().lower() == "по запросу"
                try:
                    kratnost = max(1, int(float(str(c).replace(",", "."))))
                except (ValueError, TypeError):
                    kratnost = 1
                is_two_day = str(d or "").strip().lower() == "двухдневка"
                norm = normalize_sku_name(b_str)
                if norm:
                    result[norm] = SkuMeta(
                        dough_group=current_group,
                        kratnost=kratnost,
                        is_two_day=is_two_day,
                        is_on_demand=is_on_demand,
                    )
        elif b_str and c_str in ("", "None", "кратность выпуска"):
            current_group = b_str
    return result


def _resolve_sku_meta(
    template_sku_name: object,
    sku_meta: dict[str, SkuMeta],
) -> SkuMeta | None:
    for key in sku_match_keys(template_sku_name):
        meta = sku_meta.get(key)
        if meta:
            return meta
    return None


def _detect_peaks(hourly: dict[int, float]) -> list[int]:
    vol = [hourly.get(h, 0.0) for h in _FORECAST_HOURS]
    total = sum(vol)
    if total == 0:
        return [_FORECAST_HOURS[0]]
    norm = [v / total for v in vol]
    peaks: list[int] = []
    for i in range(1, len(norm) - 1):
        if norm[i] > norm[i - 1] and norm[i] > norm[i + 1]:
            lt = min(norm[max(0, i - 2):i] or [norm[i]])
            rt = min(norm[i + 1:min(len(norm), i + 3)] or [norm[i]])
            prom = norm[i] / max(lt, rt, 0.001)
            if prom >= _PEAK_MIN_PROMINENCE and norm[i] >= _PEAK_MIN_SHARE:
                peaks.append(_FORECAST_HOURS[i])
    if not peaks:
        peaks = [_FORECAST_HOURS[int(max(range(len(norm)), key=lambda idx: norm[idx]))]]
    return peaks


def _cluster_peaks(peaks: list[int]) -> list[int]:
    if not peaks:
        return []
    sorted_h = sorted(peaks)
    clusters: list[list[int]] = [[sorted_h[0]]]
    for h in sorted_h[1:]:
        if h - clusters[-1][-1] <= _PEAK_MERGE_HOURS:
            clusters[-1].append(h)
        else:
            clusters.append([h])
    return [int(round(sum(c) / len(c))) for c in clusters]


def _centroid_to_window(
    centroid: int,
    sheet_windows: dict[int, BakingWindow],
) -> BakingWindow | None:
    # Prefer a window whose end_hour <= centroid (batch ready before demand peak)
    candidates = [w for w in sheet_windows.values() if w.end_hour <= centroid]
    if candidates:
        return max(candidates, key=lambda w: w.end_hour)
    # All windows end after centroid — fall back to earliest window
    if sheet_windows:
        return min(sheet_windows.values(), key=lambda w: w.end_hour)
    return None


def _build_group_hourly(
    sku_hour_rows: list[dict[str, Any]],
    sku_meta: dict[str, SkuMeta],
) -> dict[str, dict[int, float]]:
    group_hourly: dict[str, dict[int, float]] = {}
    for row in sku_hour_rows:
        meta = _resolve_sku_meta(row.get("product_name"), sku_meta)
        if meta is None:
            continue
        hour = row.get("hour")
        if hour is None:
            continue
        qty = float(row.get("forecast_qty") or 0.0)
        bucket = group_hourly.setdefault(meta.dough_group, {})
        bucket[int(hour)] = bucket.get(int(hour), 0.0) + qty
    return group_hourly


def _compute_group_windows(
    group_hourly: dict[str, dict[int, float]],
    sheet_windows: dict[int, BakingWindow],
) -> dict[str, list[BakingWindow]]:
    result: dict[str, list[BakingWindow]] = {}
    for group, hourly in group_hourly.items():
        if not hourly or sum(hourly.values()) == 0:
            continue
        peaks = _detect_peaks(hourly)
        centroids = _cluster_peaks(peaks)
        windows: list[BakingWindow] = []
        seen_cols: set[int] = set()
        for centroid in centroids:
            w = _centroid_to_window(centroid, sheet_windows)
            if w and w.column not in seen_cols:
                windows.append(w)
                seen_cols.add(w.column)
        if windows:
            result[group] = sorted(windows, key=lambda w: w.column)
    return result


def _build_sku_schedule(
    sku_name: object,
    sku_meta: dict[str, SkuMeta],
    group_windows: dict[str, list[BakingWindow]],
    sheet_windows: dict[int, BakingWindow],
    has_next_day: bool = False,
) -> list[ScheduledColumn] | None:
    """Build a data-driven schedule for a SKU from group windows.

    Returns None when the SKU is not in the template metadata — caller
    should fall back to reading the pre-filled template schedule.

    For two-day SKUs the baker preps the defrost batch at the END of the
    current shift (last available window).  The quantity is sized from the
    next day's early-morning forecast by ``allocate_template_row``.  That
    window is therefore reserved for defrost and excluded from today's bake
    coverage so the second-to-last bake window absorbs the remaining hours.
    """
    meta = _resolve_sku_meta(sku_name, sku_meta)
    if meta is None:
        return None

    bake_windows = group_windows.get(meta.dough_group)
    if not bake_windows:
        return None

    schedule: list[ScheduledColumn] = []
    defrost_col: int | None = None

    if meta.is_two_day and has_next_day and sheet_windows:
        last_window = max(sheet_windows.values(), key=lambda w: w.column)
        schedule.append(
            ScheduledColumn(
                window=last_window,
                is_defrost=True,
                note="(ночная дефр)",
            )
        )
        defrost_col = last_window.column

    for window in bake_windows:
        if window.column == defrost_col:
            continue  # last window reserved for defrost
        schedule.append(
            ScheduledColumn(window=window, is_defrost=False, note=None)
        )

    return schedule or None


def allocate_template_row(
    *,
    template_sku_name: object,
    schedule: list[ScheduledColumn],
    product_hour_lookup: dict[str, dict[int, float]],
    next_day_hour_lookup: dict[str, dict[int, float]] | None = None,
    round_to: int | None = None,
    available_windows: dict[int, BakingWindow] | None = None,
    first_sales_hour: int = FIRST_SALES_HOUR,
    last_sales_hour: int = LAST_SALES_HOUR,
) -> dict[int, int | str]:
    """Allocate forecast quantities into a row's scheduled columns.

    Bake windows get the summed forecast over their coverage hours. Defrost columns
    get tomorrow's early-window volume (fallback: today's early-window volume) and
    keep their original annotation text. Defrost columns never contribute to today's
    coverage.
    """
    hourly = _resolve_hourly_forecast(template_sku_name, product_hour_lookup)
    if not hourly:
        return {}

    effective_round_to = round_to or schedule_round_to(schedule)
    bake_windows = [item.window for item in schedule if not item.is_defrost]
    bake_windows = _with_midday_split_window(
        bake_windows,
        available_windows=available_windows,
        hourly=hourly,
        first_sales_hour=first_sales_hour,
        last_sales_hour=last_sales_hour,
    )
    allocated: dict[int, int | str] = {}

    coverage = coverage_hours(
        bake_windows,
        first_sales_hour=first_sales_hour,
        last_sales_hour=last_sales_hour,
    )
    carry = 0.0
    for window in sorted(
        bake_windows,
        key=lambda item: (item.start_hour, item.end_hour, item.column),
    ):
        demand = sum(hourly.get(hour, 0.0) for hour in coverage.get(window.column, []))
        net_qty = max(demand - carry, 0.0)
        if net_qty <= 0:
            carry = max(carry - demand, 0.0)
            continue
        baked_qty = _round_up(net_qty, effective_round_to)
        allocated[window.column] = baked_qty
        carry = carry + baked_qty - demand

    defrost_columns = [item for item in schedule if item.is_defrost]
    if defrost_columns:
        next_hourly = None
        if next_day_hour_lookup is not None:
            next_hourly = _resolve_hourly_forecast(
                template_sku_name, next_day_hour_lookup
            )
        source = next_hourly if next_hourly else hourly
        defrost_qty = _round_up(
            _early_window_sum(source, first_sales_hour=first_sales_hour),
            effective_round_to,
        )
        for item in defrost_columns:
            # Strip the template's leading number, keep the annotation (e.g.
            # "(ночная дефр)") and re-emit it with the forecast-sized quantity.
            annotation = re.sub(r"^\s*\d+\s*", "", item.note or "").strip()
            allocated[item.window.column] = (
                f"{defrost_qty} {annotation}".strip() if annotation else defrost_qty
            )
    return allocated


def _select_sheet_name(bucket: str | None, sheet_names: list[str]) -> str:
    if bucket and bucket in sheet_names:
        return bucket
    if bucket:
        # Per-SKU schedules differ per sheet, so a wrong/unknown bucket silently
        # picks the wrong baking schedule. Surface it instead of failing quietly.
        logger.warning(
            "baking_plan: revenue bucket %r does not match any sheet %r; "
            "falling back to default",
            bucket,
            sheet_names,
        )
    if DEFAULT_BUCKET in sheet_names:
        return DEFAULT_BUCKET
    return sheet_names[0]


def template_path_for_bakery(bakery_id: object) -> Path:
    try:
        normalized_id = int(str(bakery_id))
    except (TypeError, ValueError):
        return DEFAULT_TEMPLATE_PATH
    template_path = INDIVIDUAL_TEMPLATE_PATHS.get(normalized_id)
    if template_path and template_path.exists():
        return template_path
    return DEFAULT_TEMPLATE_PATH


def _sales_hour_bounds(
    product_hour_lookup: dict[str, dict[int, float]],
) -> tuple[int, int]:
    hours = [hour for hourly in product_hour_lookup.values() for hour in hourly]
    if not hours:
        return FIRST_SALES_HOUR, LAST_SALES_HOUR
    return min(min(hours), FIRST_SALES_HOUR), max(max(hours), LAST_SALES_HOUR)


def _sheet_windows(sheet: Worksheet) -> dict[int, BakingWindow]:
    windows = {}
    for column in WINDOW_COLUMNS:
        parsed = parse_window_label(sheet.cell(row=5, column=column).value, column)
        if parsed:
            windows[column] = parsed
    return windows


def build_baking_plan_workbook(
    *,
    bakery: dict[str, Any],
    forecast_date: str,
    sku_hour_rows: list[dict[str, Any]],
    next_day_sku_hour_rows: list[dict[str, Any]] | None = None,
    assortment_rows: list[dict[str, Any]] | None = None,
    bucket: str | None = None,
    template_path: Path = DEFAULT_TEMPLATE_PATH,
) -> bytes:
    workbook = load_workbook(template_path)
    selected_sheet_name = _select_sheet_name(bucket, workbook.sheetnames)
    sheet = workbook[selected_sheet_name]

    # Parse metadata from комментарии BEFORE removing other sheets.
    sku_meta = parse_comments_sheet(workbook)

    for worksheet in list(workbook.worksheets):
        if worksheet.title != selected_sheet_name:
            workbook.remove(worksheet)
    sheet.title = "План выпекания"

    sheet["A1"] = (
        f"План выпекания: {bakery.get('bakery_name') or bakery.get('bakery_id')} "
        f"на {forecast_date}. Шаблон: {selected_sheet_name}"
    )
    sheet["A2"] = (
        "Количество по SKU-hour прогнозу активного run, разнесено по расписанию "
        "выпекания шаблона; ночная дефростация — по утреннему прогнозу след. дня."
    )

    sheet_windows = _sheet_windows(sheet)
    group_hourly = _build_group_hourly(sku_hour_rows, sku_meta)
    group_windows = _compute_group_windows(group_hourly, sheet_windows)
    product_hour_lookup = build_product_hour_lookup(sku_hour_rows)
    product_total_lookup = build_product_total_lookup(sku_hour_rows)
    assortment_lookup, assortment_products = build_assortment_lookup(
        assortment_rows or []
    )
    next_day_hour_lookup = (
        build_product_hour_lookup(next_day_sku_hour_rows)
        if next_day_sku_hour_rows
        else None
    )
    first_sales_hour, last_sales_hour = _sales_hour_bounds(product_hour_lookup)
    plan_rows: list[dict[str, Any]] = []
    present_product_ids: set[str] = set()
    prototype: dict[str, Any] | None = None

    for row_index in range(PLAN_START_ROW, sheet.max_row + 1):
        sheet.row_dimensions[row_index].hidden = False
        sku_name = sheet.cell(row=row_index, column=2).value
        if not sku_name:
            continue

        # Build schedule from data-driven group windows; fall back to template
        # pre-filled C:L cells for SKUs not in the комментарии sheet.
        schedule = _build_sku_schedule(
            sku_name,
            sku_meta,
            group_windows,
            sheet_windows,
            has_next_day=next_day_sku_hour_rows is not None,
        )
        if schedule is None:
            schedule = read_row_schedule(sheet, row_index, sheet_windows)
        if not schedule:
            continue

        assortment_product = (
            resolve_assortment_product(sku_name, assortment_lookup)
            if assortment_rows is not None
            else None
        )
        if assortment_rows is not None and assortment_product is None:
            continue

        sku_meta_entry = _resolve_sku_meta(sku_name, sku_meta)
        round_to = sku_meta_entry.kratnost if sku_meta_entry else None

        allocated = allocate_template_row(
            template_sku_name=sku_name,
            schedule=schedule,
            product_hour_lookup=product_hour_lookup,
            next_day_hour_lookup=next_day_hour_lookup,
            round_to=round_to,
            first_sales_hour=first_sales_hour,
            last_sales_hour=last_sales_hour,
            available_windows=sheet_windows,
        )
        snapshot = _snapshot_row(sheet, row_index)
        prototype = prototype or snapshot
        product_id = str((assortment_product or {}).get("product_id") or "")
        if product_id:
            present_product_ids.add(product_id)
        category_name = (
            (assortment_product or {}).get("category_name") or "Без группы"
        )
        plan_rows.append(
            {
                "snapshot": snapshot,
                "product_id": product_id,
                "product_name": (assortment_product or {}).get("product_name")
                or sku_name,
                "category_name": category_name,
                "allocated": allocated,
                "total": _numeric_plan_total(allocated),
                "source_order": row_index,
            }
        )

    for product in assortment_products:
        product_id = str(product["product_id"])
        if product_id in present_product_ids:
            continue
        plan_rows.append(
            {
                "snapshot": None,
                "product_id": product_id,
                "product_name": product["product_name"],
                "category_name": product["category_name"],
                "allocated": {},
                "total": _resolve_product_total(
                    product["product_name"], product_total_lookup
                ),
                "source_order": 10**9,
            }
        )

    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.max_row >= PLAN_START_ROW:
            sheet.unmerge_cells(str(merged_range))
    if sheet.max_row >= PLAN_START_ROW:
        sheet.delete_rows(PLAN_START_ROW, sheet.max_row - PLAN_START_ROW + 1)

    sheet.cell(row=5, column=TOTAL_COLUMN).value = "Итого"
    sheet.cell(row=5, column=TOTAL_COLUMN)._style = copy(
        sheet.cell(row=5, column=12)._style
    )
    if sheet.max_column > TOTAL_COLUMN:
        sheet.delete_cols(
            TOTAL_COLUMN + 1,
            sheet.max_column - TOTAL_COLUMN,
        )
    sheet.column_dimensions["M"].width = 14
    sheet.column_dimensions["M"].hidden = False

    category_fill = PatternFill("solid", fgColor="D9EAF7")
    current_row = PLAN_START_ROW
    plan_rows.sort(
        key=lambda row: (
            GROUP_SORT_ORDER.get(str(row["category_name"]), 100),
            str(row["category_name"]).casefold(),
            int(row["source_order"]),
            str(row["product_name"]).casefold(),
        )
    )
    current_category: str | None = None
    for plan_row in plan_rows:
        category = str(plan_row["category_name"] or "Без группы")
        if category != current_category:
            sheet.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=TOTAL_COLUMN,
            )
            category_cell = sheet.cell(row=current_row, column=1)
            category_cell.value = category
            category_cell.fill = category_fill
            category_cell.font = Font(bold=True)
            category_cell.alignment = Alignment(horizontal="left")
            sheet.row_dimensions[current_row].hidden = False
            current_category = category
            current_row += 1

        _restore_row(
            sheet,
            current_row,
            plan_row["snapshot"],
            prototype,
        )
        if plan_row["snapshot"] is None:
            for column in range(1, TOTAL_COLUMN + 1):
                sheet.cell(row=current_row, column=column).fill = PatternFill()
        sheet.cell(row=current_row, column=2).value = plan_row["product_name"]
        for column in WINDOW_COLUMNS:
            target_cell = sheet.cell(row=current_row, column=column)
            target_cell.value = plan_row["allocated"].get(column)
        total_cell = sheet.cell(row=current_row, column=TOTAL_COLUMN)
        total_cell.value = plan_row["total"]
        total_cell.number_format = "0"
        total_cell.font = copy(sheet.cell(row=current_row, column=12).font)
        total_cell.fill = copy(sheet.cell(row=current_row, column=12).fill)
        total_cell.border = copy(sheet.cell(row=current_row, column=12).border)
        total_cell.alignment = Alignment(horizontal="center")
        sheet.row_dimensions[current_row].hidden = False
        current_row += 1

    for row_index in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row_index].hidden = False
    for column in range(1, TOTAL_COLUMN + 1):
        column_letter = sheet.cell(row=1, column=column).column_letter
        sheet.column_dimensions[column_letter].hidden = False

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
