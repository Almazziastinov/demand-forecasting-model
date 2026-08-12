"""Parse and inventory archived pilot-plan workbooks.

The module is deliberately filesystem-agnostic: callers pass one file or an
already assembled collection of manifest records.  It never scans directories
and never chooses between conflicting publication versions implicitly.
"""

from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook


PlanFormat = Literal["legacy_v1", "stock_aware_v2"]
SourceQuality = Literal["bitrix_attachment", "local_unverified", "reconstructed"]

LEGACY_HEADERS = (
    "Пекарня",
    "Категория",
    "Номенклатура",
    "Прогноз",
    "Прогноз (кратность)",
    "Кратность",
)
STOCK_AWARE_HEADERS = (
    "Пекарня",
    "Категория",
    "Номенклатура",
    "Прогноз",
    "Остаток со вчерашнего дня",
    "Чистая потребность",
    "План выпуска",
    "Итого на продажу",
    "Кратность",
)

_DATE_PATTERN = re.compile(r"(?<!\d)(\d{2})\.(\d{2})\.(\d{4})(?!\d)")


def normalize_header(value: object) -> str:
    """Normalize a Russian workbook header for stable format matching."""

    if value is None:
        return ""
    text = str(value).replace("ё", "е").replace("Ё", "Е")
    return " ".join(text.strip().casefold().split())


_NORMALIZED_LEGACY_HEADERS = tuple(normalize_header(value) for value in LEGACY_HEADERS)
_NORMALIZED_STOCK_HEADERS = tuple(
    normalize_header(value) for value in STOCK_AWARE_HEADERS
)


@dataclass(frozen=True)
class PilotPlanRow:
    """One archived plan row at bakery and SKU grain."""

    bakery: str
    category: str
    product_name: str
    issued_forecast: float
    plan_qty: float
    kratnost: float
    yesterday_stock: float | None = None
    net_need: float | None = None
    total_for_sale: float | None = None


@dataclass(frozen=True)
class ParsedPilotPlan:
    """Parsed workbook content and its immutable file identity."""

    path: Path
    target_date: date
    format_version: PlanFormat
    rows: tuple[PilotPlanRow, ...]
    sha256: str


@dataclass(frozen=True)
class PilotPlanManifestRecord:
    """Provenance record for one candidate publication artifact."""

    target_date: date
    format_version: PlanFormat
    rows: int
    sha256: str
    source_quality: SourceQuality
    path: str
    published_at: datetime | None = None
    message_id: str | None = None
    disk_id: str | None = None
    is_effective: bool = False


def _number(value: object, *, field: str, row_number: int) -> float:
    if value is None or value == "":
        raise ValueError(f"Missing {field!r} at workbook row {row_number}")
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Invalid numeric {field!r} at workbook row {row_number}: {value!r}"
        ) from exc


def _optional_number(value: object, *, field: str, row_number: int) -> float | None:
    if value is None or value == "":
        return None
    return _number(value, field=field, row_number=row_number)


def _target_date(*values: object) -> date:
    for value in values:
        match = _DATE_PATTERN.search(str(value or ""))
        if match:
            day, month, year = (int(part) for part in match.groups())
            return date(year, month, day)
    raise ValueError("Target date was not found in workbook title or filename")


def _detect_format(headers: tuple[str, ...]) -> PlanFormat:
    if headers == _NORMALIZED_LEGACY_HEADERS:
        return "legacy_v1"
    if headers == _NORMALIZED_STOCK_HEADERS:
        return "stock_aware_v2"
    raise ValueError(f"Unsupported pilot-plan headers: {headers!r}")


def parse_pilot_plan(path: str | Path) -> ParsedPilotPlan:
    """Parse one pilot-plan workbook without modifying it."""

    resolved_path = Path(path).resolve()
    file_bytes = resolved_path.read_bytes()
    workbook = load_workbook(resolved_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    if worksheet.max_row is None:
        workbook.close()
        workbook = load_workbook(resolved_path, read_only=False, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
    title = worksheet.cell(row=1, column=1).value

    header_row = 0
    format_version: PlanFormat | None = None
    expected_width = 0
    for row_number, values in enumerate(
        worksheet.iter_rows(
            min_row=1, max_row=min(10, worksheet.max_row), values_only=True
        ),
        start=1,
    ):
        candidates = (
            tuple(normalize_header(value) for value in values[: len(LEGACY_HEADERS)]),
            tuple(
                normalize_header(value) for value in values[: len(STOCK_AWARE_HEADERS)]
            ),
        )
        for candidate in candidates:
            try:
                format_version = _detect_format(candidate)
            except ValueError:
                continue
            header_row = row_number
            expected_width = len(candidate)
            break
        if format_version:
            break
    if format_version is None:
        raise ValueError("Pilot-plan header row was not found in the first 10 rows")

    parsed_rows: list[PilotPlanRow] = []
    for row_number, values in enumerate(
        worksheet.iter_rows(
            min_row=header_row + 1,
            max_col=expected_width,
            values_only=True,
        ),
        start=header_row + 1,
    ):
        if not any(value not in (None, "") for value in values):
            continue
        bakery, category, product_name = (
            str(value or "").strip() for value in values[:3]
        )
        if not bakery or not category or not product_name:
            raise ValueError(f"Incomplete row identity at workbook row {row_number}")

        if format_version == "legacy_v1":
            parsed_rows.append(
                PilotPlanRow(
                    bakery=bakery,
                    category=category,
                    product_name=product_name,
                    issued_forecast=_number(
                        values[3], field="forecast", row_number=row_number
                    ),
                    plan_qty=_number(
                        values[4], field="rounded forecast", row_number=row_number
                    ),
                    kratnost=_number(
                        values[5], field="kratnost", row_number=row_number
                    ),
                )
            )
        else:
            parsed_rows.append(
                PilotPlanRow(
                    bakery=bakery,
                    category=category,
                    product_name=product_name,
                    issued_forecast=_number(
                        values[3], field="forecast", row_number=row_number
                    ),
                    yesterday_stock=_optional_number(
                        values[4], field="yesterday stock", row_number=row_number
                    ),
                    net_need=_optional_number(
                        values[5], field="net need", row_number=row_number
                    ),
                    plan_qty=_number(values[6], field="plan", row_number=row_number),
                    total_for_sale=_optional_number(
                        values[7], field="total for sale", row_number=row_number
                    ),
                    kratnost=_number(
                        values[8], field="kratnost", row_number=row_number
                    ),
                )
            )

    if not parsed_rows:
        raise ValueError("Pilot-plan workbook contains no data rows")

    return ParsedPilotPlan(
        path=resolved_path,
        target_date=_target_date(title, resolved_path.name),
        format_version=format_version,
        rows=tuple(parsed_rows),
        sha256=hashlib.sha256(file_bytes).hexdigest(),
    )


def build_manifest_record(
    plan: ParsedPilotPlan,
    *,
    source_quality: SourceQuality,
    published_at: datetime | None = None,
    message_id: str | None = None,
    disk_id: str | None = None,
    is_effective: bool = False,
) -> PilotPlanManifestRecord:
    """Build a provenance record for an already parsed plan."""

    if source_quality == "bitrix_attachment" and not message_id:
        raise ValueError("Bitrix attachments require message_id")
    return PilotPlanManifestRecord(
        target_date=plan.target_date,
        format_version=plan.format_version,
        rows=len(plan.rows),
        sha256=plan.sha256,
        source_quality=source_quality,
        path=str(plan.path),
        published_at=published_at,
        message_id=message_id,
        disk_id=disk_id,
        is_effective=is_effective,
    )


def validate_manifest(
    records: list[PilotPlanManifestRecord] | tuple[PilotPlanManifestRecord, ...],
) -> None:
    """Reject ambiguous effective versions and conflicting duplicate identities."""

    by_date: dict[date, list[PilotPlanManifestRecord]] = {}
    identities: dict[tuple[str | None, str | None], str] = {}
    for record in records:
        by_date.setdefault(record.target_date, []).append(record)
        identity = (record.message_id, record.disk_id)
        if any(identity):
            previous_hash = identities.setdefault(identity, record.sha256)
            if previous_hash != record.sha256:
                raise ValueError(
                    f"Publication identity {identity!r} has conflicting hashes"
                )

    for target_date, candidates in by_date.items():
        effective = [record for record in candidates if record.is_effective]
        if len(effective) > 1:
            raise ValueError(f"Multiple effective plans for {target_date.isoformat()}")
        distinct_hashes = {record.sha256 for record in candidates}
        if len(distinct_hashes) > 1 and not effective:
            raise ValueError(
                f"Conflicting unselected plan versions for {target_date.isoformat()}"
            )
