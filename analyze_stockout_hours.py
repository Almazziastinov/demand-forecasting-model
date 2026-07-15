"""
Hourly stockout detection prototype — bakery 21 (Парковая 7, Казань).

Algorithm:
1. Day-level filter: продано/выпуск >= STOCKOUT_RATIO → stockout day candidate
2. Coverage window assignment: template windows -> coverage hours per window
3. Within-window dropout detection: last sale hour < window coverage end,
   while bakery was still active -> stockout hours = last_sale_h+1 .. coverage_end
4. Missed demand estimate: avg sales rate in hours before dropout * n_stockout_hours

Run: .venv/Scripts/python.exe analyze_stockout_hours.py
"""

import sys
from pathlib import Path

import pandas as pd
import clickhouse_connect
from dotenv import dotenv_values
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "apps"))

from baking_plan.templates import parse_windows, COMMENTS_SHEET_NAME
from baking_plan.allocation import coverage_hours

# ── config ───────────────────────────────────────────────────────────────────
BAKERY_ID    = "000000021"
DATE_FROM    = "2026-06-15"
DATE_TO      = "2026-07-14"
STOCKOUT_RATIO = 0.90       # продано/выпуск >= this -> stockout day
MIN_BAKERY_ACTIVE = 3.0     # bakery-total units/hour to count as "open"
TEMPLATE_PATH = ROOT / "apps/baking_plan/assets/individual/21_parkovaya_7.xlsx"
BAKEABLE = {"Пироги сытные", "Пироги сладкие", "Выпечка сытная", "Выпечка сладкая", "Фастфуд"}

# ── ClickHouse ───────────────────────────────────────────────────────────────
env = dotenv_values(ROOT / ".env.dev")
client = clickhouse_connect.get_client(
    host=env["CLICKHOUSE_HOST"],
    port=int(env["CLICKHOUSE_PORT"]),
    username=env["CLICKHOUSE_USER"],
    password=env["CLICKHOUSE_PASSWORD"],
    database=env["CLICKHOUSE_DATABASE"],
    secure=True,
)

# ─────────────────────────────────────────────────────────────────────────────
# 1. Parse template windows
# ─────────────────────────────────────────────────────────────────────────────
print("Loading template windows...")
wb = load_workbook(TEMPLATE_PATH, read_only=True, data_only=True)
data_sheets = [s for s in wb.sheetnames if s != COMMENTS_SHEET_NAME]
sheet_name = data_sheets[0]
windows = parse_windows(wb, sheet_name)
wb.close()

# Deduplicate by label (template may have same label on multiple columns;
# for coverage analysis we only need unique time spans).
seen_labels: set[str] = set()
unique_windows = []
for w in windows:
    if w.label not in seen_labels:
        unique_windows.append(w)
        seen_labels.add(w.label)

coverage = coverage_hours(unique_windows)
# Drop windows with empty coverage (tight-packed templates where two windows
# share an end_hour produce zero-length coverage ranges).
coverage = {lbl: hrs for lbl, hrs in coverage.items() if hrs}

print(f"  Sheet: {sheet_name!r}  |  {len(unique_windows)} unique windows  ({len(coverage)} with coverage)")
for w in sorted(unique_windows, key=lambda x: x.start_hour):
    hrs = coverage.get(w.label, [])
    span = f"{min(hrs)}-{max(hrs)}h" if hrs else "empty (skipped)"
    print(f"    {w.label}: bake {w.start_hour}-{w.end_hour}h  covers sales {span}")

# ─────────────────────────────────────────────────────────────────────────────
# 2. Daily production (выпуск)
# ─────────────────────────────────────────────────────────────────────────────
print("\nLoading daily production...")
prod_df = client.query_df(
    """
    SELECT release_date AS date,
           product_id,
           sum(quantity) AS vypusk
    FROM fct_production_release
    WHERE bakery_id = %(bid)s
      AND release_date BETWEEN %(d0)s AND %(d1)s
      AND is_deleted != '1'
    GROUP BY date, product_id
    """,
    parameters={"bid": BAKERY_ID, "d0": DATE_FROM, "d1": DATE_TO},
)
prod_df["date"]   = pd.to_datetime(prod_df["date"])
prod_df["vypusk"] = prod_df["vypusk"].astype(float)
print(f"  {len(prod_df)} SKU-day rows, {prod_df['date'].nunique()} days")

# ─────────────────────────────────────────────────────────────────────────────
# 3. Hourly sales
# ─────────────────────────────────────────────────────────────────────────────
print("Loading hourly sales...")
sales_df = client.query_df(
    """
    SELECT toDate(check_datetime)   AS date,
           toHour(check_datetime)   AS hour,
           product_id,
           any(product_name)        AS product_name,
           any(category_name)       AS category_name,
           sum(quantity)            AS sold
    FROM mart_sales_60d
    WHERE bakery_id = %(bid)s
      AND check_date BETWEEN %(d0)s AND %(d1)s
      AND cash_event_type = 'Продажа'
    GROUP BY date, hour, product_id
    """,
    parameters={"bid": BAKERY_ID, "d0": DATE_FROM, "d1": DATE_TO},
)
sales_df["date"] = pd.to_datetime(sales_df["date"])
print(f"  {len(sales_df)} (date, hour, product) rows")

# ─────────────────────────────────────────────────────────────────────────────
# 4. Bakery-level hourly activity (is the bakery open that hour?)
# ─────────────────────────────────────────────────────────────────────────────
bakery_hour_total = (
    sales_df.groupby(["date", "hour"])["sold"]
    .sum()
    .rename("bakery_total")
    .reset_index()
)
bakery_lookup = (
    bakery_hour_total.set_index(["date", "hour"])["bakery_total"].to_dict()
)

# ─────────────────────────────────────────────────────────────────────────────
# 5. Day-level stockout filter: продано / выпуск >= STOCKOUT_RATIO
# ─────────────────────────────────────────────────────────────────────────────
bakeable_sales = sales_df[sales_df["category_name"].isin(BAKEABLE)].copy()
daily_sold = (
    bakeable_sales.groupby(["date", "product_id"])["sold"]
    .sum()
    .reset_index()
    .rename(columns={"sold": "daily_sold"})
)
merged = daily_sold.merge(prod_df, on=["date", "product_id"], how="inner")
merged["ratio"] = merged["daily_sold"] / merged["vypusk"].replace(0.0, float("nan"))
stockout_days = merged[merged["ratio"] >= STOCKOUT_RATIO].copy()

print(f"\nStockout day candidates: {len(stockout_days)} SKU-days "
      f"({stockout_days['product_id'].nunique()} SKUs, "
      f"{stockout_days['date'].nunique()} days)")

# ─────────────────────────────────────────────────────────────────────────────
# 6. Coverage-window hourly dropout detection
# ─────────────────────────────────────────────────────────────────────────────
hourly_lookup = (
    sales_df.set_index(["date", "product_id", "hour"])["sold"].to_dict()
)
name_by_pid = (
    sales_df.drop_duplicates("product_id")
    .set_index("product_id")[["product_name", "category_name"]]
    .to_dict("index")
)

events = []

for _, row in stockout_days.iterrows():
    date       = row["date"]
    pid        = row["product_id"]
    daily_sold = row["daily_sold"]
    vypusk     = float(row["vypusk"])
    ratio      = row["ratio"]

    for win_label, hours in coverage.items():
        if not hours:
            continue

        # Sales per hour in this coverage window
        win_sales = {h: hourly_lookup.get((date, pid, h), 0.0) for h in hours}
        sold_hours = [h for h, s in win_sales.items() if s > 0]

        if not sold_hours:
            continue  # conservative: no sales in window, skip

        last_sale_h = max(sold_hours)
        if last_sale_h >= max(hours):
            continue  # sold through whole window, no stockout here

        # Candidate stockout hours: after last sale until coverage end,
        # but only where bakery was still selling other things
        candidate = [h for h in hours if h > last_sale_h]
        active = [h for h in candidate if bakery_lookup.get((date, h), 0.0) >= MIN_BAKERY_ACTIVE]
        if not active:
            continue

        # Estimate missed demand from avg rate of selling hours in same window
        selling_qty = [win_sales[h] for h in hours if h <= last_sale_h and win_sales[h] > 0]
        avg_rate = sum(selling_qty) / len(selling_qty) if selling_qty else 0.0
        missed = avg_rate * len(active)

        pinfo = name_by_pid.get(pid, {})
        events.append({
            "date":           date,
            "product_id":     pid,
            "product_name":   pinfo.get("product_name", pid),
            "category":       pinfo.get("category_name", ""),
            "window":         win_label,
            "last_sale_h":    last_sale_h,
            "coverage_end_h": max(hours),
            "stockout_hours": active,
            "n_stockout_h":   len(active),
            "win_sold":       round(sum(win_sales.values()), 1),
            "avg_rate":       round(avg_rate, 2),
            "missed_est":     round(missed, 1),
            "daily_sold":     round(daily_sold, 1),
            "vypusk":         round(vypusk, 1),
            "ratio":          round(ratio, 3),
        })

if not events:
    print("\nNo within-window stockout events detected.")
    sys.exit(0)

ev_df = pd.DataFrame(events)

# ─────────────────────────────────────────────────────────────────────────────
# 7. Report
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 80)
print(f"HOURLY STOCKOUT REPORT  bakery {BAKERY_ID}  {DATE_FROM} -> {DATE_TO}")
print("=" * 80)

# Top SKUs by missed demand
top_skus = (
    ev_df.groupby(["product_id", "product_name", "category"])
    .agg(
        events       = ("date", "count"),
        stockout_days= ("date", "nunique"),
        total_missed = ("missed_est", "sum"),
        avg_missed_ph= ("avg_rate", "mean"),
        avg_n_hours  = ("n_stockout_h", "mean"),
    )
    .reset_index()
    .sort_values("total_missed", ascending=False)
    .head(15)
)
print("\nTop 15 SKUs by estimated missed units:")
print(top_skus.to_string(index=False))

# Which windows are most problematic
win_stats = (
    ev_df.groupby("window")
    .agg(events=("date", "count"), total_missed=("missed_est", "sum"), skus=("product_id", "nunique"))
    .reset_index()
    .sort_values("total_missed", ascending=False)
)
print("\nMissed demand by window:")
print(win_stats.to_string(index=False))

# Worst individual events
print("\nTop 20 worst (date, SKU, window) events:")
detail = (
    ev_df.sort_values("missed_est", ascending=False)
    .head(20)
    [["date", "product_name", "window", "last_sale_h", "coverage_end_h",
      "n_stockout_h", "win_sold", "missed_est", "ratio"]]
    .copy()
)
detail["date"] = detail["date"].dt.strftime("%Y-%m-%d")
print(detail.to_string(index=False))

# Aggregate: how many SKU-hours per day are stockout
daily_agg = (
    ev_df.groupby("date")
    .agg(sku_window_events=("product_id", "count"),
         total_missed=("missed_est", "sum"),
         skus_affected=("product_id", "nunique"))
    .reset_index()
    .sort_values("total_missed", ascending=False)
    .head(10)
)
daily_agg["date"] = daily_agg["date"].dt.strftime("%Y-%m-%d")
print("\nTop 10 worst days:")
print(daily_agg.to_string(index=False))

print(f"\nTotal estimated missed units (all events): {ev_df['missed_est'].sum():.0f}")
print(f"Total events detected: {len(ev_df)}")

out_path = ROOT / "outputs" / "stockout_hours_bakery21.csv"
out_path.parent.mkdir(exist_ok=True)
ev_df.to_csv(out_path, index=False, encoding="utf-8-sig")
print(f"\nFull results -> {out_path}")
