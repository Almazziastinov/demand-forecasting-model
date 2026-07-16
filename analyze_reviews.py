"""
Анализ отзывов: классификация тональности + семантические категории негатива.
Использует VibeCode AI API (bitrix/bitrixgpt-5.5).
Выходной файл: reviews_analysis_2026.xlsx
"""

import json
import os
import re
import time
import requests
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

API_KEY = "vibe_api_o9eDjakzMxux55JyoJS6GINHhfqYeuc5_dc336e"
BASE_URL = "https://vibecode.bitrix24.tech/v1"
MODEL = "bitrix/bitrixgpt-5.5"
INPUT_FILE = r"C:\Users\dns\Desktop\reviews_2026.xlsx"
CHECKPOINT_FILE = r"C:\Users\dns\Desktop\reviews_classified.json"
NEG_CHECKPOINT_FILE = r"C:\Users\dns\Desktop\reviews_neg_classified.json"
OUTPUT_FILE = r"C:\Users\dns\Desktop\reviews_analysis_2026.xlsx"
BATCH_SIZE = 25

HEADERS = {"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"}


# ─── AI helpers ───────────────────────────────────────────────────────────────

def ai_complete(prompt: str, max_tokens: int = 2000) -> str:
    resp = requests.post(
        f"{BASE_URL}/chat/completions",
        headers=HEADERS,
        json={"model": MODEL, "messages": [{"role": "user", "content": prompt}],
              "max_tokens": max_tokens},
        timeout=45,
    )
    if resp.status_code == 429:
        retry = int(resp.headers.get("Retry-After", 5))
        print(f"  429 rate-limit, ждём {retry}s")
        time.sleep(retry)
        return ai_complete(prompt, max_tokens)
    resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"].strip()


def batch_classify_sentiment(reviews: list[dict]) -> list[str]:
    """Классифицирует список отзывов за один запрос. Возвращает list тональностей."""
    lines = "\n".join(
        f"{i+1}. {r['text'][:300]}" for i, r in enumerate(reviews)
    )
    prompt = f"""Ты классификатор отзывов на русском языке о пекарне.
Для каждого отзыва ответь ТОЛЬКО одним словом: позитивный / нейтральный / негативный.
Никаких пояснений — только пронумерованный список ответов.

Отзывы:
{lines}

Ответ (формат: 1. позитивный / нейтральный / негативный, каждый на новой строке):"""

    raw = ai_complete(prompt, max_tokens=len(reviews) * 15)
    results = []
    for line in raw.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        # Извлекаем слово после номера
        m = re.search(r'\d+\.\s*(.+)', line)
        word = (m.group(1) if m else line).strip().lower()
        if "негат" in word or "плох" in word or "отриц" in word:
            results.append("негативный")
        elif "нейтр" in word:
            results.append("нейтральный")
        else:
            results.append("позитивный")

    # Дополняем если модель вернула меньше ответов
    while len(results) < len(reviews):
        results.append("позитивный")

    return results[:len(reviews)]


def identify_negative_categories(sample_texts: list[str]) -> list[str]:
    """По выборке негативных отзывов выделяет ключевые категории проблем."""
    # Берём не более 25 коротких отрывков чтобы не превысить лимит
    sample = "\n".join(f"- {t[:120]}" for t in sample_texts[:25])
    prompt = f"""Пекарня. Негативные отзывы ниже. Выдели 6-8 главных категорий проблем (2-4 слова каждая).
Только пронумерованный список, без пояснений.

{sample}

Ответ:"""
    raw = ai_complete(prompt, max_tokens=300)
    cats = []
    for line in raw.strip().splitlines():
        m = re.search(r'\d+\.\s*(.+)', line.strip())
        if m:
            cat = m.group(1).strip().rstrip(".")
            if cat:
                cats.append(cat)
    if not cats:
        cats = ["Качество выпечки", "Грубость персонала", "Завышенные цены",
                "Маленький ассортимент", "Долгое обслуживание", "Другое"]
    return cats


def batch_classify_negative(reviews: list[dict], categories: list[str]) -> list[str]:
    """Классифицирует негативные отзывы по категориям."""
    cat_list = "\n".join(f"- {c}" for c in categories)
    lines = "\n".join(f"{i+1}. {r['text'][:250]}" for i, r in enumerate(reviews))
    prompt = f"""Ты классификатор жалоб. Для каждого отзыва выбери ОДНУ наиболее подходящую категорию из списка.
Отвечай ТОЛЬКО пронумерованным списком — название категории точно как в списке.

Категории:
{cat_list}

Отзывы:
{lines}

Ответ (1. <категория>, каждый на новой строке):"""

    raw = ai_complete(prompt, max_tokens=len(reviews) * 20)
    results = []
    for line in raw.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        m = re.search(r'\d+\.\s*(.+)', line)
        val = (m.group(1) if m else line).strip()
        # Найти ближайшую категорию
        best = categories[0]
        best_score = 0
        for cat in categories:
            score = sum(w in val.lower() for w in cat.lower().split())
            if score > best_score:
                best_score = score
                best = cat
        results.append(best)

    while len(results) < len(reviews):
        results.append(categories[0])
    return results[:len(reviews)]


# ─── Classification pipeline ──────────────────────────────────────────────────

def extract_city(address: str) -> str:
    if not address:
        return "Неизвестно"
    return address.split(",")[0].strip()


def run_sentiment_classification(df: pd.DataFrame) -> pd.DataFrame:
    checkpoint = {}
    if os.path.exists(CHECKPOINT_FILE):
        with open(CHECKPOINT_FILE, encoding="utf-8") as f:
            checkpoint = json.load(f)
        print(f"Загружен чекпоинт: {len(checkpoint)} классифицировано")

    records = df[["id_msg", "text"]].to_dict("records")
    total = len(records)
    classified = dict(checkpoint)

    todo = [r for r in records if str(r["id_msg"]) not in classified]
    print(f"Осталось классифицировать: {len(todo)} из {total}")

    for i in range(0, len(todo), BATCH_SIZE):
        batch = todo[i: i + BATCH_SIZE]
        try:
            sentiments = batch_classify_sentiment(batch)
            for r, s in zip(batch, sentiments):
                classified[str(r["id_msg"])] = s
            pct = (len(classified) / total) * 100
            print(f"  [{len(classified)}/{total}] {pct:.1f}%  {sentiments[:3]}...")
        except Exception as e:
            print(f"  Ошибка в батче {i}: {e}")
            time.sleep(3)
            continue

        # Сохраняем чекпоинт каждые 10 батчей
        if (i // BATCH_SIZE) % 10 == 0:
            with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f:
                json.dump(classified, f, ensure_ascii=False)

        time.sleep(0.2)

    with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f:
        json.dump(classified, f, ensure_ascii=False)

    df["тональность"] = df["id_msg"].apply(lambda x: classified.get(str(x), "позитивный"))
    return df


def run_negative_classification(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    neg_df = df[df["тональность"] == "негативный"].copy()
    print(f"\nНегативных отзывов: {len(neg_df)}")

    if len(neg_df) == 0:
        df["категория_проблемы"] = ""
        return df, []

    # Загружаем чекпоинт если есть
    neg_checkpoint = {}
    if os.path.exists(NEG_CHECKPOINT_FILE):
        with open(NEG_CHECKPOINT_FILE, encoding="utf-8") as f:
            data = json.load(f)
            neg_checkpoint = data.get("cat_map", {})
            saved_cats = data.get("categories", [])
        if neg_checkpoint and saved_cats:
            print(f"Загружен чекпоинт негатива: {len(neg_checkpoint)} размечено")
            categories = saved_cats
            print(f"Категории: {categories}")
            # Проверяем полноту
            neg_ids = set(str(r) for r in neg_df["id_msg"])
            missing = neg_ids - set(neg_checkpoint.keys())
            if not missing:
                print("Все негативные уже размечены, пропускаю.")
                df["категория_проблемы"] = df["id_msg"].apply(
                    lambda x: neg_checkpoint.get(str(x), "")
                )
                return df, categories
            print(f"Осталось разметить: {len(missing)}")
    else:
        saved_cats = []

    texts = neg_df["text"].tolist()
    if not saved_cats:
        print("Определяю категории проблем...")
        categories = identify_negative_categories(texts)
        print(f"Выделено категорий: {len(categories)}")
        for c in categories:
            print(f"  - {c}")
    else:
        categories = saved_cats

    print("Классифицирую негативные отзывы по категориям...")
    cat_map = dict(neg_checkpoint)
    neg_records = [r for r in neg_df[["id_msg", "text"]].to_dict("records")
                   if str(r["id_msg"]) not in cat_map]

    for i in range(0, len(neg_records), BATCH_SIZE):
        batch = neg_records[i: i + BATCH_SIZE]
        try:
            cats = batch_classify_negative(batch, categories)
            for r, c in zip(batch, cats):
                cat_map[str(r["id_msg"])] = c
            done = len(neg_checkpoint) + i + len(batch)
            print(f"  [{done}/{len(neg_df)}]")
        except Exception as e:
            print(f"  Ошибка: {e}")
            time.sleep(3)
            continue

        # Сохраняем чекпоинт каждые 5 батчей
        if (i // BATCH_SIZE) % 5 == 0:
            with open(NEG_CHECKPOINT_FILE, "w", encoding="utf-8") as f:
                json.dump({"categories": categories, "cat_map": cat_map}, f, ensure_ascii=False)
        time.sleep(0.15)

    with open(NEG_CHECKPOINT_FILE, "w", encoding="utf-8") as f:
        json.dump({"categories": categories, "cat_map": cat_map}, f, ensure_ascii=False)

    df["категория_проблемы"] = df["id_msg"].apply(
        lambda x: cat_map.get(str(x), "")
    )
    return df, categories


# ─── Excel report ─────────────────────────────────────────────────────────────

FONT_NAME = "Arial"

C_DARK    = "1F3A5F"
C_POS     = "27AE60"
C_NEU     = "F39C12"
C_NEG     = "E74C3C"
C_HEAD    = "2C3E50"
C_STRIPE1 = "EBF5FB"
C_STRIPE2 = "FFFFFF"

def hdr_style(cell, bg=C_DARK, fg="FFFFFF", size=11):
    cell.font = Font(name=FONT_NAME, bold=True, color=fg, size=size)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def data_style(cell, stripe=False, bold=False, align="left"):
    cell.font = Font(name=FONT_NAME, size=10, bold=bold)
    if stripe:
        cell.fill = PatternFill("solid", start_color=C_STRIPE1)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)

def pct_style(cell, stripe=False):
    cell.number_format = "0.0%"
    data_style(cell, stripe, align="center")

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_border(cell):
    cell.border = border

def write_sheet_raw(wb, df):
    ws = wb.create_sheet("Все отзывы")
    cols = ["id_msg", "дата_сообщения", "источник", "дата_отзыва",
            "адрес", "город", "тональность", "категория_проблемы", "текст_отзыва"]
    headers = ["ID", "Дата сообщения", "Источник", "Дата отзыва",
               "Адрес", "Город", "Тональность", "Категория проблемы", "Текст отзыва"]
    widths  = [10, 16, 12, 12, 38, 18, 14, 22, 80]

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h)
        hdr_style(cell)
        apply_border(cell)
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.row_dimensions[1].height = 30

    SENT_COLOR = {"позитивный": "D5F5E3", "нейтральный": "FEF9E7", "негативный": "FADBD8"}

    for row_i, (_, row) in enumerate(df.iterrows(), 2):
        stripe = (row_i % 2 == 0)
        sent = row.get("тональность", "позитивный")
        for c, col in enumerate(cols, 1):
            val = row.get(col, "")
            cell = ws.cell(row_i, c, val)
            cell.font = Font(name=FONT_NAME, size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            apply_border(cell)
            if col == "тональность":
                cell.fill = PatternFill("solid", start_color=SENT_COLOR.get(sent, "FFFFFF"))
                cell.alignment = Alignment(horizontal="center", vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_sheet_summary(wb, df):
    ws = wb.create_sheet("Общая сводка")

    total = len(df)
    pos = (df["тональность"] == "позитивный").sum()
    neu = (df["тональность"] == "нейтральный").sum()
    neg = (df["тональность"] == "негативный").sum()

    # Title
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "Общая сводка отзывов 2026"
    c.font = Font(name=FONT_NAME, bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    headers = ["Тональность", "Количество", "Доля", ""]
    for c_i, h in enumerate(headers, 1):
        cell = ws.cell(3, c_i, h)
        hdr_style(cell)
        apply_border(cell)

    rows_data = [
        ("Позитивные", pos, pos/total if total else 0, C_POS),
        ("Нейтральные", neu, neu/total if total else 0, C_NEU),
        ("Негативные",  neg, neg/total if total else 0, C_NEG),
        ("ИТОГО",      total, 1.0, C_DARK),
    ]
    for r_i, (label, cnt, pct, color) in enumerate(rows_data, 4):
        bg = color if label == "ИТОГО" else color + "33"
        fg = "FFFFFF" if label == "ИТОГО" else "000000"
        bold = label == "ИТОГО"
        for ci, val in [(1, label), (2, cnt), (3, pct)]:
            cell = ws.cell(r_i, ci, val)
            cell.font = Font(name=FONT_NAME, bold=bold, size=10, color=fg)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
            if ci == 3:
                cell.number_format = "0.0%"
            apply_border(cell)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12

    # Pie chart
    pie = PieChart()
    pie.title = "Распределение тональности"
    pie.style = 10
    pie.width = 15
    pie.height = 10
    labels = Reference(ws, min_col=1, min_row=4, max_row=6)
    data   = Reference(ws, min_col=2, min_row=3, max_row=6)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    colors_hex = ["00B050", "FFC000", "FF0000"]
    for idx, color in enumerate(colors_hex):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        pie.series[0].dPt.append(pt)
    ws.add_chart(pie, "E3")


def write_sheet_neg_categories(wb, df, categories):
    ws = wb.create_sheet("Категории негатива")

    neg_df = df[df["тональность"] == "негативный"]
    if len(neg_df) == 0 or not categories:
        ws["A1"] = "Нет негативных отзывов"
        return

    cat_counts = neg_df["категория_проблемы"].value_counts()

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = f"Категории проблем в негативных отзывах (всего: {len(neg_df)})"
    c.font = Font(name=FONT_NAME, bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=C_NEG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for ci, h in enumerate(["Категория проблемы", "Количество", "Доля от негатива"], 1):
        cell = ws.cell(3, ci, h)
        hdr_style(cell, bg=C_NEG)
        apply_border(cell)

    total_neg = len(neg_df)
    shades = ["FADBD8", "F5B7B1", "EC7063", "E74C3C", "CB4335", "B03A2E",
              "943126", "78281F", "641E16", "4A0F10"]

    for r_i, (cat, cnt) in enumerate(cat_counts.items(), 4):
        stripe = (r_i % 2 == 0)
        shade = shades[min((r_i - 4), len(shades) - 1)]
        for ci, val in enumerate([cat, cnt, cnt / total_neg], 1):
            cell = ws.cell(r_i, ci, val)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.fill = PatternFill("solid", start_color=C_STRIPE1 if stripe else C_STRIPE2)
            apply_border(cell)
            if ci == 3:
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="center")
            elif ci == 2:
                cell.alignment = Alignment(horizontal="center")

    # Total row
    r_tot = 4 + len(cat_counts)
    ws.cell(r_tot, 1, "ИТОГО").font = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
    ws.cell(r_tot, 2, total_neg).font = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
    ws.cell(r_tot, 3, 1.0).font = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
    ws.cell(r_tot, 3).number_format = "0.0%"
    for ci in range(1, 4):
        ws.cell(r_tot, ci).fill = PatternFill("solid", start_color=C_NEG)
        ws.cell(r_tot, ci).alignment = Alignment(horizontal="center")
        apply_border(ws.cell(r_tot, ci))

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20

    # Bar chart
    bar = BarChart()
    bar.type = "bar"
    bar.title = "Категории проблем"
    bar.y_axis.title = "Количество"
    bar.style = 10
    bar.width = 20
    bar.height = 12
    n = len(cat_counts)
    data_ref  = Reference(ws, min_col=2, min_row=3, max_row=3 + n)
    cats_ref  = Reference(ws, min_col=1, min_row=4, max_row=3 + n)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.series[0].graphicalProperties.solidFill = "E74C3C"
    ws.add_chart(bar, "E3")


def write_pivot_sheet(wb, df, groupby_col: str, sheet_name: str, categories: list[str]):
    ws = wb.create_sheet(sheet_name)

    # Columns: группа, всего, позитивных, нейтральных, негативных, %позитив, %негатив
    # + категории проблем (для негативных)
    sent_cols = ["позитивный", "нейтральный", "негативный"]
    base_headers = [groupby_col, "Всего", "Позитивных", "Нейтральных", "Негативных",
                    "% позитив", "% нейтраль", "% негатив"]
    all_headers = base_headers + categories

    widths = [32 if groupby_col == "адрес" else 20, 10, 12, 12, 12, 11, 11, 11]
    widths += [18] * len(categories)

    for ci, (h, w) in enumerate(zip(all_headers, widths), 1):
        cell = ws.cell(1, ci, h)
        hdr_style(cell)
        apply_border(cell)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 36

    grouped = df.groupby(groupby_col)
    pivot = grouped["тональность"].value_counts().unstack(fill_value=0)
    for s in sent_cols:
        if s not in pivot.columns:
            pivot[s] = 0
    pivot["всего"] = pivot[sent_cols].sum(axis=1)
    pivot = pivot.sort_values("всего", ascending=False)

    # Negative category breakdown per group
    neg_df = df[df["тональность"] == "негативный"]
    cat_pivot = pd.DataFrame(index=pivot.index, columns=categories).fillna(0)
    if len(neg_df) > 0 and categories:
        neg_grp = neg_df.groupby([groupby_col, "категория_проблемы"]).size().unstack(fill_value=0)
        for cat in categories:
            if cat in neg_grp.columns:
                for idx in pivot.index:
                    if idx in neg_grp.index:
                        cat_pivot.loc[idx, cat] = neg_grp.loc[idx, cat]

    for r_i, (group_val, row) in enumerate(pivot.iterrows(), 2):
        stripe = (r_i % 2 == 0)
        total = row["всего"]
        pos = row.get("позитивный", 0)
        neu = row.get("нейтральный", 0)
        neg = row.get("негативный", 0)

        vals = [group_val, total, pos, neu, neg,
                pos/total if total else 0,
                neu/total if total else 0,
                neg/total if total else 0]
        vals += [int(cat_pivot.loc[group_val, cat]) for cat in categories]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(r_i, ci, val)
            cell.font = Font(name=FONT_NAME, size=9)
            cell.fill = PatternFill("solid", start_color=C_STRIPE1 if stripe else C_STRIPE2)
            apply_border(cell)
            if ci in (6, 7, 8):
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="center")
                # Colour-code % негатив
                if ci == 8 and total > 0:
                    pct_neg = neg / total
                    if pct_neg > 0.3:
                        cell.fill = PatternFill("solid", start_color="FADBD8")
                    elif pct_neg > 0.15:
                        cell.fill = PatternFill("solid", start_color="FEF9E7")
            elif ci in (2, 3, 4, 5):
                cell.alignment = Alignment(horizontal="center")
            elif ci == 1:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center")

    # Total row
    r_tot = 2 + len(pivot)
    total_all = len(df)
    pos_all = (df["тональность"] == "позитивный").sum()
    neu_all = (df["тональность"] == "нейтральный").sum()
    neg_all = (df["тональность"] == "негативный").sum()
    tot_vals = ["ИТОГО", total_all, pos_all, neu_all, neg_all,
                pos_all/total_all if total_all else 0,
                neu_all/total_all if total_all else 0,
                neg_all/total_all if total_all else 0]
    neg_df2 = df[df["тональность"] == "негативный"]
    for cat in categories:
        tot_vals.append(int((neg_df2["категория_проблемы"] == cat).sum()))

    for ci, val in enumerate(tot_vals, 1):
        cell = ws.cell(r_tot, ci, val)
        cell.font = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=C_HEAD)
        apply_border(cell)
        if ci in (6, 7, 8):
            cell.number_format = "0.0%"
            cell.alignment = Alignment(horizontal="center")
        elif ci in (2, 3, 4, 5):
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.alignment = Alignment(horizontal="left")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}1"


def build_excel(df: pd.DataFrame, categories: list[str]):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("Пишу листы Excel...")
    write_sheet_summary(wb, df)
    write_sheet_neg_categories(wb, df, categories)
    write_pivot_sheet(wb, df, "адрес", "По пекарням", categories)
    write_pivot_sheet(wb, df, "город", "По городам", categories)
    write_sheet_raw(wb, df)

    wb.save(OUTPUT_FILE)
    print(f"Файл сохранён: {OUTPUT_FILE}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("Загружаю отзывы из Excel...")
    raw = pd.read_excel(INPUT_FILE)

    col_map = {
        raw.columns[0]: "id_msg",
        raw.columns[1]: "дата_сообщения",
        raw.columns[2]: "источник",
        raw.columns[3]: "дата_отзыва",
        raw.columns[4]: "адрес",
        raw.columns[5]: "текст_отзыва",
        raw.columns[6]: "текст_полный",
    }
    raw = raw.rename(columns=col_map)
    raw["text"] = raw["текст_полный"].fillna(raw["текст_отзыва"]).astype(str)
    raw["адрес"] = raw["адрес"].fillna("").astype(str)
    raw["город"] = raw["адрес"].apply(extract_city)

    print(f"Загружено {len(raw)} отзывов\n")

    # 1. Sentiment
    print("=== Шаг 1: классификация тональности ===")
    df = run_sentiment_classification(raw)
    sent_counts = df["тональность"].value_counts()
    print(f"Результат: {dict(sent_counts)}\n")

    # 2. Negative categories
    print("=== Шаг 2: категории негатива ===")
    df, categories = run_negative_classification(df)

    # 3. Build Excel
    print("\n=== Шаг 3: создание Excel-отчёта ===")
    build_excel(df, categories)
    print("\nГотово!")


if __name__ == "__main__":
    main()
