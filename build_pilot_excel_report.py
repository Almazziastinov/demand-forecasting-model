"""
Excel-отчёт по пилотным пекарням: лист на пекарню, строки = SKU,
колонки = Факт + 4 варианта прогноза (prod, prior14, bias_corr, base_raw)
для каждой даты 2026-06-22..2026-06-28. Первая строка -- итоги по пекарне.
Цвет ячеек прогноза = отклонение от факта (зелёный близко, красный сильно мимо).
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import clickhouse_connect
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from pipelines.forecast_publish.load_forecast_run import get_clickhouse_settings

PILOT_IDS = [20, 21, 22, 28, 80, 89, 107, 221, 222, 257]
START_DATE = "2026-06-22"
END_DATE = "2026-06-28"
DATES = pd.date_range(START_DATE, END_DATE).strftime("%Y-%m-%d").tolist()
VARIANTS = ["prod", "prior14", "bias_corr", "base_raw"]
VARIANT_LABELS = {"prod": "Прод", "prior14": "Prior14", "bias_corr": "BiasCorr", "base_raw": "BaseRaw"}

OUT_PATH = Path(__file__).resolve().parent / "outputs" / "pilot_bakeries_forecast_vs_fact.xlsx"


def connect():
    ch = get_clickhouse_settings(".env")
    return clickhouse_connect.get_client(
        host=ch["host"], port=int(ch["port"]),
        username=ch["username"], password=ch["password"],
        database=ch["database"], secure=bool(ch["secure"]), verify=False,
    )


def load_variant_forecast(client, variant: str) -> pd.DataFrame:
    if variant == "prod":
        where_extra = "and source_run_id not like 'dev_%%'"
        params = {"pilot": PILOT_IDS, "start": START_DATE, "end": END_DATE}
    else:
        where_extra = "and source_run_id like %(pattern)s"
        params = {"pilot": PILOT_IDS, "start": START_DATE, "end": END_DATE,
                  "pattern": f"dev_{variant}_%_h1"}
    df = client.query_df(
        f"""
        select
            forecast_date,
            toInt64(bakery_id)  as bakery_id,
            toInt64(product_id) as product_id,
            any(product_name)   as product_name,
            sum(forecast_qty)   as forecast_qty
        from sku_forecast_day_snapshots
        where lead_days = 1
          and toInt64(bakery_id) in %(pilot)s
          and forecast_date between %(start)s and %(end)s
          {where_extra}
        group by forecast_date, bakery_id, product_id
        """,
        parameters=params,
    )
    df["forecast_date"] = pd.to_datetime(df["forecast_date"]).dt.strftime("%Y-%m-%d")
    return df


def load_actuals(client) -> pd.DataFrame:
    df = client.query_df(
        """
        select
            check_date          as forecast_date,
            toInt64(bakery_id)  as bakery_id,
            toInt64(product_id) as product_id,
            any(product_name)   as product_name,
            sum(quantity)       as actual_qty
        from mart_sales_60d
        where toInt64(bakery_id) in %(pilot)s
          and check_date between %(start)s and %(end)s
        group by forecast_date, bakery_id, product_id
        """,
        parameters={"pilot": PILOT_IDS, "start": START_DATE, "end": END_DATE},
    )
    df["forecast_date"] = pd.to_datetime(df["forecast_date"]).dt.strftime("%Y-%m-%d")
    return df


def bias_color(actual: float, forecast: float) -> str:
    """Возвращает hex-цвет по степени отклонения прогноза от факта."""
    if actual <= 0 and forecast <= 0:
        return "FFFFFF"
    denom = max(actual, 1.0)
    rel = (forecast - actual) / denom
    arel = abs(rel)
    if arel <= 0.15:
        return "C6EFCE"   # зелёный - близко
    if arel <= 0.40:
        return "FFEB9C"   # жёлтый - умеренное отклонение
    if rel > 0:
        return "FFC7CE"   # красный - перепрогноз
    return "F8CBAD"       # оранжевый - недопрогноз


def main():
    out = sys.stdout.buffer

    def p(s: str):
        out.write((s + "\n").encode("utf-8"))
        out.flush()

    p("Подключаемся к ClickHouse...")
    client = connect()

    p("Загружаем факт...")
    actuals = load_actuals(client)
    p(f"  факт: {len(actuals):,} строк")

    forecasts: dict[str, pd.DataFrame] = {}
    for v in VARIANTS:
        df = load_variant_forecast(client, v)
        forecasts[v] = df
        p(f"  {v}: {len(df):,} строк")

    # Сводим всё в один df: ключ (bakery_id, product_id, forecast_date)
    key = ["bakery_id", "product_id", "forecast_date"]
    merged = actuals[key + ["product_name", "actual_qty"]].copy()
    for v in VARIANTS:
        df = forecasts[v]
        merged = merged.merge(
            df[key + ["forecast_qty"]].rename(columns={"forecast_qty": f"fc_{v}"}),
            on=key, how="outer",
        )
    # product_name может прийти только из факта -- подтянем из прогнозов где не было факта
    for v in VARIANTS:
        df = forecasts[v]
        name_map = df.set_index(key)["product_name"]
        missing = merged["product_name"].isna()
        if missing.any():
            merged.loc[missing, "product_name"] = merged.loc[missing].set_index(key).index.map(name_map)

    merged["actual_qty"] = merged["actual_qty"].fillna(0.0)
    for v in VARIANTS:
        merged[f"fc_{v}"] = merged[f"fc_{v}"].fillna(0.0)
    merged["product_name"] = merged["product_name"].fillna("(без названия)")

    p("Строим Excel...")
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill("solid", fgColor="D9E1F2")
    total_font = Font(bold=True)
    sku_font = Font(bold=False)

    # Цвет своей "шапки" под каждую под-колонку -- одинаковый на каждом листе и на каждой дате,
    # чтобы было видно, какая колонка за какой вариант отвечает.
    sub_header_colors = {
        "Факт": "808080",       # серый
        "Прод": "375623",       # тёмно-зелёный (текущий прод)
        "Prior14": "9DC3E6",    # голубой
        "BiasCorr": "B4A7D6",   # сиреневый
        "BaseRaw": "F4B084",    # оранжевый
    }
    thin = Side(style="thin", color="BFBFBF")
    thick = Side(style="medium", color="000000")

    # --- Лист-легенда ---
    legend = wb.create_sheet(title="Легенда")
    legend["A1"] = "Структура отчёта"
    legend["A1"].font = Font(bold=True, size=14)
    legend_rows = [
        ("", ""),
        ("Лист на каждую пилотную пекарню.", ""),
        ("Столбец A", "Название SKU. Строка 1 = заголовок 'SKU', строка 2 = ИТОГО по пекарне (сумма всех SKU), с строки 3 -- сами SKU."),
        ("", ""),
        ("Каждая дата (22-28 июня)", "блок из 5 колонок подряд, в таком порядке:"),
        ("  Факт", "фактические продажи за день (серый заголовок)"),
        ("  Прод", "текущий продовый прогноз (тёмно-зелёный заголовок)"),
        ("  Prior14", "dev-вариант: uplift-модель, окно приора 14 дней вместо 30 (голубой заголовок)"),
        ("  BiasCorr", "dev-вариант: uplift-модель + коррекция смещения для топ-5 SKU (сиреневый заголовок)"),
        ("  BaseRaw", "dev-вариант: базовая (norm) модель + raw uplift на аллокации (оранжевый заголовок)"),
        ("", ""),
        ("Цвет ячейки прогноза", "= степень отклонения от факта в этот же день:"),
        ("  зелёный", "отклонение <= 15% от факта"),
        ("  жёлтый", "отклонение 15-40%"),
        ("  красный", "перепрогноз > 40% (прогноз сильно выше факта)"),
        ("  оранжевый", "недопрогноз > 40% (прогноз сильно ниже факта)"),
        ("", ""),
        ("ВАЖНО", "у вариантов Prior14 и BiasCorr заметно меньше строк прогноза в источнике данных, "
                  "чем у Прод/BaseRaw -- много пустых/нулевых значений по ним может быть из-за "
                  "неполноты бэкфилла, а не реальной недопрогнозировки."),
    ]
    for i, (a, b) in enumerate(legend_rows, start=2):
        legend.cell(row=i, column=1, value=a).font = Font(bold=a.strip().isupper() or (a.startswith("  ") is False and a != ""))
        legend.cell(row=i, column=2, value=b)
        if a.strip() in sub_header_colors:
            legend.cell(row=i, column=1).fill = PatternFill("solid", fgColor=sub_header_colors[a.strip()])
            legend.cell(row=i, column=1).font = Font(bold=True, color="FFFFFF")
    legend.column_dimensions["A"].width = 14
    legend.column_dimensions["B"].width = 110

    for bakery_id in PILOT_IDS:
        sub = merged[merged["bakery_id"] == bakery_id]
        if sub.empty:
            continue
        sheet_name = f"Пекарня {bakery_id}"
        ws = wb.create_sheet(title=sheet_name[:31])

        # --- Заголовок: строка 1 = даты (merge по 5 колонок), строка 2 = Факт/варианты ---
        ws.cell(row=1, column=1, value="SKU").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=2, column=1, value=f"ИТОГО пекарня {bakery_id}").font = total_font
        ws.cell(row=2, column=1).fill = total_fill

        col = 2
        date_col_start: dict[str, int] = {}
        for d in DATES:
            date_col_start[d] = col
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(VARIANTS))
            c = ws.cell(row=1, column=col, value=d)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
            sub_headers = ["Факт"] + [VARIANT_LABELS[v] for v in VARIANTS]
            for i, h in enumerate(sub_headers):
                hc = ws.cell(row=2, column=col + i, value=h)
                hc.font = Font(bold=True, size=10, color="FFFFFF")
                hc.fill = PatternFill("solid", fgColor=sub_header_colors[h])
                hc.alignment = Alignment(horizontal="center")
            # толстая граница справа от блока даты, чтобы было видно где он заканчивается
            last_col_of_block = col + len(VARIANTS)
            for r in (1, 2):
                ws.cell(row=r, column=last_col_of_block).border = Border(right=thick)
            col += len(VARIANTS) + 1

        # --- Итоговая строка (row 2) по пекарне ---
        for d in DATES:
            day_sub = sub[sub["forecast_date"] == d]
            actual_sum = day_sub["actual_qty"].sum()
            base_col = date_col_start[d]
            ws.cell(row=2, column=base_col, value=round(actual_sum, 1)).font = total_font
            for i, v in enumerate(VARIANTS, start=1):
                fc_sum = day_sub[f"fc_{v}"].sum()
                cell = ws.cell(row=2, column=base_col + i, value=round(fc_sum, 1))
                cell.font = total_font
                cell.fill = PatternFill("solid", fgColor=bias_color(actual_sum, fc_sum))
            ws.cell(row=2, column=base_col + len(VARIANTS)).border = Border(right=thick)

        # --- Строки по SKU ---
        skus = (
            sub.groupby(["product_id", "product_name"])["actual_qty"].sum()
            .sort_values(ascending=False).reset_index()[["product_id", "product_name"]]
        )
        row = 3
        for _, sku_row in skus.iterrows():
            pid, pname = sku_row["product_id"], sku_row["product_name"]
            ws.cell(row=row, column=1, value=pname).font = sku_font
            sku_data = sub[sub["product_id"] == pid]
            for d in DATES:
                day_row = sku_data[sku_data["forecast_date"] == d]
                actual_v = float(day_row["actual_qty"].sum()) if not day_row.empty else 0.0
                base_col = date_col_start[d]
                ws.cell(row=row, column=base_col, value=round(actual_v, 1))
                for i, v in enumerate(VARIANTS, start=1):
                    fc_v = float(day_row[f"fc_{v}"].sum()) if not day_row.empty else 0.0
                    cell = ws.cell(row=row, column=base_col + i, value=round(fc_v, 1))
                    cell.fill = PatternFill("solid", fgColor=bias_color(actual_v, fc_v))
                ws.cell(row=row, column=base_col + len(VARIANTS)).border = Border(right=thick)
            row += 1

        ws.column_dimensions["A"].width = 38
        for d in DATES:
            base_col = date_col_start[d]
            for i in range(len(VARIANTS) + 1):
                ws.column_dimensions[get_column_letter(base_col + i)].width = 9
        ws.freeze_panes = "B3"

    OUT_PATH.parent.mkdir(exist_ok=True)
    wb.save(OUT_PATH)
    p(f"\nГотово: {OUT_PATH}")


if __name__ == "__main__":
    main()
